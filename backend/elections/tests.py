import csv
import json
from datetime import timedelta
from io import BytesIO
from io import StringIO
from unittest.mock import patch

from audit.models import AuditLog
from django.core.cache import cache
from django.contrib.auth import get_user_model
from django.test import override_settings
from django.utils import timezone
from openpyxl import load_workbook
from reports.weasyprint_utils import WeasyPrintUnavailableError
from rest_framework.test import APITestCase

from elections.models import (
    ElectionCandidate,
    ElectionCensusMember,
    ElectionProcess,
    ElectionRole,
    VoteAccessSession,
    VoteRecord,
    VoterToken,
)
from elections.services_observer import generate_observer_congratulations_for_election
from students.models import ObserverAnnotation, Student


class ElectionE2EFlowTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username="admin_elections_tests",
            password="pass1234",
            role=user_model.ROLE_ADMIN,
        )

    def _create_process_with_ballot(self, name: str = "Jornada E2E"):
        now = timezone.now()
        process = ElectionProcess.objects.create(
            name=name,
            status=ElectionProcess.Status.OPEN,
            starts_at=now - timedelta(hours=1),
            ends_at=now + timedelta(hours=2),
        )
        personero_role = ElectionRole.objects.create(
            process=process,
            code=ElectionRole.CODE_PERSONERO,
            title="Personería",
            display_order=1,
        )
        contralor_role = ElectionRole.objects.create(
            process=process,
            code=ElectionRole.CODE_CONTRALOR,
            title="Contraloría",
            display_order=2,
        )
        personero_candidate = ElectionCandidate.objects.create(
            role=personero_role,
            name="Ana Candidata",
            number="01",
            grade="11",
            is_active=True,
            display_order=1,
        )
        contralor_candidate = ElectionCandidate.objects.create(
            role=contralor_role,
            name="Luis Candidato",
            number="02",
            grade="10",
            is_active=True,
            display_order=1,
        )
        return process, personero_role, contralor_role, personero_candidate, contralor_candidate

    def _create_token(self, *, process: ElectionProcess, raw_token: str, expires_delta_hours: int = 2, **kwargs):
        now = timezone.now()
        defaults = {
            "token_hash": VoterToken.hash_token(raw_token),
            "token_prefix": raw_token[:12],
            "status": VoterToken.Status.ACTIVE,
            "expires_at": now + timedelta(hours=expires_delta_hours),
            "student_grade": "10",
            "student_shift": "Mañana",
            "metadata": {
                "student_external_id": "EXT-100",
                "document_number": "DOC-100",
            },
        }
        defaults.update(kwargs)
        return VoterToken.objects.create(process=process, **defaults)

    def _submit_vote_with_blank_for_second_role(
        self,
        *,
        process: ElectionProcess,
        personero_role: ElectionRole,
        contralor_role: ElectionRole,
        personero_candidate: ElectionCandidate,
        token_suffix: str,
    ):
        ElectionCensusMember.objects.create(
            student_external_id=f"EXT-{token_suffix}",
            document_number=f"DOC-{token_suffix}",
            full_name="Estudiante Export",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )
        raw_token = f"VOTO-E2E-{token_suffix}"
        self._create_token(
            process=process,
            raw_token=raw_token,
            metadata={
                "student_external_id": f"EXT-{token_suffix}",
                "document_number": f"DOC-{token_suffix}",
            },
        )

        self.client.force_authenticate(user=None)
        validate_response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )
        self.assertEqual(validate_response.status_code, 200)

        submit_response = self.client.post(
            "/api/elections/public/submit-vote/",
            {
                "access_session_id": validate_response.data["access_session_id"],
                "selections": [
                    {"role_id": personero_role.id, "candidate_id": personero_candidate.id, "is_blank": False},
                    {"role_id": contralor_role.id, "is_blank": True},
                ],
            },
            format="json",
        )
        self.assertEqual(submit_response.status_code, 201)

    def test_happy_path_validate_and_submit_vote_consumes_token(self):
        process, personero_role, contralor_role, personero_candidate, contralor_candidate = self._create_process_with_ballot()
        ElectionCensusMember.objects.create(
            student_external_id="EXT-100",
            document_number="DOC-100",
            full_name="Estudiante Prueba",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )
        raw_token = "VOTO-E2E-0001"
        voter_token = self._create_token(process=process, raw_token=raw_token)

        self.client.force_authenticate(user=None)
        validate_response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )

        self.assertEqual(validate_response.status_code, 200)
        access_session_id = validate_response.data["access_session_id"]
        self.assertEqual(validate_response.data["process"]["id"], process.id)
        self.assertEqual(len(validate_response.data["roles"]), 2)

        submit_response = self.client.post(
            "/api/elections/public/submit-vote/",
            {
                "access_session_id": access_session_id,
                "selections": [
                    {"role_id": personero_role.id, "candidate_id": personero_candidate.id, "is_blank": False},
                    {"role_id": contralor_role.id, "candidate_id": contralor_candidate.id, "is_blank": False},
                ],
            },
            format="json",
        )

        self.assertEqual(submit_response.status_code, 201)
        self.assertEqual(submit_response.data["process_id"], process.id)
        self.assertEqual(submit_response.data["saved_votes"], 2)

        voter_token.refresh_from_db()
        self.assertEqual(voter_token.status, VoterToken.Status.USED)
        self.assertIsNotNone(voter_token.used_at)
        self.assertEqual(VoteRecord.objects.filter(voter_token=voter_token).count(), 2)

    def test_submit_vote_is_idempotent_for_same_access_session(self):
        process, personero_role, contralor_role, personero_candidate, contralor_candidate = self._create_process_with_ballot(
            name="Jornada Idempotente"
        )
        ElectionCensusMember.objects.create(
            student_external_id="EXT-200",
            document_number="DOC-200",
            full_name="Estudiante Idempotencia",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )
        raw_token = "VOTO-E2E-IDEMP-1"
        voter_token = self._create_token(
            process=process,
            raw_token=raw_token,
            metadata={"student_external_id": "EXT-200", "document_number": "DOC-200"},
        )

        self.client.force_authenticate(user=None)
        validate_response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )
        self.assertEqual(validate_response.status_code, 200)

        payload = {
            "access_session_id": validate_response.data["access_session_id"],
            "selections": [
                {"role_id": personero_role.id, "candidate_id": personero_candidate.id, "is_blank": False},
                {"role_id": contralor_role.id, "candidate_id": contralor_candidate.id, "is_blank": False},
            ],
        }

        first_submit_response = self.client.post(
            "/api/elections/public/submit-vote/",
            payload,
            format="json",
        )
        self.assertEqual(first_submit_response.status_code, 201)
        self.assertFalse(first_submit_response.data["already_submitted"])
        self.assertEqual(first_submit_response.data["saved_votes"], 2)

        second_submit_response = self.client.post(
            "/api/elections/public/submit-vote/",
            payload,
            format="json",
        )
        self.assertEqual(second_submit_response.status_code, 200)
        self.assertTrue(second_submit_response.data["already_submitted"])
        self.assertEqual(second_submit_response.data["saved_votes"], 2)
        self.assertEqual(second_submit_response.data["process_id"], process.id)

        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_VOTE_SUBMIT",
                object_type="ElectionProcess",
                object_id=str(process.id),
                status_code=201,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_VOTE_SUBMIT_DUPLICATE",
                object_type="ElectionProcess",
                object_id=str(process.id),
                status_code=200,
            ).exists()
        )

        voter_token.refresh_from_db()
        self.assertEqual(voter_token.status, VoterToken.Status.USED)
        self.assertEqual(VoteRecord.objects.filter(voter_token=voter_token).count(), 2)

    def test_contingency_reset_reactivates_used_token(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Reset")
        raw_token = "VOTO-E2E-RESET"
        voter_token = self._create_token(
            process=process,
            raw_token=raw_token,
            status=VoterToken.Status.USED,
            used_at=timezone.now() - timedelta(minutes=10),
            expires_at=timezone.now() - timedelta(minutes=1),
        )

        self.client.force_authenticate(user=self.admin)
        reset_response = self.client.post(
            "/api/elections/tokens/reset/",
            {
                "token": raw_token,
                "reason": "Se reinicia por contingencia durante jornada.",
                "extend_hours": 6,
            },
            format="json",
        )

        self.assertEqual(reset_response.status_code, 200)
        self.assertEqual(reset_response.data["status"], VoterToken.Status.ACTIVE)
        self.assertEqual(reset_response.data["token_id"], voter_token.id)

        voter_token.refresh_from_db()
        self.assertEqual(voter_token.status, VoterToken.Status.ACTIVE)
        self.assertIsNone(voter_token.used_at)
        self.assertGreater(voter_token.expires_at, timezone.now())

    def test_scrutiny_exports_csv_xlsx_and_pdf_endpoints(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Exportes")
        self.client.force_authenticate(user=self.admin)

        csv_response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.csv")
        self.assertEqual(csv_response.status_code, 200)
        self.assertIn("text/csv", csv_response["Content-Type"])

        xlsx_response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.xlsx")
        self.assertEqual(xlsx_response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            xlsx_response["Content-Type"],
        )

        pdf_response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.pdf")
        self.assertIn(pdf_response.status_code, [200, 503])
        if pdf_response.status_code == 200:
            self.assertIn("application/pdf", pdf_response["Content-Type"])
        else:
            self.assertIn("detail", pdf_response.data)

    def test_validate_token_returns_410_when_token_is_expired(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Expirado")
        raw_token = "VOTO-E2E-EXP-1"
        self._create_token(
            process=process,
            raw_token=raw_token,
            expires_delta_hours=-1,
            status=VoterToken.Status.ACTIVE,
        )

        self.client.force_authenticate(user=None)
        response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )

        self.assertEqual(response.status_code, 410)
        self.assertEqual(response.data["status"], VoterToken.Status.EXPIRED)

    def test_validate_token_returns_403_when_census_member_not_found(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Censo")
        ElectionCensusMember.objects.create(
            student_external_id="EXT-OTRO",
            document_number="DOC-OTRO",
            full_name="Otro Estudiante",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )

        raw_token = "VOTO-E2E-CENSO-1"
        self._create_token(
            process=process,
            raw_token=raw_token,
            metadata={"student_external_id": "EXT-NO-EXISTE", "document_number": "DOC-NO-EXISTE"},
        )

        self.client.force_authenticate(user=None)
        response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertIn("No se encontró", response.data["detail"])

    @override_settings(ELECTIONS_REQUIRE_TOKEN_IDENTITY=True)
    def test_validate_token_returns_403_when_identity_is_required_and_missing(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Identidad Estricta")
        ElectionCensusMember.objects.create(
            student_external_id="EXT-STRICT-1",
            document_number="DOC-STRICT-1",
            full_name="Estudiante Estricto",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )

        raw_token = "VOTO-E2E-STRICT-1"
        self._create_token(
            process=process,
            raw_token=raw_token,
            metadata={},
            student_grade="10",
            student_shift="Mañana",
        )

        self.client.force_authenticate(user=None)
        response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertIn("identidad verificable", response.data["detail"].lower())

    def test_validate_token_returns_409_when_process_window_is_closed(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Cerrada")
        process.ends_at = timezone.now() - timedelta(minutes=1)
        process.save(update_fields=["ends_at"])

        raw_token = "VOTO-E2E-CLOSED-1"
        self._create_token(process=process, raw_token=raw_token)

        self.client.force_authenticate(user=None)
        response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )

        self.assertEqual(response.status_code, 409)
        self.assertIn("no se encuentra abierta", response.data["detail"].lower())

    def test_validate_token_returns_403_when_token_is_revoked(self):
        process, *_ = self._create_process_with_ballot(name="Jornada Revocado")
        raw_token = "VOTO-E2E-REV-1"
        self._create_token(
            process=process,
            raw_token=raw_token,
            status=VoterToken.Status.REVOKED,
            revoked_at=timezone.now() - timedelta(minutes=5),
            revoked_reason="Bloqueado por prueba",
        )

        self.client.force_authenticate(user=None)
        response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.data["status"], VoterToken.Status.REVOKED)

    def test_scrutiny_summary_reflects_total_and_blank_votes(self):
        process, personero_role, contralor_role, personero_candidate, _ = self._create_process_with_ballot(
            name="Jornada Summary"
        )
        ElectionCensusMember.objects.create(
            student_external_id="EXT-100",
            document_number="DOC-100",
            full_name="Estudiante Summary",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )

        raw_token = "VOTO-E2E-SUMMARY-1"
        self._create_token(process=process, raw_token=raw_token)

        self.client.force_authenticate(user=None)
        validate_response = self.client.post(
            "/api/elections/public/validate-token/",
            {"token": raw_token},
            format="json",
        )
        self.assertEqual(validate_response.status_code, 200)

        submit_response = self.client.post(
            "/api/elections/public/submit-vote/",
            {
                "access_session_id": validate_response.data["access_session_id"],
                "selections": [
                    {"role_id": personero_role.id, "candidate_id": personero_candidate.id, "is_blank": False},
                    {"role_id": contralor_role.id, "is_blank": True},
                ],
            },
            format="json",
        )
        self.assertEqual(submit_response.status_code, 201)

        self.client.force_authenticate(user=self.admin)
        summary_response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-summary/")

        self.assertEqual(summary_response.status_code, 200)
        summary = summary_response.data["summary"]
        self.assertEqual(summary["total_votes"], 2)
        self.assertEqual(summary["total_blank_votes"], 1)

        roles = summary_response.data["roles"]
        self.assertEqual(len(roles), 2)
        role_by_id = {row["role_id"]: row for row in roles}
        self.assertEqual(role_by_id[personero_role.id]["total_votes"], 1)
        self.assertEqual(role_by_id[personero_role.id]["blank_votes"], 0)
        self.assertEqual(role_by_id[contralor_role.id]["total_votes"], 1)
        self.assertEqual(role_by_id[contralor_role.id]["blank_votes"], 1)

    def test_scrutiny_export_csv_contains_expected_rows_and_counts(self):
        process, personero_role, contralor_role, personero_candidate, _ = self._create_process_with_ballot(
            name="Jornada CSV"
        )
        self._submit_vote_with_blank_for_second_role(
            process=process,
            personero_role=personero_role,
            contralor_role=contralor_role,
            personero_candidate=personero_candidate,
            token_suffix="CSV-1",
        )

        self.client.force_authenticate(user=self.admin)
        response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.csv")
        self.assertEqual(response.status_code, 200)

        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        self.assertIn(
            ["cargo", "codigo", "numero", "candidato", "votos", "votos_blanco_cargo", "total_cargo"],
            rows,
        )
        self.assertIn(["Personería", "PERSONERO", "01", "Ana Candidata", "1", "0", "1"], rows)
        self.assertIn(["Contraloría", "CONTRALOR", "", "", "0", "1", "1"], rows)

    def test_scrutiny_export_xlsx_contains_expected_rows_and_counts(self):
        process, personero_role, contralor_role, personero_candidate, _ = self._create_process_with_ballot(
            name="Jornada XLSX"
        )
        self._submit_vote_with_blank_for_second_role(
            process=process,
            personero_role=personero_role,
            contralor_role=contralor_role,
            personero_candidate=personero_candidate,
            token_suffix="XLSX-1",
        )

        self.client.force_authenticate(user=self.admin)
        response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.xlsx")
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        rows = [tuple("" if value is None else str(value) for value in row) for row in sheet.iter_rows(values_only=True)]

        self.assertIn(
            ("Cargo", "Código", "Número", "Candidato", "Votos", "Votos en blanco cargo", "Total cargo"),
            rows,
        )
        self.assertIn(("Personería", "PERSONERO", "01", "Ana Candidata", "1", "0", "1"), rows)
        self.assertIn(("Contraloría", "CONTRALOR", "", "", "0", "1", "1"), rows)

    def test_scrutiny_export_pdf_returns_200_with_mocked_renderer(self):
        process, *_ = self._create_process_with_ballot(name="Jornada PDF OK")
        self.client.force_authenticate(user=self.admin)

        with patch("reports.weasyprint_utils.render_pdf_bytes_from_html", return_value=b"%PDF-1.4 mocked"):
            response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.pdf")

        self.assertEqual(response.status_code, 200)
        self.assertIn("application/pdf", response["Content-Type"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_scrutiny_export_pdf_returns_503_when_weasyprint_unavailable(self):
        process, *_ = self._create_process_with_ballot(name="Jornada PDF 503")
        self.client.force_authenticate(user=self.admin)

        with patch(
            "reports.weasyprint_utils.render_pdf_bytes_from_html",
            side_effect=WeasyPrintUnavailableError("WeasyPrint no está disponible"),
        ):
            response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.pdf")

        self.assertEqual(response.status_code, 503)
        self.assertIn("detail", response.data)

    def test_scrutiny_export_pdf_returns_500_on_unexpected_error(self):
        process, *_ = self._create_process_with_ballot(name="Jornada PDF 500")
        self.client.force_authenticate(user=self.admin)

        with patch("reports.weasyprint_utils.render_pdf_bytes_from_html", side_effect=RuntimeError("boom")):
            response = self.client.get(f"/api/elections/manage/processes/{process.id}/scrutiny-export.pdf")

        self.assertEqual(response.status_code, 500)
        self.assertIn("detail", response.data)


class ElectionPermissionsTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username="admin_elections_perm",
            password="pass1234",
            role=user_model.ROLE_ADMIN,
        )
        self.coordinator = user_model.objects.create_user(
            username="coord_elections_perm",
            password="pass1234",
            role=user_model.ROLE_COORDINATOR,
        )
        self.teacher = user_model.objects.create_user(
            username="teacher_elections_perm",
            password="pass1234",
            role=user_model.ROLE_TEACHER,
        )

        now = timezone.now()
        self.process = ElectionProcess.objects.create(
            name="Jornada Permisos",
            status=ElectionProcess.Status.OPEN,
            starts_at=now - timedelta(hours=1),
            ends_at=now + timedelta(hours=2),
        )
        self.raw_token = "VOTO-E2E-PERM-1"
        self.voter_token = VoterToken.objects.create(
            process=self.process,
            token_hash=VoterToken.hash_token(self.raw_token),
            token_prefix=self.raw_token[:12],
            status=VoterToken.Status.ACTIVE,
            expires_at=now + timedelta(hours=2),
        )

    def test_reset_requires_authentication(self):
        self.client.force_authenticate(user=None)
        response = self.client.post(
            "/api/elections/tokens/reset/",
            {
                "token": self.raw_token,
                "reason": "Contingencia de prueba sin autenticación.",
                "extend_hours": 4,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 401)

    def test_reset_allows_coordinator_role(self):
        self.voter_token.status = VoterToken.Status.USED
        self.voter_token.used_at = timezone.now() - timedelta(minutes=1)
        self.voter_token.save(update_fields=["status", "used_at"])

        self.client.force_authenticate(user=self.coordinator)
        response = self.client.post(
            "/api/elections/tokens/reset/",
            {
                "token": self.raw_token,
                "reason": "Contingencia aprobada por coordinación institucional.",
                "extend_hours": 4,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["status"], VoterToken.Status.ACTIVE)

    def test_reset_denies_teacher_role(self):
        self.client.force_authenticate(user=self.teacher)
        response = self.client.post(
            "/api/elections/tokens/reset/",
            {
                "token": self.raw_token,
                "reason": "Intento de reset por usuario sin permiso docente.",
                "extend_hours": 4,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 403)

    def test_scrutiny_summary_requires_admin_role(self):
        self.client.force_authenticate(user=self.teacher)
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/scrutiny-summary/")
        self.assertEqual(response.status_code, 403)

    def test_scrutiny_summary_requires_authentication(self):
        self.client.force_authenticate(user=None)
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/scrutiny-summary/")
        self.assertEqual(response.status_code, 401)

    def test_scrutiny_exports_require_authentication(self):
        self.client.force_authenticate(user=None)
        export_paths = [
            "scrutiny-export.csv",
            "scrutiny-export.xlsx",
            "scrutiny-export.pdf",
        ]

        for export_path in export_paths:
            response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/{export_path}")
            self.assertEqual(response.status_code, 401)

    def test_scrutiny_exports_require_admin_role(self):
        self.client.force_authenticate(user=self.teacher)
        export_paths = [
            "scrutiny-export.csv",
            "scrutiny-export.xlsx",
            "scrutiny-export.pdf",
        ]

        for export_path in export_paths:
            response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/{export_path}")
            self.assertEqual(response.status_code, 403)

    def test_create_process_forces_draft_even_if_open_is_requested(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.post(
            "/api/elections/manage/processes/",
            {
                "name": "Jornada Forzada Borrador",
                "status": ElectionProcess.Status.OPEN,
                "starts_at": (timezone.now() + timedelta(hours=1)).isoformat(),
                "ends_at": (timezone.now() + timedelta(hours=3)).isoformat(),
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["status"], ElectionProcess.Status.DRAFT)

        created_process = ElectionProcess.objects.get(id=response.data["id"])
        self.assertEqual(created_process.status, ElectionProcess.Status.DRAFT)


class ElectionAuditTrailTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username="admin_elections_audit",
            password="pass1234",
            role=user_model.ROLE_ADMIN,
        )

        now = timezone.now()
        self.process = ElectionProcess.objects.create(
            name="Jornada Auditoría",
            status=ElectionProcess.Status.DRAFT,
            starts_at=now + timedelta(hours=1),
            ends_at=now + timedelta(hours=3),
        )
        self.member = ElectionCensusMember.objects.create(
            student_external_id="EXT-AUD-1",
            document_number="DOC-AUD-1",
            full_name="Estudiante Auditoría",
            grade="11",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )

        self.client.force_authenticate(user=self.admin)

    def test_open_process_creates_audit_log(self):
        response = self.client.post(f"/api/elections/manage/processes/{self.process.id}/open/", format="json")
        self.assertEqual(response.status_code, 200)

        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_PROCESS_OPEN",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            ).exists()
        )

    def test_scrutiny_csv_export_creates_audit_log(self):
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/scrutiny-export.csv")
        self.assertEqual(response.status_code, 200)

        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_SCRUTINY_EXPORT_CSV",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            ).exists()
        )

    def test_census_exclude_and_include_create_audit_logs(self):
        exclude_response = self.client.post(
            f"/api/elections/manage/processes/{self.process.id}/census/exclusions/",
            {"member_id": self.member.id, "reason": "Prueba de exclusión."},
            format="json",
        )
        self.assertEqual(exclude_response.status_code, 200)

        include_response = self.client.delete(
            f"/api/elections/manage/processes/{self.process.id}/census/exclusions/{self.member.id}/",
        )
        self.assertEqual(include_response.status_code, 204)

        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_CENSUS_MEMBER_EXCLUDE",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_CENSUS_MEMBER_INCLUDE",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=204,
            ).exists()
        )

    def test_census_exports_create_audit_logs(self):
        xlsx_response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/census/manual-codes.xlsx")
        self.assertEqual(xlsx_response.status_code, 200)

        qr_response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/census/qr-print/")
        self.assertEqual(qr_response.status_code, 200)

        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_CENSUS_MANUAL_CODES_EXPORT",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_CENSUS_QR_PRINT",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            ).exists()
        )

    def test_manual_codes_regenerate_requires_confirmation_and_reason(self):
        response_without_confirmation = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/census/manual-codes.xlsx",
            {"mode": "regenerate"},
        )
        self.assertEqual(response_without_confirmation.status_code, 400)
        self.assertIn("confirmar", response_without_confirmation.data["detail"].lower())

        response_without_reason = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/census/manual-codes.xlsx",
            {"mode": "regenerate", "confirm_regeneration": "true"},
        )
        self.assertEqual(response_without_reason.status_code, 400)
        self.assertIn("motivo", response_without_reason.data["detail"].lower())

    def test_manual_codes_existing_mode_reuses_codes_without_regeneration(self):
        now = timezone.now()
        existing_code = "VOTO-EXISTING-001"
        VoterToken.objects.create(
            process=self.process,
            token_hash=VoterToken.hash_token(existing_code),
            token_prefix=existing_code[:12],
            status=VoterToken.Status.ACTIVE,
            expires_at=now + timedelta(hours=2),
            student_grade=self.member.grade,
            student_shift=self.member.shift,
            metadata={
                "student_external_id": self.member.student_external_id,
                "document_number": self.member.document_number,
                "full_name": self.member.full_name,
                "manual_code": existing_code,
                "issued_from": "process_census",
            },
        )

        tokens_before = VoterToken.objects.filter(process=self.process).count()
        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/census/manual-codes.xlsx",
            {"mode": "existing"},
        )
        self.assertEqual(response.status_code, 200)

        tokens_after = VoterToken.objects.filter(process=self.process).count()
        self.assertEqual(tokens_after, tokens_before)

        export_log = (
            AuditLog.objects.filter(
                event_type="ELECTION_CENSUS_MANUAL_CODES_EXPORT",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            )
            .order_by("-created_at", "-id")
            .first()
        )
        self.assertIsNotNone(export_log)
        metadata = export_log.metadata if isinstance(export_log.metadata, dict) else {}
        self.assertEqual(metadata.get("mode"), "existing")
        self.assertEqual(int(metadata.get("generated_count") or 0), 0)


class ElectionObserverCongratsTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username="admin_elections_observer",
            password="pass1234",
            role=user_model.ROLE_ADMIN,
        )

        now = timezone.now()
        self.process = ElectionProcess.objects.create(
            name="Jornada Felicitaciones",
            status=ElectionProcess.Status.OPEN,
            starts_at=now - timedelta(hours=1),
            ends_at=now + timedelta(hours=2),
        )
        self.personero_role = ElectionRole.objects.create(
            process=self.process,
            code=ElectionRole.CODE_PERSONERO,
            title="Personería",
            display_order=1,
        )
        self.contralor_role = ElectionRole.objects.create(
            process=self.process,
            code=ElectionRole.CODE_CONTRALOR,
            title="Contraloría",
            display_order=2,
        )

        self.student_personero_a = self._create_student("stud_personero_a", "Pérez", "Ana", "DOC-P-1")
        self.student_personero_b = self._create_student("stud_personero_b", "Gómez", "Bruno", "DOC-P-2")
        self.student_contralor = self._create_student("stud_contralor", "Díaz", "Carla", "DOC-C-1")
        self.student_inactive = self._create_student("stud_inactive", "Rojas", "Diego", "DOC-I-1")

        self.personero_a = ElectionCandidate.objects.create(
            role=self.personero_role,
            name="Ana Personera",
            student_id_ref=self.student_personero_a.user_id,
            student_document_number="DOC-P-1",
            number="01",
            grade="11",
            is_active=True,
            display_order=1,
        )
        self.personero_b = ElectionCandidate.objects.create(
            role=self.personero_role,
            name="Bruno Personero",
            student_id_ref=self.student_personero_b.user_id,
            student_document_number="DOC-P-2",
            number="02",
            grade="11",
            is_active=True,
            display_order=2,
        )
        self.contralor = ElectionCandidate.objects.create(
            role=self.contralor_role,
            name="Carla Contralora",
            student_id_ref=self.student_contralor.user_id,
            student_document_number="DOC-C-1",
            number="03",
            grade="10",
            is_active=True,
            display_order=1,
        )
        self.inactive_candidate = ElectionCandidate.objects.create(
            role=self.contralor_role,
            name="Diego Inactivo",
            student_id_ref=self.student_inactive.user_id,
            student_document_number="DOC-I-1",
            number="04",
            grade="10",
            is_active=False,
            display_order=2,
        )

        self._create_vote(self.personero_role, self.personero_a, suffix="P1")
        self._create_vote(self.personero_role, self.personero_b, suffix="P2")
        self._create_vote(self.contralor_role, self.contralor, suffix="C1")
        self._create_vote(self.contralor_role, self.contralor, suffix="C2")

        self.client.force_authenticate(user=self.admin)

    def _create_student(self, username: str, last_name: str, first_name: str, document_number: str) -> Student:
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username=username,
            password="pass1234",
            role=user_model.ROLE_STUDENT,
            first_name=first_name,
            last_name=last_name,
        )
        return Student.objects.create(user=user, document_number=document_number)

    def _create_vote(self, role: ElectionRole, candidate: ElectionCandidate, *, suffix: str) -> None:
        raw_token = f"VOTO-OBS-{suffix}"
        token = VoterToken.objects.create(
            process=self.process,
            token_hash=VoterToken.hash_token(raw_token),
            token_prefix=raw_token[:12],
            status=VoterToken.Status.USED,
            expires_at=timezone.now() + timedelta(hours=1),
            used_at=timezone.now(),
            student_grade="11",
            student_shift="Mañana",
        )
        access_session = VoteAccessSession.objects.create(
            voter_token=token,
            expires_at=timezone.now() + timedelta(minutes=10),
            consumed_at=timezone.now(),
        )
        VoteRecord.objects.create(
            process=self.process,
            role=role,
            candidate=candidate,
            voter_token=token,
            access_session=access_session,
            is_blank=False,
        )

    @patch("elections.services_observer.AIService.improve_text", side_effect=lambda text: f"IA::{text}")
    def test_close_process_generates_ai_observer_annotations(self, _mock_improve):
        response = self.client.post(f"/api/elections/manage/processes/{self.process.id}/close/", format="json")
        self.assertEqual(response.status_code, 200)

        self.process.refresh_from_db()
        self.assertEqual(self.process.status, ElectionProcess.Status.CLOSED)

        self.assertEqual(
            ObserverAnnotation.objects.filter(
                student_id=self.student_personero_a.user_id,
                is_deleted=False,
            ).count(),
            2,
        )
        self.assertEqual(
            ObserverAnnotation.objects.filter(
                student_id=self.student_personero_b.user_id,
                is_deleted=False,
            ).count(),
            2,
        )
        self.assertEqual(
            ObserverAnnotation.objects.filter(
                student_id=self.student_contralor.user_id,
                is_deleted=False,
            ).count(),
            2,
        )
        self.assertEqual(
            ObserverAnnotation.objects.filter(
                student_id=self.student_inactive.user_id,
                is_deleted=False,
            ).count(),
            0,
        )

        self.assertEqual(ObserverAnnotation.objects.filter(is_deleted=False).count(), 6)
        self.assertTrue(
            ObserverAnnotation.objects.filter(rule_key=f"ELECTION_WINNER:{self.process.id}:PERSONERO:{self.personero_a.id}").exists()
        )
        self.assertTrue(
            ObserverAnnotation.objects.filter(rule_key=f"ELECTION_WINNER:{self.process.id}:PERSONERO:{self.personero_b.id}").exists()
        )
        self.assertTrue(
            ObserverAnnotation.objects.filter(rule_key=f"ELECTION_WINNER:{self.process.id}:CONTRALOR:{self.contralor.id}").exists()
        )

        self.assertTrue(
            ObserverAnnotation.objects.filter(
                created_by=self.admin,
                text__startswith="IA::",
                is_deleted=False,
            ).exists()
        )

        self.assertTrue(
            AuditLog.objects.filter(
                event_type="ELECTION_PROCESS_CLOSE",
                object_type="ElectionProcess",
                object_id=str(self.process.id),
                actor=self.admin,
                status_code=200,
            ).exists()
        )

    @patch("elections.services_observer.AIService.improve_text", side_effect=lambda text: f"IA::{text}")
    def test_list_processes_includes_persisted_congrats_summary(self, _mock_improve):
        close_response = self.client.post(f"/api/elections/manage/processes/{self.process.id}/close/", format="json")
        self.assertEqual(close_response.status_code, 200)

        list_response = self.client.get("/api/elections/manage/processes/")
        self.assertEqual(list_response.status_code, 200)

        process_item = next((item for item in list_response.data["results"] if item["id"] == self.process.id), None)
        self.assertIsNotNone(process_item)
        assert process_item is not None

        self.assertTrue(process_item.get("observer_congrats_generated"))
        summary = process_item.get("observer_congrats_summary") or {}
        self.assertEqual(summary.get("process_id"), self.process.id)
        self.assertEqual(summary.get("winner_annotations_created"), 3)
        self.assertEqual(summary.get("participant_annotations_created"), 3)

    @patch("elections.services_observer.AIService.improve_text", side_effect=Exception("provider down"))
    def test_service_fallback_and_idempotency_for_observer_annotations(self, _mock_improve):
        first_run = generate_observer_congratulations_for_election(
            process_id=self.process.id,
            created_by_id=self.admin.id,
        )
        self.assertEqual(first_run["winner_annotations_created"], 3)
        self.assertEqual(first_run["participant_annotations_created"], 3)
        self.assertEqual(first_run["fallback_messages"], 6)
        self.assertEqual(ObserverAnnotation.objects.filter(is_deleted=False).count(), 6)

        second_run = generate_observer_congratulations_for_election(
            process_id=self.process.id,
            created_by_id=self.admin.id,
        )
        self.assertEqual(second_run["winner_annotations_created"], 0)
        self.assertEqual(second_run["participant_annotations_created"], 0)
        self.assertEqual(second_run["winner_annotations_updated"], 3)
        self.assertEqual(second_run["participant_annotations_updated"], 3)
        self.assertEqual(ObserverAnnotation.objects.filter(is_deleted=False).count(), 6)

        self.assertFalse(
            ObserverAnnotation.objects.filter(
                meta__generated_by_ai=True,
                is_deleted=False,
            ).exists()
        )


class ElectionLiveDashboardTests(APITestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username="admin_elections_live",
            password="pass1234",
            role=user_model.ROLE_ADMIN,
        )
        self.teacher = user_model.objects.create_user(
            username="teacher_elections_live",
            password="pass1234",
            role=user_model.ROLE_TEACHER,
        )

        now = timezone.now()
        self.process = ElectionProcess.objects.create(
            name="Jornada Live",
            status=ElectionProcess.Status.OPEN,
            starts_at=now - timedelta(hours=1),
            ends_at=now + timedelta(hours=2),
        )
        self.personero_role = ElectionRole.objects.create(
            process=self.process,
            code=ElectionRole.CODE_PERSONERO,
            title="Personería",
            display_order=1,
        )
        self.personero_candidate = ElectionCandidate.objects.create(
            role=self.personero_role,
            name="Ana Live",
            number="01",
            grade="11",
            is_active=True,
            display_order=1,
        )

        ElectionCensusMember.objects.create(
            student_external_id="EXT-LIVE-1",
            document_number="DOC-LIVE-1",
            full_name="Estudiante Live Uno",
            grade="11",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )
        ElectionCensusMember.objects.create(
            student_external_id="EXT-LIVE-2",
            document_number="DOC-LIVE-2",
            full_name="Estudiante Live Dos",
            grade="10",
            shift="Mañana",
            is_active=True,
            status=ElectionCensusMember.Status.ACTIVE,
        )

        token = VoterToken.objects.create(
            process=self.process,
            token_hash=VoterToken.hash_token("VOTO-LIVE-0001"),
            token_prefix="VOTO-LIVE-00",
            status=VoterToken.Status.USED,
            expires_at=now + timedelta(hours=1),
            used_at=now - timedelta(minutes=5),
            student_grade="11",
            student_shift="Mañana",
        )
        access_session = VoteAccessSession.objects.create(
            voter_token=token,
            expires_at=now + timedelta(minutes=10),
            consumed_at=now - timedelta(minutes=5),
        )
        VoteRecord.objects.create(
            process=self.process,
            role=self.personero_role,
            candidate=self.personero_candidate,
            voter_token=token,
            access_session=access_session,
            is_blank=False,
            created_at=now - timedelta(minutes=2),
        )

    def test_live_dashboard_requires_authentication(self):
        self.client.force_authenticate(user=None)
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/live-dashboard/")
        self.assertEqual(response.status_code, 401)

    def test_live_dashboard_requires_admin_role(self):
        self.client.force_authenticate(user=self.teacher)
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/live-dashboard/")
        self.assertEqual(response.status_code, 403)

    def test_live_dashboard_returns_expected_payload(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?window_minutes=60")

        self.assertEqual(response.status_code, 200)
        self.assertIn("process", response.data)
        self.assertIn("kpis", response.data)
        self.assertIn("operational_kpis", response.data)
        self.assertIn("ranking", response.data)
        self.assertIn("minute_series", response.data)
        self.assertIn("alerts", response.data)

        kpis = response.data["kpis"]
        self.assertEqual(kpis["total_votes"], 1)
        self.assertEqual(kpis["total_blank_votes"], 0)
        self.assertEqual(kpis["enabled_census_count"], 2)
        self.assertEqual(kpis["unique_voters_count"], 1)
        self.assertEqual(kpis["participation_percent"], 50.0)

        operational_kpis = response.data["operational_kpis"]
        self.assertIn("audited_events", operational_kpis)
        self.assertIn("client_errors", operational_kpis)
        self.assertIn("server_errors", operational_kpis)
        self.assertIn("duplicate_submits", operational_kpis)

    def test_live_dashboard_applies_custom_alert_config(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?window_minutes=90&blank_rate_threshold=0.3&inactivity_minutes=7&spike_threshold=12&series_limit=20"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("config", response.data)
        config = response.data["config"]
        self.assertEqual(config["window_minutes"], 90)
        self.assertEqual(config["blank_rate_threshold"], 0.3)
        self.assertEqual(config["inactivity_minutes"], 7)
        self.assertEqual(config["spike_threshold"], 12)
        self.assertEqual(config["series_limit"], 20)

    def test_live_dashboard_rejects_invalid_threshold_params(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?blank_rate_threshold=1.5"
        )
        self.assertEqual(response.status_code, 400)

        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?inactivity_minutes=0"
        )
        self.assertEqual(response.status_code, 400)

        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?spike_threshold=0"
        )
        self.assertEqual(response.status_code, 400)

        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?series_limit=2"
        )
        self.assertEqual(response.status_code, 400)

    def test_live_dashboard_rejects_invalid_since_param(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/?since=invalid-date"
        )
        self.assertEqual(response.status_code, 400)

    def test_live_dashboard_supports_incremental_without_ranking(self):
        self.client.force_authenticate(user=self.admin)
        since = (timezone.now() - timedelta(minutes=30)).isoformat()
        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/",
            {
                "since": since,
                "include_ranking": "false",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["is_incremental"])
        self.assertIn("cursor", response.data)
        self.assertEqual(response.data["ranking"], [])

    def test_live_dashboard_stream_requires_authentication(self):
        self.client.force_authenticate(user=None)
        response = self.client.get(f"/api/elections/manage/processes/{self.process.id}/live-dashboard/stream/")
        self.assertEqual(response.status_code, 401)

    def test_live_dashboard_stream_emits_snapshot_event(self):
        self.client.force_authenticate(user=self.admin)
        response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/stream/",
            {"window_minutes": 60},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/event-stream")

        content = "".join(chunk.decode("utf-8") if isinstance(chunk, bytes) else chunk for chunk in response.streaming_content)
        self.assertIn("event: snapshot", content)

        data_lines = [line for line in content.splitlines() if line.startswith("data: ")]
        self.assertGreaterEqual(len(data_lines), 1)
        payload = json.loads(data_lines[0][len("data: "):])
        self.assertEqual(payload["process"]["id"], self.process.id)
        self.assertIn("kpis", payload)

    @patch("elections.views_management.build_live_dashboard_payload")
    def test_live_dashboard_uses_cached_snapshot_for_same_params(self, mock_build_live_payload):
        now = timezone.now()
        mock_build_live_payload.return_value = {
            "generated_at": now,
            "cursor": now,
            "is_incremental": False,
            "process": {"id": self.process.id, "name": self.process.name, "status": self.process.status},
            "config": {
                "window_minutes": 60,
                "blank_rate_threshold": 0.25,
                "inactivity_minutes": 10,
                "spike_threshold": 8,
                "series_limit": 60,
            },
            "kpis": {
                "total_votes": 1,
                "total_blank_votes": 0,
                "blank_vote_percent": 0,
                "enabled_census_count": 2,
                "unique_voters_count": 1,
                "participation_percent": 50.0,
            },
            "ranking": [],
            "minute_series": [],
            "alerts": [],
        }

        self.client.force_authenticate(user=self.admin)
        first_response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/",
            {"window_minutes": 60},
        )
        second_response = self.client.get(
            f"/api/elections/manage/processes/{self.process.id}/live-dashboard/",
            {"window_minutes": 60},
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        self.assertEqual(mock_build_live_payload.call_count, 1)
