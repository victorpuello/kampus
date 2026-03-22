from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.conf import settings
from django.http import HttpResponse, StreamingHttpResponse
from django.template.loader import render_to_string
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from difflib import SequenceMatcher
import io
import csv
import json
import re
import unicodedata

from audit.models import AuditLog
from audit.services import log_event


def _period_end_of_day(period: "Period"):
    """Fallback deadline when explicit edit_until is not configured.

    Uses the period end_date at 23:59:59 in the current timezone.
    """

    if getattr(period, "end_date", None) is None:
        return None
    tz = timezone.get_current_timezone()
    dt = datetime.combine(period.end_date, time(23, 59, 59))
    return timezone.make_aware(dt, tz)


TOPIC_IMPORT_COLUMNS = [
    "academic_year",
    "period_name",
    "grade_name",
    "subject_name",
    "sequence_order",
    "title",
    "description",
]


def _normalize_topic_import_header(value):
    return str(value or "").strip().lower()


def _normalize_topic_lookup_value(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.replace("&", " y ")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_topic_lookup_loose(value):
    stop_words = {"y", "e", "de", "del", "la", "las", "el", "los"}
    normalized = _normalize_topic_lookup_value(value)
    tokens = [token for token in normalized.split() if token not in stop_words]
    return " ".join(tokens)


def _grade_lookup_keys(grade_name):
    keys = {
        _normalize_topic_lookup_value(grade_name),
        _normalize_topic_lookup_loose(grade_name),
    }
    ordinal = guess_ordinal(str(grade_name or ""))
    if ordinal is not None:
        keys.add(str(ordinal))
    return {key for key in keys if key}


def _resolve_topic_import_academic_load(grade_name, subject_name, loads_by_grade_key):
    grade_keys = _grade_lookup_keys(grade_name)
    candidate_loads = []
    seen_load_ids = set()
    for grade_key in grade_keys:
        for load in loads_by_grade_key.get(grade_key, []):
            if load.id in seen_load_ids:
                continue
            candidate_loads.append(load)
            seen_load_ids.add(load.id)

    if not candidate_loads:
        return None, None

    subject_exact = _normalize_topic_lookup_value(subject_name)
    subject_loose = _normalize_topic_lookup_loose(subject_name)

    for load in candidate_loads:
        if _normalize_topic_lookup_value(load.subject.name) == subject_exact:
            return load, None
    for load in candidate_loads:
        if _normalize_topic_lookup_loose(load.subject.name) == subject_loose:
            return load, None

    area_matches = []
    for load in candidate_loads:
        area_name = getattr(getattr(load, "subject", None), "area", None)
        area_label = getattr(area_name, "name", "") if area_name is not None else ""
        if not area_label:
            continue
        if _normalize_topic_lookup_value(area_label) == subject_exact or _normalize_topic_lookup_loose(area_label) == subject_loose:
            area_matches.append(load)
    if len(area_matches) == 1:
        return area_matches[0], None

    scored_candidates = []
    for load in candidate_loads:
        subject_score = SequenceMatcher(None, subject_loose, _normalize_topic_lookup_loose(load.subject.name)).ratio()
        area_name = getattr(getattr(load, "subject", None), "area", None)
        area_label = getattr(area_name, "name", "") if area_name is not None else ""
        area_score = SequenceMatcher(None, subject_loose, _normalize_topic_lookup_loose(area_label)).ratio() if area_label else 0.0
        score = max(subject_score, area_score)
        scored_candidates.append((score, load))

    scored_candidates.sort(key=lambda item: item[0], reverse=True)
    if not scored_candidates:
        return None, None

    best_score, best_load = scored_candidates[0]
    second_score = scored_candidates[1][0] if len(scored_candidates) > 1 else 0.0
    if best_score >= 0.84 and (best_score - second_score >= 0.08 or len(scored_candidates) == 1):
        return best_load, None

    suggestions = []
    for _, load in scored_candidates[:3]:
        suggestion = load.subject.name
        if suggestion not in suggestions:
            suggestions.append(suggestion)
    return None, suggestions


def _build_topic_import_context():
    period_map = {
        (str(period.academic_year.year).strip().lower(), str(period.name).strip().lower()): period
        for period in Period.objects.select_related("academic_year").all()
    }
    all_loads = list(AcademicLoad.objects.select_related("grade", "subject", "subject__area").all())
    loads_by_grade_key: dict[str, list[AcademicLoad]] = {}
    for load in all_loads:
        for grade_key in _grade_lookup_keys(load.grade.name):
            loads_by_grade_key.setdefault(grade_key, []).append(load)
    return period_map, loads_by_grade_key


def _parse_topic_import_corrections(raw_corrections):
    if not raw_corrections:
        return {}
    if isinstance(raw_corrections, str):
        try:
            raw_corrections = json.loads(raw_corrections)
        except Exception:
            return {}
    if not isinstance(raw_corrections, list):
        return {}

    corrections: dict[int, dict[str, str]] = {}
    for item in raw_corrections:
        if not isinstance(item, dict):
            continue
        row_number = item.get("row_number")
        try:
            parsed_row_number = int(row_number)
        except Exception:
            continue
        subject_name = str(item.get("subject_name") or "").strip()
        if not subject_name:
            continue
        corrections[parsed_row_number] = {"subject_name": subject_name}
    return corrections


def _apply_topic_import_correction(row, correction):
    corrected_row = dict(row)
    if correction and correction.get("subject_name"):
        corrected_row["subject_name"] = correction["subject_name"]
    return corrected_row


def _validate_topic_import_row(row_number, row, period_map, loads_by_grade_key, corrections=None):
    corrected_row = _apply_topic_import_correction(row, (corrections or {}).get(row_number))

    academic_year = str(corrected_row.get("academic_year") or "").strip()
    period_name = str(corrected_row.get("period_name") or "").strip()
    grade_name = str(corrected_row.get("grade_name") or "").strip()
    subject_name = str(corrected_row.get("subject_name") or "").strip()
    sequence_order_raw = str(corrected_row.get("sequence_order") or "").strip()
    title = str(corrected_row.get("title") or "").strip()
    description = str(corrected_row.get("description") or "").strip()

    result = {
        "row_number": row_number,
        "academic_year": academic_year,
        "period_name": period_name,
        "grade_name": grade_name,
        "subject_name": subject_name,
        "sequence_order": sequence_order_raw,
        "title": title,
        "description": description,
        "status": "ready",
        "message": "Lista para importar.",
        "suggestions": [],
        "resolved_subject_name": None,
    }

    if not all([academic_year, period_name, grade_name, subject_name, sequence_order_raw, title]):
        result["status"] = "error"
        result["message"] = "Faltan columnas obligatorias."
        return result, None, None

    try:
        sequence_order = int(sequence_order_raw)
    except Exception:
        result["status"] = "error"
        result["message"] = "sequence_order inválido."
        return result, None, None

    period = period_map.get((academic_year.lower(), period_name.lower()))
    if period is None:
        result["status"] = "error"
        result["message"] = f"No se encontró el periodo '{period_name}' para el año {academic_year}."
        return result, None, None

    academic_load, suggestions = _resolve_topic_import_academic_load(grade_name, subject_name, loads_by_grade_key)
    if academic_load is None:
        result["status"] = "review" if suggestions else "error"
        result["message"] = (
            f"Se requiere confirmar la asignatura para '{subject_name}'."
            if suggestions
            else f"No se encontró la asignatura o área '{subject_name}' para el grado '{grade_name}'."
        )
        result["suggestions"] = suggestions or []
        return result, None, None

    result["resolved_subject_name"] = academic_load.subject.name
    return result, period, {
        "academic_load": academic_load,
        "sequence_order": sequence_order,
        "title": title,
        "description": description,
    }


def _build_topic_import_validation(data_rows, period_map, loads_by_grade_key, corrections=None):
    rows = []
    ready_rows = 0
    review_rows = 0
    error_rows = 0
    resolved_rows = []

    for row_number, row in enumerate(data_rows, start=2):
        row_result, period, resolved = _validate_topic_import_row(
            row_number,
            row,
            period_map,
            loads_by_grade_key,
            corrections=corrections,
        )
        rows.append(row_result)
        if row_result["status"] == "ready":
            ready_rows += 1
            resolved_rows.append((row_result, period, resolved))
        elif row_result["status"] == "review":
            review_rows += 1
        else:
            error_rows += 1

    return {
        "total_rows": len(data_rows),
        "ready_rows": ready_rows,
        "review_rows": review_rows,
        "error_rows": error_rows,
        "rows": rows,
        "resolved_rows": resolved_rows,
    }


def _read_topic_import_rows(upload):
    filename = str(getattr(upload, "name", "") or "").lower()
    raw_bytes = upload.read()

    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook  # noqa: PLC0415

        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    elif filename.endswith(".xls"):
        import xlrd  # noqa: PLC0415

        workbook = xlrd.open_workbook(file_contents=raw_bytes)
        sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
    else:
        raw_text = raw_bytes.decode("utf-8-sig")
        reader = csv.reader(raw_text.splitlines())
        rows = list(reader)

    if not rows:
        return [], []

    headers = [_normalize_topic_import_header(cell) for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if row is None:
            continue
        values = list(row)
        if not any(str(cell or "").strip() for cell in values):
            continue
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = "" if index >= len(values) or values[index] is None else str(values[index]).strip()
        data_rows.append(record)

    return headers, data_rows

from .models import (
    AcademicLevel,
    AcademicYear,
    Achievement,
    AchievementDefinition,
    Area,
    Assessment,
    Dimension,
    EvaluationComponent,
    EvaluationScale,
    Grade,
    Group,
    PerformanceIndicator,
    Period,
    StudentGrade,
    Subject,
    TeacherAssignment,
    AcademicLoad,
    GradeSheet,
    AchievementGrade,
    AchievementActivityColumn,
    AchievementActivityGrade,
    EditRequest,
    EditRequestItem,
    EditGrant,
    EditGrantItem,
    PeriodTopic,
    ClassPlan,
)

from students.academic_period_report import (
    generate_academic_period_group_report_pdf,
    generate_preschool_academic_period_group_report_pdf,
)
from students.models import Enrollment, FamilyMember
from core.models import Institution
from core.permissions import KampusModelPermissions
from .serializers import (
    AcademicLevelSerializer,
    AcademicYearSerializer,
    AchievementDefinitionSerializer,
    AchievementSerializer,
    AreaSerializer,
    AssessmentSerializer,
    DimensionSerializer,
    EvaluationComponentSerializer,
    EvaluationScaleSerializer,
    GradeSerializer,
    GroupSerializer,
    PerformanceIndicatorSerializer,
    PeriodSerializer,
    StudentGradeSerializer,
    SubjectSerializer,
    TeacherAssignmentSerializer,
    AcademicLoadSerializer,
    GradeSheetSerializer,
    GradeSheetListSerializer,
    GradebookBulkUpsertSerializer,
    PreschoolGradebookBulkUpsertSerializer,
    PreschoolGradebookLabelSerializer,
    GradeSheetGradingModeSerializer,
    AchievementActivityColumnSerializer,
    ActivityColumnsBulkUpsertSerializer,
    ActivityGradesBulkUpsertSerializer,
    EditRequestSerializer,
    EditRequestDecisionSerializer,
    EditGrantSerializer,
    PeriodTopicSerializer,
    ClassPlanSerializer,
)
from .ai import AIService, AIConfigError, AIParseError, AIProviderError
from .throttles import AcademicAIUserRateThrottle
from .grade_ordinals import guess_ordinal
from .grading import (
    DEFAULT_EMPTY_SCORE,
    coalesce_score,
    final_grade_from_dimensions,
    match_scale,
    weighted_average,
)
from .promotion import compute_promotions_for_year, PASSING_SCORE_DEFAULT
from communications.email_service import send_email
from reports.models import ReportJob
from reports.serializers import ReportJobSerializer
from reports.tasks import generate_report_job_pdf


class AcademicYearViewSet(viewsets.ModelViewSet):
    queryset = AcademicYear.objects.all()
    serializer_class = AcademicYearSerializer
    permission_classes = [KampusModelPermissions]

    @action(detail=True, methods=["get"], url_path="promotion-preview")
    def promotion_preview(self, request, pk=None):
        year: AcademicYear = self.get_object()

        grade_id_raw = request.query_params.get("grade_id")
        grade_id = None
        if grade_id_raw not in (None, ""):
            try:
                grade_id = int(str(grade_id_raw))
            except Exception:
                return Response({"detail": "grade_id inválido"}, status=status.HTTP_400_BAD_REQUEST)

        passing_score_raw = request.query_params.get("passing_score")
        try:
            passing_score = PASSING_SCORE_DEFAULT if not passing_score_raw else Decimal(str(passing_score_raw))
        except Exception:
            return Response({"detail": "passing_score inválido"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            computed = compute_promotions_for_year(academic_year=year, passing_score=passing_score)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        enrollments_qs = Enrollment.objects.filter(id__in=list(computed.keys())).select_related(
            "student",
            "student__user",
            "grade",
        )
        if grade_id is not None:
            enrollments_qs = enrollments_qs.filter(grade_id=grade_id)

        # Infer grade progression for UI (target grade per decision)
        grades = list(Grade.objects.all().only("id", "ordinal", "name", "level_id"))
        grade_name_by_id = {int(g.id): getattr(g, "name", "") for g in grades}
        ordinal_to_grade_id_by_level: dict[int, dict[int, int]] = {}
        ordinal_to_grade_id_global: dict[int, int] = {}
        for g in grades:
            ord_val = getattr(g, "ordinal", None)
            if ord_val is None:
                ord_val = guess_ordinal(getattr(g, "name", ""))
            if ord_val is None:
                continue
            o = int(ord_val)
            gid = int(g.id)
            lvl = int(getattr(g, "level_id", 0) or 0)
            ordinal_to_grade_id_by_level.setdefault(lvl, {}).setdefault(o, gid)
            ordinal_to_grade_id_global.setdefault(o, gid)

        def _grade_ordinal(current_grade: Grade) -> int | None:
            ord_val = getattr(current_grade, "ordinal", None)
            if ord_val is not None:
                return int(ord_val)
            guessed = guess_ordinal(getattr(current_grade, "name", ""))
            return int(guessed) if guessed is not None else None

        def next_grade_id(current_grade: Grade) -> int | None:
            ord_val = _grade_ordinal(current_grade)
            if ord_val is None:
                return None
            lvl = int(getattr(current_grade, "level_id", 0) or 0)
            per_level = ordinal_to_grade_id_by_level.get(lvl) or {}
            return per_level.get(ord_val + 1) or ordinal_to_grade_id_global.get(ord_val + 1)

        def is_last_grade(current_grade: Grade) -> bool:
            ord_val = _grade_ordinal(current_grade)
            if ord_val is not None:
                return ord_val >= 13
            return str(getattr(current_grade, "name", "")).strip().lower() in {"11", "11°", "once", "undecimo", "undécimo"}

        results = []
        for e in enrollments_qs.order_by("id"):
            c = computed.get(e.id)
            if not c:
                continue
            u = getattr(e.student, "user", None)
            student_name = ""
            if u is not None:
                student_name = u.get_full_name()

            grade_ord = _grade_ordinal(e.grade)
            target_grade_id = None
            if c.decision in {"PROMOTED", "CONDITIONAL"}:
                if not is_last_grade(e.grade):
                    target_grade_id = next_grade_id(e.grade)
            elif c.decision == "REPEATED":
                target_grade_id = int(e.grade_id)

            results.append(
                {
                    "enrollment_id": e.id,
                    "decision": c.decision,
                    "failed_subjects_count": len(c.failed_subject_ids),
                    "failed_areas_count": len(c.failed_area_ids),
                    "failed_subjects_distinct_areas_count": c.failed_subjects_distinct_areas_count,
                    "failed_subject_ids": c.failed_subject_ids,
                    "failed_area_ids": c.failed_area_ids,
                    "student_id": int(e.student_id),
                    "student_name": student_name,
                    "student_document_number": getattr(e.student, "document_number", ""),
                    "grade_id": int(e.grade_id),
                    "grade_name": getattr(e.grade, "name", ""),
                    "grade_ordinal": grade_ord,
                    "target_grade_id": int(target_grade_id) if target_grade_id is not None else None,
                    "target_grade_name": grade_name_by_id.get(int(target_grade_id)) if target_grade_id is not None else None,
                }
            )

        results.sort(key=lambda x: (x["decision"], x["enrollment_id"]))
        return Response(
            {
                "academic_year": {"id": year.id, "year": year.year, "status": year.status},
                "passing_score": str(passing_score),
                "count": len(results),
                "results": results,
            }
        )

    @action(detail=True, methods=["post"], url_path="close-with-promotion")
    def close_with_promotion(self, request, pk=None):
        from decimal import Decimal

        from students.models import Enrollment
        from .models import EnrollmentPromotionSnapshot, Period

        year: AcademicYear = self.get_object()

        # Require all periods closed before closing academic year.
        if Period.objects.filter(academic_year=year, is_closed=False).exists():
            return Response(
                {"detail": "No se puede cerrar el año: hay periodos abiertos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        passing_score_raw = request.data.get("passing_score")
        try:
            passing_score = PASSING_SCORE_DEFAULT if passing_score_raw in (None, "") else Decimal(str(passing_score_raw))
        except Exception:
            return Response({"detail": "passing_score inválido"}, status=status.HTTP_400_BAD_REQUEST)

        computed = compute_promotions_for_year(academic_year=year, passing_score=passing_score)

        with transaction.atomic():
            # Persist decisions: store a human-readable string in Enrollment.final_status (legacy field)
            enrollments = Enrollment.objects.filter(id__in=list(computed.keys())).select_related("grade", "student", "student__user")

            updated = 0
            created = 0

            for e in enrollments:
                c = computed.get(e.id)
                if not c:
                    continue

                # Map decisions to legacy final_status text
                if c.decision == "PROMOTED":
                    final_status = "PROMOCIÓN PLENA"
                elif c.decision == "CONDITIONAL":
                    final_status = "PROMOCIÓN CONDICIONAL"
                elif c.decision == "REPEATED":
                    final_status = "REPROBÓ / REPITE"
                else:
                    final_status = c.decision

                # Grade 11 (último) => Graduado when promoted and no pending
                is_grade_11 = False
                if getattr(e.grade, "ordinal", None) is not None:
                    is_grade_11 = int(e.grade.ordinal) >= 13
                else:
                    is_grade_11 = str(e.grade.name).strip().lower() in {"11", "11°", "once", "undecimo", "undécimo"}

                if is_grade_11 and c.decision == "PROMOTED":
                    final_status = "GRADUADO"
                    e.status = "GRADUATED"

                e.final_status = final_status
                e.save(update_fields=["final_status", "status"] if e.status == "GRADUATED" else ["final_status"])

                if e.status == "GRADUATED":
                    try:
                        student = getattr(e, "student", None)
                        if student is not None and not Enrollment.objects.filter(student=student, status="ACTIVE").exists():
                            user = getattr(student, "user", None)
                            if user is not None and getattr(user, "is_active", True):
                                user.is_active = False
                                user.save(update_fields=["is_active"])
                    except Exception:
                        pass

                details = {
                    "passing_score": str(passing_score),
                    "subject_finals": {str(k): str(v) for k, v in c.subject_finals.items()},
                    "failed_subject_ids": c.failed_subject_ids,
                    "failed_area_ids": c.failed_area_ids,
                }

                snap_values = {
                    "decision": "GRADUATED" if final_status == "GRADUADO" else c.decision,
                    "failed_subjects_count": len(c.failed_subject_ids),
                    "failed_areas_count": len(c.failed_area_ids),
                    "failed_subjects_distinct_areas_count": c.failed_subjects_distinct_areas_count,
                    "details": details,
                }

                snap, was_created = EnrollmentPromotionSnapshot.objects.update_or_create(
                    enrollment=e,
                    defaults=snap_values,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

            # Close the year
            year.status = AcademicYear.STATUS_CLOSED
            year.save(update_fields=["status"])

        return Response(
            {
                "academic_year": {"id": year.id, "year": year.year, "status": year.status},
                "passing_score": str(passing_score),
                "snapshots": {"created": created, "updated": updated},
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="apply-promotions")
    def apply_promotions(self, request, pk=None):
        """Creates next-year enrollments based on promotion snapshots.

        Body: {"target_academic_year": <id>}
        """

        from students.models import ConditionalPromotionPlan, Enrollment
        from .models import EnrollmentPromotionSnapshot

        source_year: AcademicYear = self.get_object()
        target_year_id = request.data.get("target_academic_year")
        if not target_year_id:
            return Response({"detail": "target_academic_year es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        passing_score_raw = request.data.get("passing_score")
        try:
            passing_score = PASSING_SCORE_DEFAULT if passing_score_raw in (None, "") else Decimal(str(passing_score_raw))
        except Exception:
            return Response({"detail": "passing_score inválido"}, status=status.HTTP_400_BAD_REQUEST)

        enrollment_ids_raw = request.data.get("enrollment_ids")
        enrollment_ids = None
        if enrollment_ids_raw not in (None, ""):
            if not isinstance(enrollment_ids_raw, list):
                return Response({"detail": "enrollment_ids debe ser una lista"}, status=status.HTTP_400_BAD_REQUEST)
            try:
                enrollment_ids = [int(x) for x in enrollment_ids_raw]
            except Exception:
                return Response({"detail": "enrollment_ids inválido"}, status=status.HTTP_400_BAD_REQUEST)

        source_grade_id_raw = request.data.get("source_grade_id")
        source_grade_id = None
        if source_grade_id_raw not in (None, ""):
            try:
                source_grade_id = int(str(source_grade_id_raw))
            except Exception:
                return Response({"detail": "source_grade_id inválido"}, status=status.HTTP_400_BAD_REQUEST)

        exclude_repeated_raw = request.data.get("exclude_repeated")
        exclude_repeated = False
        if exclude_repeated_raw not in (None, ""):
            if isinstance(exclude_repeated_raw, bool):
                exclude_repeated = exclude_repeated_raw
            else:
                s = str(exclude_repeated_raw).strip().lower()
                exclude_repeated = s in {"1", "true", "t", "yes", "y", "si", "sí"}

        target_group_id_raw = request.data.get("target_group_id")
        target_group_id = None
        if target_group_id_raw not in (None, ""):
            try:
                target_group_id = int(str(target_group_id_raw))
            except Exception:
                return Response({"detail": "target_group_id inválido"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            target_year = AcademicYear.objects.get(id=int(target_year_id))
        except Exception:
            return Response({"detail": "target_academic_year inválido"}, status=status.HTTP_400_BAD_REQUEST)

        if source_year.id == target_year.id:
            return Response({"detail": "El año destino debe ser diferente"}, status=status.HTTP_400_BAD_REQUEST)

        if source_year.status != AcademicYear.STATUS_CLOSED:
            return Response(
                {"detail": "El año origen debe estar FINALIZADO (CLOSED) para aplicar promociones."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Determine the first period of the target year for PAP deadlines (optional)
        first_period = (
            Period.objects.filter(academic_year=target_year).order_by("start_date").only("id").first()
        )

        # Grade progression mapping using ordinal (with fallback inference from Grade.name)
        grades = list(Grade.objects.all().only("id", "ordinal", "name", "level_id"))
        ordinal_to_grade_id_by_level: dict[int, dict[int, int]] = {}
        ordinal_to_grade_id_global: dict[int, int] = {}

        for g in grades:
            ord_val = getattr(g, "ordinal", None)
            if ord_val is None:
                ord_val = guess_ordinal(getattr(g, "name", ""))
            if ord_val is None:
                continue
            o = int(ord_val)
            gid = int(g.id)
            lvl = int(getattr(g, "level_id", 0) or 0)
            ordinal_to_grade_id_by_level.setdefault(lvl, {}).setdefault(o, gid)
            ordinal_to_grade_id_global.setdefault(o, gid)

        def _grade_ordinal(current_grade: Grade) -> int | None:
            ord_val = getattr(current_grade, "ordinal", None)
            if ord_val is not None:
                return int(ord_val)
            guessed = guess_ordinal(getattr(current_grade, "name", ""))
            return int(guessed) if guessed is not None else None

        def next_grade_id(current_grade: Grade) -> int | None:
            ord_val = _grade_ordinal(current_grade)
            if ord_val is None:
                return None
            lvl = int(getattr(current_grade, "level_id", 0) or 0)
            per_level = ordinal_to_grade_id_by_level.get(lvl) or {}
            return per_level.get(ord_val + 1) or ordinal_to_grade_id_global.get(ord_val + 1)

        def is_last_grade(current_grade: Grade) -> bool:
            ord_val = _grade_ordinal(current_grade)
            if ord_val is not None:
                return ord_val >= 13
            return str(getattr(current_grade, "name", "")).strip().lower() in {"11", "11°", "once", "undecimo", "undécimo"}

        # When a year is CLOSED, AcademicYear.save retires ACTIVE enrollments.
        # Applying promotions must therefore consider both ACTIVE (open year) and RETIRED (closed year) enrollments.
        source_enrollments = Enrollment.objects.filter(academic_year=source_year, status__in=["ACTIVE", "RETIRED"])
        if source_grade_id is not None:
            source_enrollments = source_enrollments.filter(grade_id=source_grade_id)
        if enrollment_ids is not None:
            source_enrollments = source_enrollments.filter(id__in=enrollment_ids)

        source_enrollments = source_enrollments.select_related("student", "grade", "grade__level").order_by("id")

        # Preload groups in target year for auto-assignment / validation
        target_groups = list(
            Group.objects.filter(academic_year=target_year)
            .only("id", "name", "grade_id", "campus_id", "shift")
            .order_by("grade_id", "name", "id")
        )
        groups_by_key: dict[tuple[int, int | None], list[Group]] = {}
        for g in target_groups:
            key = (int(g.grade_id), int(g.campus_id) if getattr(g, "campus_id", None) is not None else None)
            groups_by_key.setdefault(key, []).append(g)

        selected_group_obj = None
        if target_group_id is not None:
            selected_group_obj = next((g for g in target_groups if int(g.id) == int(target_group_id)), None)
            if selected_group_obj is None:
                return Response(
                    {"detail": "El grupo seleccionado no existe en el año destino."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        snapshots = {
            s.enrollment_id: s
            for s in EnrollmentPromotionSnapshot.objects.filter(enrollment__in=source_enrollments)
        }

        # Fallback: compute decisions on-the-fly when snapshots are missing.
        # This supports environments where the year is CLOSED but snapshots were never generated.
        computed = {}
        try:
            computed = compute_promotions_for_year(academic_year=source_year, passing_score=passing_score)
        except Exception:
            computed = {}

        created = 0
        skipped_existing = 0
        skipped_graduated = 0
        skipped_missing_grade_ordinal = 0
        skipped_repeated = 0

        # Pre-check ambiguity when there are multiple groups for the target grade.
        # Avoid partial creations: fail fast with the available options.
        source_student_ids = list(source_enrollments.values_list("student_id", flat=True))
        existing_in_target = set(
            Enrollment.objects.filter(academic_year=target_year, student_id__in=source_student_ids).values_list(
                "student_id", flat=True
            )
        )
        ambiguous_groups = []
        if target_group_id is None:
            for e in source_enrollments:
                if int(e.student_id) in existing_in_target:
                    continue

                snap = snapshots.get(e.id)
                decision = None
                if snap is not None:
                    decision = snap.decision
                else:
                    c = computed.get(e.id)
                    if c is None:
                        continue
                    decision = c.decision

                if exclude_repeated and decision == "REPEATED":
                    continue
                if e.status == "GRADUATED" or decision == "GRADUATED":
                    continue
                if decision == "PROMOTED" and is_last_grade(e.grade):
                    continue

                if decision in {"PROMOTED", "CONDITIONAL"}:
                    ngid = next_grade_id(e.grade)
                    if ngid is None:
                        continue
                    target_grade_id = int(ngid)
                else:
                    target_grade_id = int(e.grade_id)

                campus_id = int(e.campus_id) if getattr(e, "campus_id", None) is not None else None
                candidates = groups_by_key.get((target_grade_id, campus_id), [])
                if not candidates and campus_id is not None:
                    candidates = groups_by_key.get((target_grade_id, None), [])

                if len(candidates) > 1:
                    ambiguous_groups.append(
                        {
                            "enrollment_id": int(e.id),
                            "target_grade_id": int(target_grade_id),
                            "campus_id": campus_id,
                            "groups": [
                                {
                                    "id": int(g.id),
                                    "name": g.name,
                                    "shift": getattr(g, "shift", None),
                                    "campus_id": int(g.campus_id) if getattr(g, "campus_id", None) is not None else None,
                                }
                                for g in candidates
                            ],
                        }
                    )

            if ambiguous_groups:
                return Response(
                    {
                        "detail": "Hay más de un grupo para el grado destino. Selecciona un grupo para aplicar promociones.",
                        "ambiguous_groups": ambiguous_groups,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        with transaction.atomic():
            for e in source_enrollments:
                snap = snapshots.get(e.id)
                decision = None
                details = {}
                if snap is not None:
                    decision = snap.decision
                    details = snap.details or {}
                else:
                    c = computed.get(e.id)
                    if c is None:
                        continue
                    decision = c.decision
                    details = {"failed_subject_ids": c.failed_subject_ids, "failed_area_ids": c.failed_area_ids}

                # Exclude REPEATED when requested
                if exclude_repeated and decision == "REPEATED":
                    skipped_repeated += 1
                    continue

                # Skip already graduated students
                if e.status == "GRADUATED" or decision == "GRADUATED":
                    skipped_graduated += 1
                    continue

                # Last grade promoted => treated as graduated (no next-year enrollment)
                if decision == "PROMOTED" and is_last_grade(e.grade):
                    skipped_graduated += 1
                    continue

                # Decide target grade
                if decision in {"PROMOTED", "CONDITIONAL"}:
                    ngid = next_grade_id(e.grade)
                    if ngid is None:
                        skipped_missing_grade_ordinal += 1
                        continue
                    target_grade_id = ngid
                else:
                    # REPEATED => stays in same grade
                    target_grade_id = e.grade_id

                # Determine target group
                assigned_group_id = None
                if selected_group_obj is not None:
                    # Ensure chosen group matches the target grade and (if present) campus
                    if int(selected_group_obj.grade_id) != int(target_grade_id):
                        return Response(
                            {"detail": "El grupo seleccionado no corresponde al grado destino."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    if getattr(e, "campus_id", None) is not None and getattr(selected_group_obj, "campus_id", None) is not None:
                        if int(selected_group_obj.campus_id) != int(e.campus_id):
                            return Response(
                                {"detail": "El grupo seleccionado no corresponde a la sede (campus) de la matrícula."},
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                    assigned_group_id = int(selected_group_obj.id)
                else:
                    campus_id = int(e.campus_id) if getattr(e, "campus_id", None) is not None else None
                    candidates = groups_by_key.get((int(target_grade_id), campus_id), [])
                    if not candidates and campus_id is not None:
                        candidates = groups_by_key.get((int(target_grade_id), None), [])
                    if len(candidates) == 1:
                        assigned_group_id = int(candidates[0].id)

                # Ensure uniqueness (student, academic_year)
                if Enrollment.objects.filter(student=e.student, academic_year=target_year).exists():
                    skipped_existing += 1
                    continue

                new_enrollment = Enrollment.objects.create(
                    student=e.student,
                    academic_year=target_year,
                    grade_id=target_grade_id,
                    group_id=assigned_group_id,
                    campus=e.campus,
                    status="ACTIVE",
                    origin_school=e.origin_school,
                    final_status="",
                    enrolled_at=None,
                )
                created += 1

                # Create conditional promotion plan (PAP placeholder)
                if decision == "CONDITIONAL":
                    ConditionalPromotionPlan.objects.create(
                        enrollment=new_enrollment,
                        source_enrollment=e,
                        due_period=first_period,
                        pending_subject_ids=list(details.get("failed_subject_ids", [])),
                        pending_area_ids=list(details.get("failed_area_ids", [])),
                        status=ConditionalPromotionPlan.STATUS_OPEN,
                        notes="Generado automáticamente por promoción condicional (SIEE).",
                    )

        return Response(
            {
                "source_academic_year": {"id": source_year.id, "year": source_year.year},
                "target_academic_year": {"id": target_year.id, "year": target_year.year},
                "created": created,
                "skipped_existing": skipped_existing,
                "skipped_graduated": skipped_graduated,
                "skipped_missing_grade_ordinal": skipped_missing_grade_ordinal,
                "skipped_repeated": skipped_repeated,
            },
            status=status.HTTP_200_OK,
        )


class PeriodViewSet(viewsets.ModelViewSet):
    queryset = Period.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [KampusModelPermissions]
    filterset_fields = ['academic_year']

    @action(detail=True, methods=["post"], url_path="close")
    def close(self, request, pk=None):
        period: Period = self.get_object()

        if period.is_closed:
            return Response(
                {"period": {"id": period.id, "name": period.name, "is_closed": True}},
                status=status.HTTP_200_OK,
            )

        academic_year = getattr(period, "academic_year", None)
        if academic_year is not None and getattr(academic_year, "status", None) == AcademicYear.STATUS_CLOSED:
            return Response(
                {"detail": "No se pueden cerrar periodos de un año lectivo finalizado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Block closure if there are OPEN PAP plans due on this period.
        from students.models import ConditionalPromotionPlan

        pending_qs = ConditionalPromotionPlan.objects.filter(due_period=period, status=ConditionalPromotionPlan.STATUS_OPEN)
        pending_count = pending_qs.count()
        if pending_count > 0:
            enrollment_ids = list(pending_qs.values_list("enrollment_id", flat=True)[:50])
            return Response(
                {
                    "detail": "No se puede cerrar el periodo: hay PAP pendientes.",
                    "pending_pap_count": pending_count,
                    "pending_enrollment_ids_sample": enrollment_ids,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        period.is_closed = True
        period.save(update_fields=["is_closed"])
        return Response({"period": {"id": period.id, "name": period.name, "is_closed": True}}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="reopen")
    def reopen(self, request, pk=None):
        period: Period = self.get_object()
        academic_year = getattr(period, "academic_year", None)
        if academic_year is not None and getattr(academic_year, "status", None) == AcademicYear.STATUS_CLOSED:
            return Response(
                {"detail": "No se pueden reabrir periodos de un año lectivo finalizado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        period.is_closed = False
        period.save(update_fields=["is_closed"])
        return Response({"period": {"id": period.id, "name": period.name, "is_closed": False}}, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        academic_year = getattr(instance, "academic_year", None)
        if academic_year is not None and getattr(academic_year, "status", None) == AcademicYear.STATUS_CLOSED:
            return Response(
                {"detail": "No se pueden eliminar periodos de un año lectivo finalizado."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)


class AcademicLevelViewSet(viewsets.ModelViewSet):
    queryset = AcademicLevel.objects.all()
    serializer_class = AcademicLevelSerializer
    permission_classes = [KampusModelPermissions]


class GradeViewSet(viewsets.ModelViewSet):
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [KampusModelPermissions]


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [KampusModelPermissions]
    filterset_fields = ['grade', 'academic_year', 'director']

    @action(
        detail=True,
        methods=["get"],
        url_path="academic-report",
        permission_classes=[IsAuthenticated],
    )
    def academic_report(self, request, pk=None):
        """Genera informe académico por periodos para un grupo.

        GET /api/groups/{id}/academic-report/?period=<period_id>
        """

        async_raw = (request.query_params.get("async") or "").strip().lower()
        async_requested = async_raw in {"1", "true", "yes"}

        def _to_int_or_none(value):
            if value is None:
                return None
            raw = str(value).strip()
            if raw == "":
                return None
            try:
                return int(raw)
            except Exception:
                return None

        period_id = _to_int_or_none(request.query_params.get("period"))
        if period_id is None:
            return Response({"detail": "period is required"}, status=status.HTTP_400_BAD_REQUEST)

        group: Group = self.get_object()

        user = getattr(request, "user", None)
        role = getattr(user, "role", None)
        if role not in {"SUPERADMIN", "ADMIN", "COORDINATOR"}:
            if role == "TEACHER":
                teacher_id = getattr(user, "id", None)
                is_director = group.director_id == teacher_id
                is_assigned = (
                    TeacherAssignment.objects.filter(teacher_id=teacher_id, group_id=group.id).exists()
                    if teacher_id is not None
                    else False
                )
                if not (is_director or is_assigned):
                    return Response(
                        {"detail": "No tienes permisos para ver este informe."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            else:
                return Response(
                    {"detail": "No tienes permisos para ver este informe."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        try:
            period = Period.objects.select_related("academic_year").get(id=period_id)
        except Period.DoesNotExist:
            return Response({"detail": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != group.academic_year_id:
            return Response(
                {"detail": "El periodo no corresponde al año lectivo del grupo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if async_requested:
            from reports.models import ReportJob
            from reports.serializers import ReportJobSerializer
            from reports.tasks import generate_report_job_pdf

            from django.utils import timezone
            from datetime import timedelta

            ttl_hours = int(getattr(settings, "REPORT_JOBS_TTL_HOURS", 24))
            expires_at = timezone.now() + timedelta(hours=ttl_hours)

            job = ReportJob.objects.create(
                created_by=request.user,
                report_type=ReportJob.ReportType.ACADEMIC_PERIOD_GROUP,
                params={"group_id": group.id, "period_id": period.id},
                expires_at=expires_at,
            )
            generate_report_job_pdf.delay(job.id)
            out = ReportJobSerializer(job, context={"request": request}).data
            return Response(out, status=status.HTTP_202_ACCEPTED)

        enrollments = (
            Enrollment.objects.select_related(
                "student",
                "student__user",
                "grade",
                "group",
                "group__director",
                "academic_year",
            )
            .filter(group_id=group.id, academic_year_id=period.academic_year_id, status="ACTIVE")
            .order_by("student__user__last_name", "student__user__first_name", "student__user__id")
        )

        if not enrollments.exists():
            return Response(
                {"detail": "No hay matrículas activas en este grupo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            level_type = None
            try:
                level_type = getattr(getattr(getattr(group.grade, "level", None), "level_type", None), "upper", lambda: None)()
            except Exception:
                level_type = None

            if level_type in {"PRESCHOOL", "PREESCOLAR"}:
                pdf_bytes = generate_preschool_academic_period_group_report_pdf(enrollments=enrollments, period=period)
            else:
                pdf_bytes = generate_academic_period_group_report_pdf(enrollments=enrollments, period=period)
            filename = f"informe-academico-grupo-{group.id}-period-{period.id}.pdf"
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'inline; filename="{filename}"'

            # Sync PDF is kept for backward compatibility, but jobs are preferred.
            response["Deprecation"] = "true"
            sunset = getattr(settings, "REPORTS_SYNC_SUNSET_DATE", "2026-06-30")
            response["Sunset"] = str(sunset)
            response["Link"] = '</api/reports/jobs/>; rel="alternate"'
            return response
        except Exception as e:
            payload = {"detail": "Error generating PDF", "error": str(e)}
            from django.conf import settings as _settings
            import traceback as _traceback

            if getattr(_settings, "DEBUG", False):
                payload["traceback"] = _traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(
        detail=True,
        methods=["get"],
        url_path="grade-report-sheet",
        permission_classes=[IsAuthenticated],
    )
    def grade_report_sheet(self, request, pk=None):
        """Genera planilla imprimible de notas (en blanco) para un grupo.

        GET /api/groups/{id}/grade-report-sheet/?format=pdf|html&period=<period_id>&subject=<text>&teacher=<text>&columns=<int>
        """

        def _to_int_or_none(value):
            if value is None:
                return None
            raw = str(value).strip()
            if raw == "":
                return None
            try:
                return int(raw)
            except Exception:
                return None

        def _upper(s: str) -> str:
            return (s or "").strip().upper()

        group: Group = self.get_object()
        async_raw = (request.query_params.get("async") or "").strip().lower()
        async_requested = async_raw in {"1", "true", "yes"}

        user = getattr(request, "user", None)
        role = getattr(user, "role", None)
        if role == "TEACHER":
            if not (group.director_id == getattr(user, "id", None) or TeacherAssignment.objects.filter(teacher_id=user.id, group_id=group.id).exists()):
                return Response({"detail": "No autorizado"}, status=status.HTTP_403_FORBIDDEN)

        fmt = (request.query_params.get("format") or "pdf").strip().lower()

        try:
            note_cols = int(request.query_params.get("columns") or "3")
        except Exception:
            note_cols = 3
        note_cols = max(1, min(note_cols, 12))

        period_id = _to_int_or_none(request.query_params.get("period"))

        subject_name = (request.query_params.get("subject") or "").strip()

        teacher_name = ""
        teacher_param = (request.query_params.get("teacher") or "").strip()
        if teacher_param:
            teacher_name = teacher_param
        else:
            try:
                if getattr(user, "is_authenticated", False) and getattr(user, "role", None) == "TEACHER":
                    teacher_name = _upper(user.get_full_name())
            except Exception:
                teacher_name = ""

        from django.template.loader import render_to_string

        from academic.reports import build_grade_report_sheet_context  # noqa: PLC0415

        ctx = build_grade_report_sheet_context(
            group=group,
            user=user,
            columns=note_cols,
            period_id=period_id,
            subject_name=subject_name,
            teacher_name=teacher_name,
        )

        html_string = render_to_string(
            "academic/reports/grade_report_sheet_pdf.html",
            ctx,
        )

        if fmt == "html":
            return HttpResponse(html_string, content_type="text/html; charset=utf-8")

        if async_requested:
            from datetime import timedelta  # noqa: PLC0415
            from reports.models import ReportJob  # noqa: PLC0415
            from reports.serializers import ReportJobSerializer  # noqa: PLC0415
            from reports.tasks import generate_report_job_pdf  # noqa: PLC0415

            ttl_hours = int(getattr(settings, "REPORT_JOBS_TTL_HOURS", 24))
            expires_at = timezone.now() + timedelta(hours=ttl_hours)

            job = ReportJob.objects.create(
                created_by=request.user,
                report_type=ReportJob.ReportType.GRADE_REPORT_SHEET,
                params={
                    "group_id": group.id,
                    "period_id": period_id,
                    "subject_name": subject_name,
                    "teacher_name": teacher_name,
                    "columns": note_cols,
                },
                expires_at=expires_at,
            )
            generate_report_job_pdf.delay(job.id)
            out = ReportJobSerializer(job, context={"request": request}).data
            return Response(out, status=status.HTTP_202_ACCEPTED)

        from reports.weasyprint_utils import WeasyPrintUnavailableError, render_pdf_bytes_from_html  # noqa: PLC0415

        try:
            group_label = str(ctx.get("group_label") or "grupo")
            filename = f"planilla_notas_{group_label}.pdf".replace(" ", "_")

            pdf_bytes = render_pdf_bytes_from_html(html=html_string, base_url=str(settings.BASE_DIR))

            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'inline; filename="{filename}"'
            response["Deprecation"] = "true"
            sunset = getattr(settings, "REPORTS_SYNC_SUNSET_DATE", "2026-06-30")
            response["Sunset"] = str(sunset)
            response["Link"] = '</api/reports/jobs/>; rel="alternate"'
            return response
        except WeasyPrintUnavailableError as e:
            from django.conf import settings as _settings

            payload = {"detail": str(e)}
            if getattr(_settings, "DEBUG", False):
                import traceback  # noqa: PLC0415

                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as e:
            from django.conf import settings as _settings

            payload = {"detail": "Error generando PDF", "error": str(e)}
            if getattr(_settings, "DEBUG", False):
                import traceback  # noqa: PLC0415

                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(
        detail=True,
        methods=["get"],
        url_path="family-directory-report",
        permission_classes=[IsAuthenticated],
    )
    def family_directory_report(self, request, pk=None):
        """Genera directorio de padres de familia (XLSX) por grupo.

        GET /api/groups/{id}/family-directory-report/
        """
        group: Group = self.get_object()

        user = getattr(request, "user", None)
        role = getattr(user, "role", None)
        if role not in {"SUPERADMIN", "ADMIN", "COORDINATOR"}:
            if role == "TEACHER":
                teacher_id = getattr(user, "id", None)
                is_director = group.director_id == teacher_id
                is_assigned = (
                    TeacherAssignment.objects.filter(teacher_id=teacher_id, group_id=group.id).exists()
                    if teacher_id is not None
                    else False
                )
                if not (is_director or is_assigned):
                    return Response(
                        {"detail": "No tienes permisos para ver este informe."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            else:
                return Response(
                    {"detail": "No tienes permisos para ver este informe."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        enrollments = (
            Enrollment.objects.select_related("student", "student__user")
            .filter(group_id=group.id, academic_year_id=group.academic_year_id, status="ACTIVE")
            .order_by("student__user__last_name", "student__user__first_name", "student__user__id")
        )

        student_ids = [enrollment.student_id for enrollment in enrollments]
        family_members = (
            FamilyMember.objects.filter(student_id__in=student_ids)
            .order_by("student_id", "-is_main_guardian", "id")
        )

        guardian_by_student_id = {}
        for fm in family_members:
            if fm.student_id not in guardian_by_student_id:
                guardian_by_student_id[fm.student_id] = fm

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Directorio"

        headers = [
            "Nombre del Estudiante",
            "Nombre del acudiente",
            "identificación",
            "telefono",
            "dirección",
            "parentezco",
        ]
        worksheet.append(headers)

        for enrollment in enrollments:
            student = enrollment.student
            student_user = student.user
            student_name = ""
            if student_user is not None:
                student_name = (student_user.get_full_name() or "").strip()
            if not student_name:
                student_name = str(student)

            guardian = guardian_by_student_id.get(student.pk)
            worksheet.append(
                [
                    student_name,
                    (guardian.full_name if guardian else "") or "",
                    (guardian.document_number if guardian else "") or "",
                    (guardian.phone if guardian else "") or "",
                    (guardian.address if guardian else "") or "",
                    (guardian.relationship if guardian else "") or "",
                ]
            )

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for cell in worksheet[get_column_letter(col_idx)]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)

        out = io.BytesIO()
        workbook.save(out)
        out.seek(0)

        filename = f"directorio_padres_familia_grupo_{group.id}.xlsx"
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'], url_path='copy_from_year')
    def copy_from_year(self, request):
        source_year_id = request.data.get('source_year_id')
        target_year_id = request.data.get('target_year_id')

        if not source_year_id or not target_year_id:
            return Response(
                {"error": "source_year_id and target_year_id are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            source_year_id = int(source_year_id)
            target_year_id = int(target_year_id)
        except (TypeError, ValueError):
            return Response(
                {"error": "source_year_id and target_year_id must be integers"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if source_year_id == target_year_id:
            return Response(
                {"error": "source_year_id and target_year_id must be different"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        source_groups = Group.objects.filter(academic_year_id=source_year_id)
        if not source_groups.exists():
            return Response(
                {"error": "No groups found in source year"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if Group.objects.filter(academic_year_id=target_year_id).exists():
            return Response(
                {
                    "error": "El año destino ya tiene grupos configurados. Elimina o ajusta antes de importar."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        with transaction.atomic():
            for group in source_groups:
                Group.objects.create(
                    academic_year_id=target_year_id,
                    name=group.name,
                    grade_id=group.grade_id,
                    campus_id=group.campus_id,
                    director_id=group.director_id,
                    shift=group.shift,
                    classroom=group.classroom,
                    capacity=group.capacity,
                )
                created_count += 1

        return Response({"message": f"Se copiaron {created_count} grupos correctamente"})


class AreaViewSet(viewsets.ModelViewSet):
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    permission_classes = [KampusModelPermissions]


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [KampusModelPermissions]


class AcademicLoadViewSet(viewsets.ModelViewSet):
    queryset = AcademicLoad.objects.all()
    serializer_class = AcademicLoadSerializer
    permission_classes = [KampusModelPermissions]


class TeacherAssignmentViewSet(viewsets.ModelViewSet):
    queryset = TeacherAssignment.objects.all()
    serializer_class = TeacherAssignmentSerializer
    permission_classes = [KampusModelPermissions]

    def _deny_write_if_teacher(self, request):
        user = getattr(request, "user", None)
        if user is not None and getattr(user, "role", None) == "TEACHER":
            return Response({"detail": "No tienes permisos para modificar asignaciones."}, status=status.HTTP_403_FORBIDDEN)
        return None

    def create(self, request, *args, **kwargs):
        denied = self._deny_write_if_teacher(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._deny_write_if_teacher(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = self._deny_write_if_teacher(request)
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._deny_write_if_teacher(request)
        if denied is not None:
            return denied
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.select_related(
            "teacher",
            "academic_year",
            "group__grade",
            "academic_load__subject",
            "academic_load__grade",
        )

    @action(detail=False, methods=["get"], url_path="me", permission_classes=[IsAuthenticated])
    def me(self, request):
        """Return only the authenticated teacher's assignments."""
        user = getattr(request, "user", None)
        if user is None or getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)

        qs = (
            self.get_queryset()
            .filter(teacher=user)
            .order_by("-academic_year__year", "group__grade__name", "group__name")
        )

        academic_year = request.query_params.get("academic_year")
        if academic_year:
            qs = qs.filter(academic_year_id=academic_year)

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="assign_grade_plan")
    def assign_grade_plan(self, request):
        """Bulk-assign all AcademicLoads for the group's grade to a teacher in a given year.

        Intended for PRIMARY teachers where a single teacher often covers the full study plan.
        Skips loads already assigned to another teacher and reports conflicts.
        """

        denied = self._deny_write_if_teacher(request)
        if denied is not None:
            return denied

        from django.contrib.auth import get_user_model  # noqa: PLC0415
        from teachers.models import Teacher as TeacherProfile  # noqa: PLC0415

        teacher_id = request.data.get("teacher")
        group_id = request.data.get("group")
        academic_year_id = request.data.get("academic_year")
        force = bool(request.data.get("force", False))

        if not teacher_id or not group_id or not academic_year_id:
            return Response(
                {"detail": "teacher, group, academic_year son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        User = get_user_model()
        teacher_user = User.objects.filter(id=teacher_id, role="TEACHER").first()
        if teacher_user is None:
            return Response({"detail": "Docente inválido."}, status=status.HTTP_400_BAD_REQUEST)

        group = Group.objects.filter(id=group_id).select_related("grade").first()
        if group is None:
            return Response({"detail": "Grupo inválido."}, status=status.HTTP_400_BAD_REQUEST)

        target_year = AcademicYear.objects.filter(id=academic_year_id).first()
        if target_year is None:
            return Response({"detail": "Año académico inválido."}, status=status.HTTP_400_BAD_REQUEST)

        if group.academic_year_id != target_year.id:
            return Response(
                {"detail": "El grupo no pertenece al año académico seleccionado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not force:
            teacher_profile = TeacherProfile.objects.filter(user_id=teacher_user.id).first()
            if teacher_profile is None or teacher_profile.teaching_level != "PRIMARY":
                return Response(
                    {
                        "detail": "Esta acción está habilitada solo para docentes de Primaria (use force=true para omitir)."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        loads = list(
            AcademicLoad.objects.filter(grade_id=group.grade_id)
            .select_related("subject")
            .order_by("subject__name")
        )

        if not loads:
            return Response(
                {"detail": "No hay cargas académicas definidas para este grado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        existing = (
            TeacherAssignment.objects.filter(group_id=group.id, academic_year_id=target_year.id)
            .values_list("academic_load_id", "teacher_id")
        )
        existing_by_load: dict[int, int] = {int(load_id): int(tid) for load_id, tid in existing}

        to_create: list[TeacherAssignment] = []
        skipped_existing = 0
        skipped_taken = 0
        conflicts = []

        conflict_teacher_ids = set(existing_by_load.values())
        conflict_teachers = {
            u.id: u.get_full_name() or u.username
            for u in User.objects.filter(id__in=conflict_teacher_ids)
        }

        for load in loads:
            existing_teacher_id = existing_by_load.get(load.id)
            if existing_teacher_id is not None:
                if existing_teacher_id == teacher_user.id:
                    skipped_existing += 1
                else:
                    skipped_taken += 1
                    conflicts.append(
                        {
                            "academic_load_id": load.id,
                            "subject": load.subject.name,
                            "assigned_teacher_id": existing_teacher_id,
                            "assigned_teacher_name": conflict_teachers.get(existing_teacher_id),
                        }
                    )
                continue

            to_create.append(
                TeacherAssignment(
                    teacher=teacher_user,
                    academic_load=load,
                    group=group,
                    academic_year=target_year,
                )
            )

        with transaction.atomic():
            if to_create:
                TeacherAssignment.objects.bulk_create(to_create)

        return Response(
            {
                "created": len(to_create),
                "skipped_existing": skipped_existing,
                "skipped_taken": skipped_taken,
                "conflicts": conflicts,
            },
            status=status.HTTP_200_OK,
        )


class EvaluationScaleViewSet(viewsets.ModelViewSet):
    queryset = EvaluationScale.objects.all()
    serializer_class = EvaluationScaleSerializer
    permission_classes = [KampusModelPermissions]

    @action(detail=False, methods=['post'])
    def copy_from_year(self, request):
        source_year_id = request.data.get('source_year_id')
        target_year_id = request.data.get('target_year_id')
        
        if not source_year_id or not target_year_id:
            return Response(
                {"error": "source_year_id and target_year_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        source_scales = EvaluationScale.objects.filter(academic_year_id=source_year_id)
        
        if not source_scales.exists():
            return Response(
                {"error": "No scales found in source year"}, 
                status=status.HTTP_404_NOT_FOUND
            )
            
        created_count = 0
        for scale in source_scales:
            # Check if similar scale exists in target year to avoid duplicates
            if not EvaluationScale.objects.filter(
                academic_year_id=target_year_id, 
                name=scale.name
            ).exists():
                EvaluationScale.objects.create(
                    academic_year_id=target_year_id,
                    name=scale.name,
                    min_score=scale.min_score,
                    max_score=scale.max_score,
                    description=scale.description,
                    scale_type=scale.scale_type
                )
                created_count += 1
                
        return Response({"message": f"Se copiaron {created_count} escalas correctamente"})


class EvaluationComponentViewSet(viewsets.ModelViewSet):
    queryset = EvaluationComponent.objects.all()
    serializer_class = EvaluationComponentSerializer
    permission_classes = [KampusModelPermissions]


class AssessmentViewSet(viewsets.ModelViewSet):
    queryset = Assessment.objects.all()
    serializer_class = AssessmentSerializer
    permission_classes = [KampusModelPermissions]


class StudentGradeViewSet(viewsets.ModelViewSet):
    queryset = StudentGrade.objects.all()
    serializer_class = StudentGradeSerializer
    permission_classes = [KampusModelPermissions]


class AchievementDefinitionViewSet(viewsets.ModelViewSet):
    queryset = AchievementDefinition.objects.all()
    serializer_class = AchievementDefinitionSerializer
    permission_classes = [KampusModelPermissions]
    filterset_fields = ['area', 'subject', 'is_active', 'dimension']

    def get_queryset(self):
        qs = super().get_queryset()
        user = getattr(self.request, "user", None)
        role = getattr(user, "role", None)

        # Requirement: teachers should only see bank achievements they created.
        # Admin-like roles can see everything.
        if role == "TEACHER":
            # Backward compatibility: backups created before the `created_by` field existed
            # will restore definitions with created_by=NULL. In that case, we allow a teacher
            # to see only the legacy definitions that match their teaching assignments.
            from django.db.models import Exists, OuterRef, Q

            # 1) Direct match by academic_load (preferred when present)
            assigned_load = TeacherAssignment.objects.filter(
                teacher=user,
                academic_load_id=OuterRef("academic_load_id"),
            )

            # 2) Fallback match by (subject, grade) when academic_load is not set
            assigned_subject_grade = TeacherAssignment.objects.filter(
                teacher=user,
                academic_load__subject_id=OuterRef("subject_id"),
                academic_load__grade_id=OuterRef("grade_id"),
            )

            qs = qs.annotate(
                _legacy_visible_by_load=Exists(assigned_load),
                _legacy_visible_by_subject_grade=Exists(assigned_subject_grade),
            )

            return qs.filter(
                Q(created_by=user)
                | (
                    Q(created_by__isnull=True)
                    & (Q(_legacy_visible_by_load=True) | Q(_legacy_visible_by_subject_grade=True))
                )
            )
        return qs

    def perform_create(self, serializer):
        user = getattr(self.request, "user", None)
        serializer.save(created_by=user)

    def get_permissions(self):
        # The achievement bank is a teacher workflow in the UI, but teachers may not
        # have Django model permissions assigned. We gate by role instead of model perms.
        if getattr(self, "action", None) in {"list", "retrieve", "create", "update", "partial_update", "destroy"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def _ensure_can_manage_definitions(self, request):
        role = getattr(getattr(request, 'user', None), 'role', None)
        if role in {'TEACHER', 'COORDINATOR', 'ADMIN', 'SUPERADMIN'}:
            return None
        return Response({"detail": "No tienes permisos para gestionar el banco de logros."}, status=status.HTTP_403_FORBIDDEN)

    def list(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_definitions(request)
        if denied is not None:
            return denied
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_definitions(request)
        if denied is not None:
            return denied
        return super().retrieve(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_definitions(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_definitions(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_definitions(request)
        if denied is not None:
            return denied
        return super().destroy(request, *args, **kwargs)

    @action(
        detail=False,
        methods=['post'],
        url_path='improve-wording',
        permission_classes=[IsAuthenticated],
        throttle_classes=[AcademicAIUserRateThrottle],
    )
    def improve_wording(self, request):
        """
        Mejora la redacción de un texto usando IA.
        Body: { "text": "..." }
        """
        role = getattr(getattr(request, 'user', None), 'role', None)
        if role not in {'TEACHER', 'COORDINATOR', 'ADMIN', 'SUPERADMIN'}:
            return Response({"detail": "No tienes permisos para usar esta función."}, status=status.HTTP_403_FORBIDDEN)

        text = request.data.get('text')
        if not text:
            return Response({"error": "Text is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            ai_service = AIService()
            improved_text = ai_service.improve_text(text)
            return Response({"improved_text": improved_text})
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AchievementViewSet(viewsets.ModelViewSet):
    queryset = Achievement.objects.all()
    serializer_class = AchievementSerializer
    permission_classes = [KampusModelPermissions]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subject', 'period', 'group']

    def get_permissions(self):
        # Creating/editing achievements is a teacher workflow in the UI, but teachers may not
        # have Django model add/change permissions assigned. We gate it by role instead.
        if getattr(self, "action", None) in {"create", "update", "partial_update", "destroy"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def _ensure_can_manage_achievements(self, request):
        role = getattr(getattr(request, 'user', None), 'role', None)
        if role in {'TEACHER', 'COORDINATOR', 'ADMIN', 'SUPERADMIN'}:
            return None
        return Response({"detail": "No tienes permisos para gestionar logros."}, status=status.HTTP_403_FORBIDDEN)

    def _teacher_can_edit_planning(self, user, period: Period) -> bool:
        effective_deadline = period.planning_edit_until or _period_end_of_day(period)
        if effective_deadline is None:
            return True
        if timezone.now() <= effective_deadline:
            return True
        return EditGrant.objects.filter(
            granted_to=user,
            scope=EditRequest.SCOPE_PLANNING,
            period_id=period.id,
            valid_until__gte=timezone.now(),
        ).exists()

    def _get_period_for_create(self, request):
        period_id = request.data.get("period")
        if not period_id:
            return None
        return Period.objects.filter(id=period_id).first()

    def _ensure_can_use_ai(self, request):
        role = getattr(getattr(request, 'user', None), 'role', None)
        if role in {'TEACHER', 'COORDINATOR', 'ADMIN', 'SUPERADMIN'}:
            return None
        return Response({"detail": "No tienes permisos para usar esta función."}, status=status.HTTP_403_FORBIDDEN)

    def create(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_achievements(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if user is not None and getattr(user, "role", None) == "TEACHER":
            period = self._get_period_for_create(request)
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_achievements(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if user is not None and getattr(user, "role", None) == "TEACHER":
            instance = self.get_object()
            period = getattr(instance, "period", None)
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_achievements(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if user is not None and getattr(user, "role", None) == "TEACHER":
            instance = self.get_object()
            period = getattr(instance, "period", None)
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        return super().destroy(request, *args, **kwargs)

    @action(
        detail=False,
        methods=['post'],
        url_path='generate-indicators',
        permission_classes=[IsAuthenticated],
        throttle_classes=[AcademicAIUserRateThrottle],
    )
    def generate_indicators(self, request):
        """
        Genera sugerencias de indicadores usando IA.
        Body: { "description": "..." }
        """
        denied = self._ensure_can_use_ai(request)
        if denied is not None:
            return denied

        description = request.data.get('description')
        if not description:
            return Response({"detail": "Description is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            ai_service = AIService()
            indicators = ai_service.generate_indicators(description)
            return Response(indicators)
        except AIConfigError as e:
            return Response(
                {"detail": str(e), "code": "AI_NOT_CONFIGURED"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except AIParseError:
            # The provider responded, but it wasn't usable JSON. Encourage retry.
            return Response(
                {
                    "detail": "La IA devolvió una respuesta inválida. Intenta nuevamente.",
                    "code": "AI_INVALID_RESPONSE",
                },
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except AIProviderError:
            return Response(
                {
                    "detail": "No se pudo generar con IA en este momento. Intenta nuevamente.",
                    "code": "AI_PROVIDER_ERROR",
                },
                status=status.HTTP_502_BAD_GATEWAY,
            )

    @action(detail=True, methods=['post'], url_path='create-indicators', permission_classes=[IsAuthenticated])
    def create_indicators(self, request, pk=None):
        """
        Crea indicadores masivamente para un logro existente.
        Body: { "indicators": [ {"level": "LOW", "description": "..."}, ... ] }
        """
        achievement = self.get_object()

        denied = self._ensure_can_use_ai(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if user is not None and getattr(user, "role", None) == "TEACHER":
            period = getattr(achievement, "period", None)
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )

        indicators_data = request.data.get('indicators', [])
        
        created_indicators = []
        errors = []
        
        for ind_data in indicators_data:
            serializer = PerformanceIndicatorSerializer(data={
                'achievement': achievement.id,
                'level': ind_data.get('level'),
                'description': ind_data.get('description')
            })
            if serializer.is_valid():
                serializer.save()
                created_indicators.append(serializer.data)
            else:
                errors.append(serializer.errors)
        
        if errors:
             return Response({"created": created_indicators, "errors": errors}, status=status.HTTP_207_MULTI_STATUS)

        return Response(created_indicators, status=status.HTTP_201_CREATED)


class PerformanceIndicatorViewSet(viewsets.ModelViewSet):
    queryset = PerformanceIndicator.objects.all()
    serializer_class = PerformanceIndicatorSerializer
    permission_classes = [KampusModelPermissions]


class DimensionViewSet(viewsets.ModelViewSet):
    queryset = Dimension.objects.all()
    serializer_class = DimensionSerializer
    permission_classes = [KampusModelPermissions]
    filterset_fields = ["academic_year", "is_active"]

    @action(detail=False, methods=['post'], url_path='copy_from_year')
    def copy_from_year(self, request):
        source_year_id = request.data.get('source_year_id')
        target_year_id = request.data.get('target_year_id')

        if not source_year_id or not target_year_id:
            return Response(
                {"error": "source_year_id and target_year_id are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            source_year_id = int(source_year_id)
            target_year_id = int(target_year_id)
        except (TypeError, ValueError):
            return Response(
                {"error": "source_year_id and target_year_id must be integers"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if source_year_id == target_year_id:
            return Response(
                {"error": "source_year_id and target_year_id must be different"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        source_dims = Dimension.objects.filter(academic_year_id=source_year_id)
        if not source_dims.exists():
            return Response(
                {"error": "No dimensions found in source year"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if Dimension.objects.filter(academic_year_id=target_year_id).exists():
            return Response(
                {
                    "error": "El año destino ya tiene dimensiones configuradas. Elimina o ajusta antes de copiar."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        with transaction.atomic():
            for dim in source_dims:
                Dimension.objects.create(
                    academic_year_id=target_year_id,
                    name=dim.name,
                    description=dim.description,
                    percentage=dim.percentage,
                    is_active=dim.is_active,
                )
                created_count += 1

        return Response({"message": f"Se copiaron {created_count} dimensiones correctamente"})


class PeriodTopicViewSet(viewsets.ModelViewSet):
    queryset = PeriodTopic.objects.all()
    serializer_class = PeriodTopicSerializer
    permission_classes = [KampusModelPermissions]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["period", "academic_load", "is_active", "source"]
    search_fields = ["title", "description", "academic_load__subject__name", "academic_load__grade__name"]

    def get_permissions(self):
        if getattr(self, "action", None) in {"list", "retrieve", "create", "update", "partial_update", "destroy", "for_me"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def _ensure_can_manage_topics(self, request):
        role = getattr(getattr(request, "user", None), "role", None)
        if role in {"TEACHER", "COORDINATOR", "ADMIN", "SUPERADMIN"}:
            return None
        return Response({"detail": "No tienes permisos para gestionar temáticas."}, status=status.HTTP_403_FORBIDDEN)

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            "period",
            "period__academic_year",
            "academic_load",
            "academic_load__subject",
            "academic_load__grade",
            "created_by",
        )
        user = getattr(self.request, "user", None)
        role = getattr(user, "role", None)
        if role == "TEACHER":
            allowed_loads = TeacherAssignment.objects.filter(teacher=user).values_list("academic_load_id", flat=True)
            qs = qs.filter(academic_load_id__in=allowed_loads)

        subject_id = self.request.query_params.get("subject")
        if subject_id:
            qs = qs.filter(academic_load__subject_id=subject_id)

        grade_id = self.request.query_params.get("grade")
        if grade_id:
            qs = qs.filter(academic_load__grade_id=grade_id)

        academic_year_id = self.request.query_params.get("academic_year")
        if academic_year_id:
            qs = qs.filter(period__academic_year_id=academic_year_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def list(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied
        return super().retrieve(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="for-me", permission_classes=[IsAuthenticated])
    def for_me(self, request):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Esta vista aplica solo para docentes."}, status=status.HTTP_400_BAD_REQUEST)

        qs = self.filter_queryset(self.get_queryset()).filter(is_active=True)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="import-template", permission_classes=[IsAuthenticated])
    def import_template(self, request):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied

        from openpyxl import Workbook  # noqa: PLC0415

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tematicas"
        worksheet.append(TOPIC_IMPORT_COLUMNS)
        worksheet.append(["2026", "Primer Periodo", "10", "Biología", "1", "La célula", "Conceptos base y observación microscópica"])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_tematicas_periodo.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="validate-import-file", permission_classes=[IsAuthenticated])
    def validate_import_file(self, request):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied

        upload = request.FILES.get("file")
        if upload is None:
            return Response({"detail": "Debes adjuntar un archivo Excel o CSV."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            headers, data_rows = _read_topic_import_rows(upload)
        except Exception:
            return Response({"detail": "No se pudo leer el archivo cargado. Usa .xlsx, .xls o .csv."}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = set(TOPIC_IMPORT_COLUMNS)
        if not headers or not required_columns.issubset(set(headers)):
            return Response(
                {"detail": "El archivo no contiene las columnas requeridas.", "required_columns": sorted(required_columns)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        corrections = _parse_topic_import_corrections(request.data.get("corrections"))
        period_map, loads_by_grade_key = _build_topic_import_context()
        validation = _build_topic_import_validation(data_rows, period_map, loads_by_grade_key, corrections=corrections)
        validation.pop("resolved_rows", None)
        return Response(validation, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="import-file", permission_classes=[IsAuthenticated])
    def import_file(self, request):
        denied = self._ensure_can_manage_topics(request)
        if denied is not None:
            return denied

        upload = request.FILES.get("file")
        if upload is None:
            return Response({"detail": "Debes adjuntar un archivo Excel o CSV."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            headers, data_rows = _read_topic_import_rows(upload)
        except Exception:
            return Response({"detail": "No se pudo leer el archivo cargado. Usa .xlsx, .xls o .csv."}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = set(TOPIC_IMPORT_COLUMNS)
        if not headers or not required_columns.issubset(set(headers)):
            return Response(
                {"detail": "El archivo no contiene las columnas requeridas.", "required_columns": sorted(required_columns)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        corrections = _parse_topic_import_corrections(request.data.get("corrections"))
        period_map, loads_by_grade_key = _build_topic_import_context()
        validation = _build_topic_import_validation(data_rows, period_map, loads_by_grade_key, corrections=corrections)

        created = 0
        updated = 0
        errors: list[str] = [f"Fila {row['row_number']}: {row['message']}" for row in validation["rows"] if row["status"] != "ready"]

        with transaction.atomic():
            for row_result, period, resolved in validation["resolved_rows"]:
                academic_load = resolved["academic_load"]
                sequence_order = resolved["sequence_order"]
                title = resolved["title"]
                description = resolved["description"]

                existing = PeriodTopic.objects.filter(
                    period=period,
                    academic_load=academic_load,
                    title__iexact=title,
                ).first()

                if existing is None:
                    PeriodTopic.objects.create(
                        period=period,
                        academic_load=academic_load,
                        title=title,
                        description=description,
                        sequence_order=sequence_order,
                        source=PeriodTopic.SOURCE_IMPORT,
                        is_active=True,
                        created_by=request.user,
                    )
                    created += 1
                else:
                    existing.title = title
                    existing.description = description
                    existing.sequence_order = sequence_order
                    existing.source = PeriodTopic.SOURCE_IMPORT
                    existing.is_active = True
                    if existing.created_by_id is None:
                        existing.created_by = request.user
                    existing.save(update_fields=["title", "description", "sequence_order", "source", "is_active", "created_by", "updated_at"])
                    updated += 1

        return Response({"created": created, "updated": updated, "errors": errors}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="import-csv", permission_classes=[IsAuthenticated])
    def import_csv(self, request):
        return self.import_file(request)


class ClassPlanViewSet(viewsets.ModelViewSet):
    queryset = ClassPlan.objects.all()
    serializer_class = ClassPlanSerializer
    permission_classes = [KampusModelPermissions]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["period", "teacher_assignment", "status", "topic"]
    search_fields = ["title", "learning_result", "topic__title"]

    def get_permissions(self):
        if getattr(self, "action", None) in {"list", "retrieve", "create", "update", "partial_update", "destroy", "my_plans", "my_summary", "generate_draft", "generate_draft_stream", "generate_section", "export_pdf"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def _log_class_plan_event(self, request, *, event_type, plan=None, status_code=200, metadata=None):
        metadata = metadata or {}
        if plan is not None:
            metadata = {
                **metadata,
                "title": getattr(plan, "title", ""),
                "period_id": getattr(plan, "period_id", None),
                "academic_year_id": getattr(getattr(plan, "period", None), "academic_year_id", None),
                "teacher_assignment_id": getattr(plan, "teacher_assignment_id", None),
                "status": getattr(plan, "status", None),
            }
        log_event(
            request,
            event_type=event_type,
            object_type="class_plan",
            object_id=getattr(plan, "id", "") if plan is not None else "",
            status_code=status_code,
            metadata=metadata,
        )

    def _ensure_can_manage_class_plans(self, request):
        role = getattr(getattr(request, "user", None), "role", None)
        if role in {"TEACHER", "COORDINATOR", "ADMIN", "SUPERADMIN"}:
            return None
        return Response({"detail": "No tienes permisos para gestionar planes de clase."}, status=status.HTTP_403_FORBIDDEN)

    def _teacher_can_edit_planning(self, user, period: Period) -> bool:
        effective_deadline = period.planning_edit_until or _period_end_of_day(period)
        if effective_deadline is None:
            return True
        if timezone.now() <= effective_deadline:
            return True
        return EditGrant.objects.filter(
            granted_to=user,
            scope=EditRequest.SCOPE_PLANNING,
            period_id=period.id,
            valid_until__gte=timezone.now(),
        ).exists()

    def _resolve_generate_draft_context(self, request):
        teacher_assignment_id = request.data.get("teacher_assignment")
        topic_id = request.data.get("topic")
        duration_minutes = request.data.get("duration_minutes") or 55
        title = str(request.data.get("title") or "").strip()

        assignment = None
        topic = None

        if teacher_assignment_id:
            assignment = TeacherAssignment.objects.select_related(
                "teacher",
                "group",
                "group__grade",
                "academic_load",
                "academic_load__subject",
            ).filter(id=teacher_assignment_id).first()
            if assignment is None:
                return None, None, Response({"detail": "Asignación docente no encontrada."}, status=status.HTTP_400_BAD_REQUEST)

            user = getattr(request, "user", None)
            if getattr(user, "role", None) == "TEACHER" and assignment.teacher_id != getattr(user, "id", None):
                return None, None, Response({"detail": "No puedes generar un plan para otra asignación docente."}, status=status.HTTP_403_FORBIDDEN)

        if topic_id:
            topic = PeriodTopic.objects.select_related(
                "period",
                "academic_load",
                "academic_load__subject",
                "academic_load__grade",
            ).filter(id=topic_id).first()
            if topic is None:
                return None, None, Response({"detail": "Temática no encontrada."}, status=status.HTTP_400_BAD_REQUEST)

        if assignment is not None and topic is not None and assignment.academic_load_id != topic.academic_load_id:
            return None, None, Response(
                {"detail": "La temática no corresponde a la asignación docente seleccionada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            duration_minutes = int(duration_minutes)
        except Exception:
            return None, None, Response({"detail": "duration_minutes inválido."}, status=status.HTTP_400_BAD_REQUEST)

        context = {
            "topic_title": getattr(topic, "title", "") or title,
            "topic_description": getattr(topic, "description", ""),
            "subject_name": getattr(getattr(getattr(assignment, "academic_load", None), "subject", None), "name", "")
            or getattr(getattr(getattr(topic, "academic_load", None), "subject", None), "name", ""),
            "grade_name": getattr(getattr(getattr(assignment, "group", None), "grade", None), "name", "")
            or getattr(getattr(getattr(topic, "academic_load", None), "grade", None), "name", ""),
            "group_name": getattr(getattr(assignment, "group", None), "name", ""),
            "teacher_name": getattr(getattr(assignment, "teacher", None), "get_full_name", lambda: "")(),
            "period_name": getattr(getattr(topic, "period", None), "name", "") or request.data.get("period_name") or "",
            "duration_minutes": duration_minutes,
            "title_hint": title,
        }

        metadata = {
            "teacher_assignment_id": teacher_assignment_id,
            "topic_id": topic_id,
            "title": title,
        }

        return context, metadata, None

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            "period",
            "period__academic_year",
            "teacher_assignment",
            "teacher_assignment__teacher",
            "teacher_assignment__group",
            "teacher_assignment__group__grade",
            "teacher_assignment__academic_load",
            "teacher_assignment__academic_load__subject",
            "topic",
            "created_by",
            "updated_by",
        )
        user = getattr(self.request, "user", None)
        role = getattr(user, "role", None)
        if role == "TEACHER":
            qs = qs.filter(teacher_assignment__teacher=user)

        subject_id = self.request.query_params.get("subject")
        if subject_id:
            qs = qs.filter(teacher_assignment__academic_load__subject_id=subject_id)

        group_id = self.request.query_params.get("group")
        if group_id:
            qs = qs.filter(teacher_assignment__group_id=group_id)

        academic_year_id = self.request.query_params.get("academic_year")
        if academic_year_id:
            qs = qs.filter(period__academic_year_id=academic_year_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def list(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied
        return super().retrieve(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if getattr(user, "role", None) == "TEACHER":
            period_id = request.data.get("period")
            period = Period.objects.filter(id=period_id).first() if period_id else None
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        response = super().create(request, *args, **kwargs)
        if response.status_code < 400:
            plan = ClassPlan.objects.select_related("period").filter(id=response.data.get("id")).first()
            if plan is not None:
                self._log_class_plan_event(request, event_type="class_plan.created", plan=plan, status_code=response.status_code)
                if plan.status == ClassPlan.STATUS_FINALIZED:
                    self._log_class_plan_event(request, event_type="class_plan.finalized", plan=plan, status_code=response.status_code)
        return response

    def update(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        previous_status = None
        instance = None
        if getattr(user, "role", None) == "TEACHER":
            instance = self.get_object()
            period = getattr(instance, "period", None)
            previous_status = getattr(instance, "status", None)
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        if instance is None:
            instance = self.get_object()
            previous_status = getattr(instance, "status", None)
        response = super().update(request, *args, **kwargs)
        if response.status_code < 400:
            plan = ClassPlan.objects.select_related("period").filter(id=instance.id).first()
            if plan is not None:
                self._log_class_plan_event(request, event_type="class_plan.updated", plan=plan, status_code=response.status_code)
                if previous_status != ClassPlan.STATUS_FINALIZED and plan.status == ClassPlan.STATUS_FINALIZED:
                    self._log_class_plan_event(request, event_type="class_plan.finalized", plan=plan, status_code=response.status_code)
        return response

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        instance = self.get_object()
        if getattr(user, "role", None) == "TEACHER":
            period = getattr(instance, "period", None)
            if period is not None and not self._teacher_can_edit_planning(user, period):
                return Response(
                    {"detail": "La edición de planeación está cerrada para este periodo.", "code": "EDIT_WINDOW_CLOSED"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        self._log_class_plan_event(request, event_type="class_plan.deleted", plan=instance, status_code=status.HTTP_204_NO_CONTENT)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="my", permission_classes=[IsAuthenticated])
    def my_plans(self, request):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Esta vista aplica solo para docentes."}, status=status.HTTP_400_BAD_REQUEST)

        qs = self.filter_queryset(self.get_queryset()).order_by("-updated_at")
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="my-summary", permission_classes=[IsAuthenticated])
    def my_summary(self, request):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        user = getattr(request, "user", None)
        if getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Esta vista aplica solo para docentes."}, status=status.HTTP_400_BAD_REQUEST)

        plans_qs = self.filter_queryset(self.get_queryset())
        plan_ids = list(plans_qs.values_list("id", flat=True))
        total_plans = len(plan_ids)
        draft_plans = plans_qs.filter(status=ClassPlan.STATUS_DRAFT).count()
        finalized_plans = plans_qs.filter(status=ClassPlan.STATUS_FINALIZED).count()
        ai_assisted_plans = sum(1 for sections in plans_qs.values_list("ai_assisted_sections", flat=True) if sections)
        planned_topic_ids = set(plans_qs.exclude(topic_id__isnull=True).values_list("topic_id", flat=True).distinct())

        topics_qs = PeriodTopic.objects.filter(is_active=True)
        academic_year_id = request.query_params.get("academic_year")
        if academic_year_id:
            topics_qs = topics_qs.filter(period__academic_year_id=academic_year_id)
        period_id = request.query_params.get("period")
        if period_id:
            topics_qs = topics_qs.filter(period_id=period_id)

        assignment_load_ids = list(
            TeacherAssignment.objects.filter(teacher=user).values_list("academic_load_id", flat=True).distinct()
        )
        if academic_year_id:
            assignment_load_ids = list(
                TeacherAssignment.objects.filter(teacher=user, academic_year_id=academic_year_id)
                .values_list("academic_load_id", flat=True)
                .distinct()
            )
        topics_qs = topics_qs.filter(academic_load_id__in=assignment_load_ids)
        available_topic_ids = set(topics_qs.values_list("id", flat=True).distinct())

        from reports.models import ReportJob  # noqa: PLC0415

        class_plan_jobs = [
            job for job in ReportJob.objects.filter(created_by=user, report_type=ReportJob.ReportType.CLASS_PLAN).order_by("-created_at")
            if int((job.params or {}).get("class_plan_id") or 0) in plan_ids
        ]
        export_pending = sum(1 for job in class_plan_jobs if job.status in {ReportJob.Status.PENDING, ReportJob.Status.RUNNING})
        export_completed = sum(1 for job in class_plan_jobs if job.status == ReportJob.Status.SUCCEEDED)

        recent_logs = AuditLog.objects.filter(actor=user, object_type="class_plan", event_type__startswith="class_plan.")
        if academic_year_id:
            recent_logs = recent_logs.filter(metadata__academic_year_id=int(academic_year_id))
        if period_id:
            recent_logs = recent_logs.filter(metadata__period_id=int(period_id))

        recent_activity = [
            {
                "id": item.id,
                "event_type": item.event_type,
                "object_id": item.object_id,
                "created_at": item.created_at,
                "metadata": item.metadata,
            }
            for item in recent_logs.order_by("-created_at", "-id")[:6]
        ]

        return Response(
            {
                "summary": {
                    "total_plans": total_plans,
                    "draft_plans": draft_plans,
                    "finalized_plans": finalized_plans,
                    "ai_assisted_plans": ai_assisted_plans,
                    "available_topics": len(available_topic_ids),
                    "topics_without_plan": max(len(available_topic_ids - planned_topic_ids), 0),
                    "completion_rate": round((finalized_plans / total_plans) * 100, 1) if total_plans else 0.0,
                    "export_pending": export_pending,
                    "export_completed": export_completed,
                },
                "recent_activity": recent_activity,
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=False,
        methods=["post"],
        url_path="generate-draft",
        permission_classes=[IsAuthenticated],
        throttle_classes=[AcademicAIUserRateThrottle],
    )
    def generate_draft(self, request):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        context, metadata, error_response = self._resolve_generate_draft_context(request)
        if error_response is not None:
            return error_response

        try:
            ai_service = AIService()
            payload = ai_service.generate_class_plan_draft(context)
            self._log_class_plan_event(
                request,
                event_type="class_plan.generate_draft",
                status_code=status.HTTP_200_OK,
                metadata=metadata,
            )
            return Response(payload)
        except AIConfigError as e:
            return Response({"detail": str(e), "code": "AI_NOT_CONFIGURED"}, status=status.HTTP_400_BAD_REQUEST)
        except AIParseError:
            return Response(
                {"detail": "La IA devolvió un borrador inválido. Intenta nuevamente.", "code": "AI_INVALID_RESPONSE"},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except AIProviderError:
            return Response(
                {"detail": "No se pudo generar el borrador con IA en este momento.", "code": "AI_PROVIDER_ERROR"},
                status=status.HTTP_502_BAD_GATEWAY,
            )

    @action(
        detail=False,
        methods=["post"],
        url_path="generate-draft-stream",
        permission_classes=[IsAuthenticated],
        throttle_classes=[AcademicAIUserRateThrottle],
    )
    def generate_draft_stream(self, request):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        context, metadata, error_response = self._resolve_generate_draft_context(request)
        if error_response is not None:
            return error_response

        ai_service = AIService()
        try:
            ai_service._ensure_available()
        except AIConfigError as exc:
            return Response({"detail": str(exc), "code": "AI_NOT_CONFIGURED"}, status=status.HTTP_400_BAD_REQUEST)

        def event_stream():
            has_done = False
            try:
                yield json.dumps({"event": "start", "progress": 5}, ensure_ascii=False) + "\n"

                for event in ai_service.generate_class_plan_draft_stream_events(context):
                    if event.get("event") == "done":
                        has_done = True
                    yield json.dumps(event, ensure_ascii=False) + "\n"

                if has_done:
                    self._log_class_plan_event(
                        request,
                        event_type="class_plan.generate_draft",
                        status_code=status.HTTP_200_OK,
                        metadata={**metadata, "mode": "stream"},
                    )
                else:
                    self._log_class_plan_event(
                        request,
                        event_type="class_plan.generate_draft",
                        status_code=status.HTTP_502_BAD_GATEWAY,
                        metadata={**metadata, "mode": "stream", "error": "stream_terminated_without_done"},
                    )
                    yield json.dumps(
                        {
                            "event": "error",
                            "code": "AI_STREAM_INCOMPLETE",
                            "detail": "La generación en tiempo real se interrumpió antes de finalizar.",
                        },
                        ensure_ascii=False,
                    ) + "\n"
            except AIProviderError:
                self._log_class_plan_event(
                    request,
                    event_type="class_plan.generate_draft",
                    status_code=status.HTTP_502_BAD_GATEWAY,
                    metadata={**metadata, "mode": "stream", "error": "provider_error"},
                )
                yield json.dumps(
                    {
                        "event": "error",
                        "code": "AI_PROVIDER_ERROR",
                        "detail": "No se pudo generar el borrador con IA en este momento.",
                    },
                    ensure_ascii=False,
                ) + "\n"
            except Exception:
                self._log_class_plan_event(
                    request,
                    event_type="class_plan.generate_draft",
                    status_code=status.HTTP_502_BAD_GATEWAY,
                    metadata={**metadata, "mode": "stream", "error": "unexpected_stream_error"},
                )
                yield json.dumps(
                    {
                        "event": "error",
                        "code": "AI_STREAM_ERROR",
                        "detail": "No se pudo completar la generación en tiempo real.",
                    },
                    ensure_ascii=False,
                ) + "\n"

        response = StreamingHttpResponse(event_stream(), content_type="application/x-ndjson")
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="generate-section",
        permission_classes=[IsAuthenticated],
        throttle_classes=[AcademicAIUserRateThrottle],
    )
    def generate_section(self, request):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        section = str(request.data.get("section") or "").strip()
        if not section:
            return Response({"detail": "section es requerido."}, status=status.HTTP_400_BAD_REQUEST)

        context = {
            "topic_title": str(request.data.get("topic_title") or "").strip(),
            "topic_description": str(request.data.get("topic_description") or "").strip(),
            "subject_name": str(request.data.get("subject_name") or "").strip(),
            "grade_name": str(request.data.get("grade_name") or "").strip(),
            "group_name": str(request.data.get("group_name") or "").strip(),
            "teacher_name": str(request.data.get("teacher_name") or "").strip(),
            "period_name": str(request.data.get("period_name") or "").strip(),
            "duration_minutes": request.data.get("duration_minutes") or 55,
            "learning_result": str(request.data.get("learning_result") or "").strip(),
            "title": str(request.data.get("title") or "").strip(),
        }

        try:
            context["duration_minutes"] = int(context["duration_minutes"])
        except Exception:
            return Response({"detail": "duration_minutes inválido."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            ai_service = AIService()
            payload = ai_service.generate_class_plan_section(section, context)
            self._log_class_plan_event(
                request,
                event_type="class_plan.generate_section",
                status_code=status.HTTP_200_OK,
                metadata={
                    "section": section,
                    "title": context.get("title"),
                    "topic_title": context.get("topic_title"),
                },
            )
            return Response(payload)
        except AIConfigError as e:
            return Response({"detail": str(e), "code": "AI_NOT_CONFIGURED"}, status=status.HTTP_400_BAD_REQUEST)
        except AIParseError:
            return Response(
                {"detail": "La IA devolvió una sección inválida. Intenta nuevamente.", "code": "AI_INVALID_RESPONSE"},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except AIProviderError:
            return Response(
                {"detail": "No se pudo generar la sección con IA en este momento.", "code": "AI_PROVIDER_ERROR"},
                status=status.HTTP_502_BAD_GATEWAY,
            )

    @action(detail=True, methods=["post"], url_path="export-pdf", permission_classes=[IsAuthenticated])
    def export_pdf(self, request, pk=None):
        denied = self._ensure_can_manage_class_plans(request)
        if denied is not None:
            return denied

        plan = self.get_object()
        if plan.status != ClassPlan.STATUS_FINALIZED:
            return Response({"detail": "Solo se pueden exportar planes finalizados."}, status=status.HTTP_400_BAD_REQUEST)

        public_base = (getattr(settings, "PUBLIC_SITE_URL", "") or "").strip().rstrip("/")
        if not public_base:
            try:
                public_base = request.build_absolute_uri("/").strip().rstrip("/")
            except Exception:
                public_base = ""

        ttl_hours = int(getattr(settings, "REPORT_JOBS_TTL_HOURS", 24))
        params = {"class_plan_id": plan.id}
        if public_base:
            params["public_site_url"] = public_base

        job = ReportJob.objects.create(
            created_by=request.user,
            report_type=ReportJob.ReportType.CLASS_PLAN,
            params=params,
            expires_at=timezone.now() + timedelta(hours=ttl_hours),
        )

        self._log_class_plan_event(
            request,
            event_type="class_plan.export_requested",
            plan=plan,
            status_code=status.HTTP_202_ACCEPTED,
            metadata={"job_id": job.id},
        )

        generate_report_job_pdf.delay(job.id)

        out = ReportJobSerializer(job, context={"request": request}).data
        return Response(out, status=status.HTTP_202_ACCEPTED)

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GradeSheetViewSet(viewsets.ModelViewSet):
    queryset = GradeSheet.objects.all()
    serializer_class = GradeSheetSerializer
    permission_classes = [KampusModelPermissions]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["teacher_assignment", "period"]
    search_fields = [
        "teacher_assignment__teacher__first_name",
        "teacher_assignment__teacher__last_name",
        "teacher_assignment__teacher__username",
        "teacher_assignment__group__name",
        "teacher_assignment__group__grade__name",
        "teacher_assignment__academic_load__subject__name",
        "period__name",
        "period__academic_year__year",
    ]
    ordering_fields = ["updated_at", "created_at", "id"]
    ordering = ["-updated_at", "-id"]

    def get_serializer_class(self):
        if getattr(self, "action", None) == "list":
            return GradeSheetListSerializer
        return super().get_serializer_class()

    def _enforce_teacher_current_period(self, *, user, period: Period):
        """Teachers can only work with the *current* period.

        Current period means: today's date is within [start_date, end_date].
        This prevents exposing future-period grade sheets to teachers.
        """

        is_teacher = user is not None and getattr(user, "role", None) == "TEACHER"
        if not is_teacher:
            return None

        today = timezone.localdate()
        if today < period.start_date or today > period.end_date:
            return Response(
                {
                    "error": "Solo se pueden diligenciar planillas del periodo actual.",
                    "code": "PERIOD_NOT_CURRENT",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return None

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            "teacher_assignment",
            "teacher_assignment__teacher",
            "teacher_assignment__group",
            "teacher_assignment__group__grade",
            "teacher_assignment__academic_load",
            "teacher_assignment__academic_load__subject",
            "period",
            "period__academic_year",
        )
        user = getattr(self.request, "user", None)
        if not user or not user.is_authenticated:
            return qs.none()
        if getattr(user, "role", None) == "TEACHER":
            return qs.filter(teacher_assignment__teacher=user)
        return qs

    @action(detail=True, methods=["post"], url_path="reset")
    def reset(self, request, pk=None):
        """Teacher can reset (clear) all grades/activities for a grade sheet.

        - Only allowed for the current period.
        - Respects the edit window: after deadline, requires FULL grant.
        - Resets grading_mode back to ACHIEVEMENT.
        """

        gradesheet: GradeSheet = self.get_object()
        period: Period = gradesheet.period
        teacher_assignment: TeacherAssignment = gradesheet.teacher_assignment

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        if period.is_closed:
            return Response(
                {"error": "El periodo está cerrado; no se puede restablecer la planilla."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not self._teacher_can_edit_structure_after_deadline(
            user=request.user,
            period=period,
            teacher_assignment=teacher_assignment,
        ):
            return Response(
                {"detail": "La edición está cerrada y no tienes permiso para restablecer la planilla."},
                status=status.HTTP_403_FORBIDDEN,
            )

        with transaction.atomic():
            col_ids = list(
                AchievementActivityColumn.objects.filter(gradesheet=gradesheet)
                .values_list("id", flat=True)
            )

            deleted_activity_grades = 0
            if col_ids:
                deleted_activity_grades = AchievementActivityGrade.objects.filter(column_id__in=col_ids).delete()[0]

            deleted_activity_columns = AchievementActivityColumn.objects.filter(gradesheet=gradesheet).delete()[0]
            deleted_achievement_grades = AchievementGrade.objects.filter(gradesheet=gradesheet).delete()[0]

            reset_mode = False
            if getattr(gradesheet, "grading_mode", None) != GradeSheet.GRADING_MODE_ACHIEVEMENT:
                gradesheet.grading_mode = GradeSheet.GRADING_MODE_ACHIEVEMENT
                gradesheet.save(update_fields=["grading_mode"])
                reset_mode = True

        return Response(
            {
                "detail": "Planilla restablecida.",
                "deleted": {
                    "achievement_grades": deleted_achievement_grades,
                    "activity_columns": deleted_activity_columns,
                    "activity_grades": deleted_activity_grades,
                },
                "reset_grading_mode": reset_mode,
            },
            status=status.HTTP_200_OK,
        )

    def _get_teacher_assignment(self, teacher_assignment_id: int):
        qs = TeacherAssignment.objects.all()
        if getattr(self.request.user, "role", None) == "TEACHER":
            qs = qs.filter(teacher=self.request.user)
        return qs.select_related("academic_year", "group", "academic_load").get(id=teacher_assignment_id)

    def _teacher_allowed_enrollment_ids_after_deadline(
        self,
        *,
        user,
        period: Period,
        teacher_assignment: TeacherAssignment,
    ) -> set[int] | None:
        """Returns None when unrestricted, else the enrollment_id set allowed.

        Mirrors the Gradebook bulk-upsert behavior.
        """

        is_teacher = user is not None and getattr(user, "role", None) == "TEACHER"
        if not is_teacher:
            return None

        effective_deadline = period.grades_edit_until
        if effective_deadline is None:
            return None
        if timezone.now() <= effective_deadline:
            return None

        active_grants = EditGrant.objects.filter(
            granted_to=user,
            scope=EditRequest.SCOPE_GRADES,
            period_id=period.id,
            teacher_assignment_id=teacher_assignment.id,
            valid_until__gte=timezone.now(),
        )

        has_full = active_grants.filter(grant_type=EditRequest.TYPE_FULL).exists()
        if has_full:
            return None

        return set(
            EditGrantItem.objects.filter(grant__in=active_grants).values_list("enrollment_id", flat=True)
        )

    def _teacher_can_edit_structure_after_deadline(
        self,
        *,
        user,
        period: Period,
        teacher_assignment: TeacherAssignment,
    ) -> bool:
        """Whether a teacher can edit gradebook structure (activity columns) after deadline.

        We require a FULL grant when the deadline has passed.
        """

        is_teacher = user is not None and getattr(user, "role", None) == "TEACHER"
        if not is_teacher:
            return True
        if period.grades_edit_until is None or timezone.now() <= period.grades_edit_until:
            return True

        return EditGrant.objects.filter(
            granted_to=user,
            scope=EditRequest.SCOPE_GRADES,
            period_id=period.id,
            teacher_assignment_id=teacher_assignment.id,
            valid_until__gte=timezone.now(),
            grant_type=EditRequest.TYPE_FULL,
        ).exists()

    def _valid_achievements_for_assignment_period(
        self,
        *,
        teacher_assignment: TeacherAssignment,
        period: Period,
    ):
        base_achievements = Achievement.objects.filter(
            academic_load=teacher_assignment.academic_load,
            period=period,
        )

        group_achievements = base_achievements.filter(group=teacher_assignment.group)
        if group_achievements.exists():
            return group_achievements
        return base_achievements.filter(group__isnull=True)

    def _recompute_achievement_grades_from_activities(
        self,
        *,
        gradesheet: GradeSheet,
        achievement_ids: set[int],
        enrollment_ids: set[int],
    ) -> int:
        """Recompute (and persist) logro grades from activity grades.

        Rule: simple average; missing activity scores count as 1.00.
        Returns the number of AchievementGrade rows upserted.
        """

        if not achievement_ids or not enrollment_ids:
            return 0

        columns = list(
            AchievementActivityColumn.objects.filter(
                gradesheet=gradesheet,
                achievement_id__in=list(achievement_ids),
                is_active=True,
            ).only("id", "achievement_id")
        )
        if not columns:
            return 0

        column_ids = [c.id for c in columns]
        columns_by_achievement: dict[int, list[int]] = {}
        for c in columns:
            columns_by_achievement.setdefault(int(c.achievement_id), []).append(int(c.id))

        grades = list(
            AchievementActivityGrade.objects.filter(
                column_id__in=column_ids,
                enrollment_id__in=list(enrollment_ids),
            ).only("column_id", "enrollment_id", "score")
        )

        score_by_col_enr: dict[tuple[int, int], Decimal | None] = {
            (int(g.column_id), int(g.enrollment_id)): g.score for g in grades
        }

        to_upsert: list[AchievementGrade] = []
        for achievement_id, ach_column_ids in columns_by_achievement.items():
            denom = Decimal(len(ach_column_ids))
            for enrollment_id in enrollment_ids:
                total = sum(
                    coalesce_score(score_by_col_enr.get((col_id, int(enrollment_id))))
                    for col_id in ach_column_ids
                )
                avg = (total / denom).quantize(Decimal("0.01"))
                to_upsert.append(
                    AchievementGrade(
                        gradesheet=gradesheet,
                        enrollment_id=int(enrollment_id),
                        achievement_id=int(achievement_id),
                        score=avg,
                    )
                )

        AchievementGrade.objects.bulk_create(
            to_upsert,
            update_conflicts=True,
            unique_fields=["gradesheet", "enrollment", "achievement"],
            update_fields=["score", "updated_at"],
        )
        return len(to_upsert)

    def _active_enrollment_ids_for_assignment(self, *, teacher_assignment: TeacherAssignment) -> set[int]:
        from students.models import Enrollment

        return set(
            Enrollment.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                group_id=teacher_assignment.group_id,
                status="ACTIVE",
            ).values_list("id", flat=True)
        )

    def _compute_gradebook_completion_stats(
        self,
        *,
        teacher_assignment: TeacherAssignment,
        period: Period,
        gradesheet: GradeSheet,
    ) -> dict:
        from students.models import Enrollment

        enrollments_qs = Enrollment.objects.filter(
            academic_year_id=teacher_assignment.academic_year_id,
            group_id=teacher_assignment.group_id,
            status="ACTIVE",
        )
        students_count = enrollments_qs.count()

        achievements_qs = self._valid_achievements_for_assignment_period(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        achievement_ids = list(achievements_qs.values_list("id", flat=True))
        achievements_count = len(achievement_ids)

        total = students_count * achievements_count
        filled = 0
        if total > 0 and achievement_ids:
            filled = (
                AchievementGrade.objects.filter(
                    gradesheet=gradesheet,
                    enrollment__in=enrollments_qs,
                    achievement_id__in=achievement_ids,
                    score__isnull=False,
                )
                .only("id")
                .count()
            )

        percent = int(round((filled / total) * 100)) if total > 0 else 0
        is_complete = total > 0 and filled >= total
        return {
            "filled": filled,
            "total": total,
            "percent": percent,
            "is_complete": is_complete,
        }

    def _build_gradebook_payload(
        self,
        *,
        gradesheet: GradeSheet,
        teacher_assignment: TeacherAssignment,
        period: Period,
    ) -> dict:
        achievements = self._valid_achievements_for_assignment_period(
            teacher_assignment=teacher_assignment,
            period=period,
        ).select_related("dimension").order_by("id")

        from students.models import Enrollment

        enrollments = (
            Enrollment.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                group_id=teacher_assignment.group_id,
                status="ACTIVE",
            )
            .select_related("student__user")
            .order_by("student__user__last_name", "student__user__first_name")
        )

        existing_grades = AchievementGrade.objects.filter(
            gradesheet=gradesheet,
            enrollment__in=enrollments,
            achievement__in=achievements,
        ).only("enrollment_id", "achievement_id", "score")

        score_by_cell = {
            (g.enrollment_id, g.achievement_id): g.score for g in existing_grades
        }

        achievement_payload = [
            {
                "id": a.id,
                "description": a.description,
                "dimension": a.dimension_id,
                "dimension_name": a.dimension.name if a.dimension else None,
                "percentage": a.percentage,
            }
            for a in achievements
        ]

        student_payload = [
            {
                "enrollment_id": e.id,
                "student_id": e.student_id,
                "student_name": e.student.user.get_full_name(),
            }
            for e in enrollments
        ]

        cell_payload = [
            {
                "enrollment": e.id,
                "achievement": a.id,
                "score": score_by_cell.get((e.id, a.id)),
            }
            for e in enrollments
            for a in achievements
        ]

        achievements_by_dimension: dict[int, list[Achievement]] = {}
        for a in achievements:
            if not a.dimension_id:
                continue
            achievements_by_dimension.setdefault(a.dimension_id, []).append(a)

        dimensions = (
            Dimension.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                id__in=list(achievements_by_dimension.keys()),
            )
            .only("id", "name", "percentage")
            .order_by("id")
        )
        dim_percentage_by_id = {d.id: int(d.percentage) for d in dimensions}

        dimensions_payload = [
            {"id": d.id, "name": d.name, "percentage": int(d.percentage)} for d in dimensions
        ]

        computed = []
        for e in enrollments:
            dim_items = []
            for dim_id, dim_achievements in achievements_by_dimension.items():
                items = [
                    (score_by_cell.get((e.id, a.id)), int(a.percentage) if a.percentage else 1)
                    for a in dim_achievements
                ]
                dim_grade = weighted_average(items) if items else DEFAULT_EMPTY_SCORE
                dim_items.append((dim_grade, dim_percentage_by_id.get(dim_id, 0)))

            final_score = final_grade_from_dimensions(dim_items)
            scale_match = match_scale(teacher_assignment.academic_year_id, final_score)
            computed.append(
                {
                    "enrollment_id": e.id,
                    "final_score": final_score,
                    "scale": scale_match.name if scale_match else None,
                }
            )

        payload = {
            "gradesheet": GradeSheetSerializer(gradesheet).data,
            "period": {"id": period.id, "name": period.name, "is_closed": period.is_closed},
            "teacher_assignment": {
                "id": teacher_assignment.id,
                "group": teacher_assignment.group_id,
                "academic_load": teacher_assignment.academic_load_id,
            },
            "dimensions": dimensions_payload,
            "achievements": achievement_payload,
            "students": student_payload,
            "cells": cell_payload,
            "computed": computed,
        }

        if getattr(gradesheet, "grading_mode", None) == GradeSheet.GRADING_MODE_ACTIVITIES:
            activity_columns = AchievementActivityColumn.objects.filter(
                gradesheet=gradesheet,
                achievement__in=achievements,
            ).order_by("achievement_id", "order", "id")

            activity_grades = AchievementActivityGrade.objects.filter(
                column__in=activity_columns,
                enrollment__in=enrollments,
            ).only("column_id", "enrollment_id", "score")
            activity_score_by_cell = {
                (g.enrollment_id, g.column_id): g.score for g in activity_grades
            }

            payload["activity_columns"] = AchievementActivityColumnSerializer(activity_columns, many=True).data
            payload["activity_cells"] = [
                {
                    "enrollment": e.id,
                    "column": c.id,
                    "score": activity_score_by_cell.get((e.id, c.id)),
                }
                for e in enrollments
                for c in activity_columns
            ]

        return payload

    def _build_gradebook_filled_report_context(
        self,
        *,
        payload: dict,
        teacher_assignment: TeacherAssignment,
        period: Period,
    ) -> tuple[dict | None, str | None]:
        def _fmt_score(value):
            if value is None:
                return ""
            try:
                return f"{Decimal(str(value)):.2f}"
            except Exception:
                return str(value)

        students = payload.get("students") or []
        computed_by_enrollment = {
            int(item.get("enrollment_id")): item for item in (payload.get("computed") or [])
        }
        achievements = payload.get("achievements") or []
        achievement_label_by_id = {int(a.get("id")): f"L{idx + 1}" for idx, a in enumerate(achievements)}
        achievement_desc_by_id = {
            int(a.get("id")): (a.get("description") or "").strip() for a in achievements
        }

        grading_mode = (payload.get("gradesheet") or {}).get("grading_mode")
        columns = []
        column_groups = []
        rows = []

        if grading_mode == GradeSheet.GRADING_MODE_ACTIVITIES and payload.get("activity_columns"):
            activity_columns = payload.get("activity_columns") or []
            activity_cells = payload.get("activity_cells") or []

            filled_column_ids = {
                int(c.get("column"))
                for c in activity_cells
                if c.get("score") is not None and c.get("column") is not None
            }

            ordered_columns = [c for c in activity_columns if int(c.get("id")) in filled_column_ids]
            columns = [
                {
                    "id": int(c.get("id")),
                    "group_label": achievement_label_by_id.get(int(c.get("achievement")), "L?"),
                    "label": (c.get("label") or "Actividad").strip(),
                    "description": achievement_desc_by_id.get(int(c.get("achievement")), ""),
                }
                for c in ordered_columns
            ]

            for col in columns:
                if not column_groups or column_groups[-1]["label"] != col["group_label"]:
                    column_groups.append({"label": col["group_label"], "count": 1})
                else:
                    column_groups[-1]["count"] += 1

            score_by_cell = {
                (int(c.get("enrollment")), int(c.get("column"))): c.get("score")
                for c in activity_cells
                if c.get("enrollment") is not None and c.get("column") is not None
            }

            for s in students:
                enrollment_id = int(s.get("enrollment_id"))
                raw_scores = [score_by_cell.get((enrollment_id, col["id"])) for col in columns]
                if all(score is None for score in raw_scores):
                    continue

                rows.append(
                    {
                        "index": len(rows) + 1,
                        "student_name": s.get("student_name") or "",
                        "scores": [_fmt_score(score) for score in raw_scores],
                        "final_score": _fmt_score((computed_by_enrollment.get(enrollment_id) or {}).get("final_score")),
                    }
                )
        else:
            cells = payload.get("cells") or []

            filled_achievement_ids = {
                int(c.get("achievement"))
                for c in cells
                if c.get("score") is not None and c.get("achievement") is not None
            }

            ordered_achievements = [a for a in achievements if int(a.get("id")) in filled_achievement_ids]
            columns = [
                {
                    "id": int(a.get("id")),
                    "group_label": "Logro",
                    "label": f"L{idx + 1}",
                    "description": (a.get("description") or "").strip(),
                }
                for idx, a in enumerate(ordered_achievements)
            ]

            score_by_cell = {
                (int(c.get("enrollment")), int(c.get("achievement"))): c.get("score")
                for c in cells
                if c.get("enrollment") is not None and c.get("achievement") is not None
            }

            for s in students:
                enrollment_id = int(s.get("enrollment_id"))
                raw_scores = [score_by_cell.get((enrollment_id, col["id"])) for col in columns]
                if all(score is None for score in raw_scores):
                    continue

                rows.append(
                    {
                        "index": len(rows) + 1,
                        "student_name": s.get("student_name") or "",
                        "scores": [_fmt_score(score) for score in raw_scores],
                        "final_score": _fmt_score((computed_by_enrollment.get(enrollment_id) or {}).get("final_score")),
                    }
                )

        if not columns or not rows:
            return None, "La planilla actual no tiene notas diligenciadas para generar el informe."

        generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M")
        grade_name = ""
        try:
            if teacher_assignment.group and teacher_assignment.group.grade:
                grade_name = teacher_assignment.group.grade.name or ""
        except Exception:
            grade_name = ""

        context = {
            "title": "Informe de planilla diligenciada",
            "grading_mode": grading_mode or GradeSheet.GRADING_MODE_ACHIEVEMENT,
            "group_name": teacher_assignment.group.name if teacher_assignment.group else "",
            "grade_name": grade_name,
            "subject_name": teacher_assignment.academic_load.subject.name if teacher_assignment.academic_load else "",
            "period_name": period.name,
            "generated_at": generated_at,
            "columns": columns,
            "column_groups": column_groups,
            "rows": rows,
        }

        institution = Institution.objects.first() or Institution()
        institution_logo_src = ""
        try:
            if getattr(institution, "pdf_show_logo", True) and getattr(institution, "logo", None):
                logo_field = institution.logo
                if getattr(logo_field, "path", None) and Path(logo_field.path).exists():
                    institution_logo_src = Path(logo_field.path).resolve().as_uri()
                elif getattr(logo_field, "url", None):
                    institution_logo_src = logo_field.url
        except Exception:
            institution_logo_src = ""

        context["institution_logo_src"] = institution_logo_src
        return context, None

    def _send_gradebook_completed_email(
        self,
        *,
        teacher,
        teacher_assignment: TeacherAssignment,
        period: Period,
        gradesheet: GradeSheet,
        payload: dict,
    ) -> None:
        from reports.weasyprint_utils import render_pdf_bytes_from_html

        recipient_email = (getattr(teacher, "email", "") or "").strip()
        if not recipient_email:
            return

        context, error_message = self._build_gradebook_filled_report_context(
            payload=payload,
            teacher_assignment=teacher_assignment,
            period=period,
        )
        if error_message:
            return

        html_string = render_to_string("academic/reports/gradebook_filled_report_pdf.html", context)
        pdf_bytes = render_pdf_bytes_from_html(html=html_string, base_url=str(settings.BASE_DIR))

        group_name = getattr(getattr(teacher_assignment, "group", None), "name", "") or "grupo"
        grade_name = ""
        try:
            if teacher_assignment.group and teacher_assignment.group.grade:
                grade_name = teacher_assignment.group.grade.name or ""
        except Exception:
            grade_name = ""
        period_name = getattr(period, "name", "periodo") or "periodo"
        attachment_name = f"planilla-diligenciada-{group_name}-{period_name}.pdf".replace(" ", "_")

        teacher_name = ""
        try:
            teacher_name = teacher.get_full_name() or getattr(teacher, "username", "")
        except Exception:
            teacher_name = getattr(teacher, "username", "") or "Docente"

        subject_name = ""
        try:
            if teacher_assignment.academic_load and teacher_assignment.academic_load.subject:
                subject_name = teacher_assignment.academic_load.subject.name or ""
        except Exception:
            subject_name = ""

        detail_parts = [p for p in [subject_name, period_name] if p]
        detail_text = " | ".join(detail_parts)
        body_text = (
            f"Hola {teacher_name},\n\n"
            "Tu planilla de calificaciones fue completada al 100%.\n"
            f"Grado: {grade_name or 'N/A'}\n"
            f"Grupo: {group_name or 'N/A'}\n"
            f"Detalle: {detail_text}\n\n"
            "Adjuntamos el informe en PDF, generado con el mismo formato del botón 'Descargar informe de planilla'."
        )

        send_email(
            recipient_email=recipient_email,
            subject="[Kampus] Tu planilla de calificaciones está completa (100%)",
            body_text=body_text,
            attachments=[(attachment_name, pdf_bytes, "application/pdf")],
            category="gradebook-complete",
            idempotency_key=f"gradebook-complete-email:sheet={gradesheet.id}",
        )

    @action(detail=False, methods=["get"], url_path="available")
    def available(self, request):
        period_id = request.query_params.get("period")
        if not period_id:
            return Response({"error": "period es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        tas = TeacherAssignment.objects.filter(academic_year_id=period.academic_year_id)
        if getattr(request.user, "role", None) == "TEACHER":
            tas = tas.filter(teacher=request.user)

        tas = tas.select_related(
            "group",
            "group__grade",
            "academic_load",
            "academic_load__subject",
        ).order_by("group__grade__name", "group__name", "academic_load__subject__name")

        from students.models import Enrollment

        items = []
        for ta in tas:
            enrollments_qs = Enrollment.objects.filter(
                academic_year_id=ta.academic_year_id,
                group_id=ta.group_id,
                status="ACTIVE",
            )
            students_count = enrollments_qs.count()

            base_achievements = Achievement.objects.filter(
                academic_load_id=ta.academic_load_id,
                period_id=period.id,
            )
            group_achievements = base_achievements.filter(group_id=ta.group_id)
            if group_achievements.exists():
                achievements_qs = group_achievements
            else:
                achievements_qs = base_achievements.filter(group__isnull=True)

            achievement_ids = list(achievements_qs.values_list("id", flat=True))
            achievements_count = len(achievement_ids)

            total = students_count * achievements_count

            gradesheet_id = (
                GradeSheet.objects.filter(teacher_assignment_id=ta.id, period_id=period.id)
                .values_list("id", flat=True)
                .first()
            )

            filled = 0
            if gradesheet_id and total > 0 and achievement_ids:
                filled = (
                    AchievementGrade.objects.filter(
                        gradesheet_id=gradesheet_id,
                        enrollment__in=enrollments_qs,
                        achievement_id__in=achievement_ids,
                        score__isnull=False,
                    )
                    .only("id")
                    .count()
                )

            percent = int(round((filled / total) * 100)) if total > 0 else 0
            is_complete = total > 0 and filled >= total

            items.append(
                {
                    "teacher_assignment_id": ta.id,
                    "group_id": ta.group_id,
                    "group_name": ta.group.name,
                    "grade_id": ta.group.grade_id,
                    "grade_name": ta.group.grade.name,
                    "academic_load_id": ta.academic_load_id,
                    "subject_name": ta.academic_load.subject.name if ta.academic_load_id else None,
                    "period": {"id": period.id, "name": period.name, "is_closed": period.is_closed},
                    "students_count": students_count,
                    "achievements_count": achievements_count,
                    "completion": {
                        "filled": filled,
                        "total": total,
                        "percent": percent,
                        "is_complete": is_complete,
                    },
                }
            )

        return Response({"results": items})

    @action(detail=False, methods=["get"], url_path="gradebook")
    def gradebook(self, request):
        teacher_assignment_id = request.query_params.get("teacher_assignment")
        period_id = request.query_params.get("period")
        if not teacher_assignment_id or not period_id:
            return Response(
                {"error": "teacher_assignment y period son requeridos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        payload = self._build_gradebook_payload(
            gradesheet=gradesheet,
            teacher_assignment=teacher_assignment,
            period=period,
        )

        return Response(payload)

    @action(detail=False, methods=["get"], url_path="gradebook-filled-report")
    def gradebook_filled_report(self, request):
        """Genera PDF de la planilla *actual* con notas diligenciadas.

        A diferencia de la sabana o boletin, este reporte usa exactamente la
        planilla del docente/asignacion/periodo actual y muestra solo columnas
        y filas con notas diligenciadas.

        GET /api/grade-sheets/gradebook-filled-report/?teacher_assignment=<id>&period=<id>&format=pdf|html
        """

        teacher_assignment_id = request.query_params.get("teacher_assignment")
        period_id = request.query_params.get("period")
        if not teacher_assignment_id or not period_id:
            return Response(
                {"error": "teacher_assignment y period son requeridos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )

        if gradesheet.grading_mode == GradeSheet.GRADING_MODE_ACTIVITIES:
            achievement_ids = set(
                self._valid_achievements_for_assignment_period(
                    teacher_assignment=teacher_assignment,
                    period=period,
                ).values_list("id", flat=True)
            )
            enrollment_ids = self._active_enrollment_ids_for_assignment(teacher_assignment=teacher_assignment)
            self._recompute_achievement_grades_from_activities(
                gradesheet=gradesheet,
                achievement_ids=achievement_ids,
                enrollment_ids=enrollment_ids,
            )

        payload = self._build_gradebook_payload(
            gradesheet=gradesheet,
            teacher_assignment=teacher_assignment,
            period=period,
        )

        fmt = (request.query_params.get("format") or "pdf").strip().lower()

        context, error_message = self._build_gradebook_filled_report_context(
            payload=payload,
            teacher_assignment=teacher_assignment,
            period=period,
        )
        if error_message:
            return Response({"error": error_message}, status=status.HTTP_400_BAD_REQUEST)

        html_string = render_to_string("academic/reports/gradebook_filled_report_pdf.html", context)

        if fmt == "html":
            return HttpResponse(html_string, content_type="text/html; charset=utf-8")

        from reports.weasyprint_utils import WeasyPrintUnavailableError, render_pdf_bytes_from_html

        try:
            filename = f"planilla-diligenciada-{teacher_assignment.group.name}-{period.name}.pdf".replace(" ", "_")
            pdf_bytes = render_pdf_bytes_from_html(html=html_string, base_url=str(settings.BASE_DIR))
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except WeasyPrintUnavailableError as e:
            return Response({"detail": str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as e:
            payload = {"detail": "Error generating PDF", "error": str(e)}
            from django.conf import settings as _settings
            import traceback as _traceback

            if getattr(_settings, "DEBUG", False):
                payload["traceback"] = _traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=["post"], url_path="set-grading-mode")
    def set_grading_mode(self, request):
        serializer = GradeSheetGradingModeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        teacher_assignment_id = serializer.validated_data["teacher_assignment"]
        period_id = serializer.validated_data["period"]
        grading_mode = serializer.validated_data["grading_mode"]
        default_columns = serializer.validated_data.get("default_columns")

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        if period.is_closed:
            return Response(
                {"error": "El periodo está cerrado; no se puede modificar la planilla."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = getattr(request, "user", None)
        if not self._teacher_can_edit_structure_after_deadline(
            user=user,
            period=period,
            teacher_assignment=teacher_assignment,
        ):
            return Response(
                {"detail": "La ventana de edición está cerrada para modificar columnas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )

        created_columns = 0
        with transaction.atomic():
            if gradesheet.grading_mode != grading_mode:
                gradesheet.grading_mode = grading_mode
                gradesheet.save(update_fields=["grading_mode", "updated_at"])

            if grading_mode == GradeSheet.GRADING_MODE_ACTIVITIES:
                n = 2 if default_columns is None else int(default_columns)
                if n > 0:
                    achievements = self._valid_achievements_for_assignment_period(
                        teacher_assignment=teacher_assignment,
                        period=period,
                    ).only("id")
                    existing_pairs = set(
                        AchievementActivityColumn.objects.filter(
                            gradesheet=gradesheet,
                            achievement__in=achievements,
                        ).values_list("achievement_id", flat=True)
                    )

                    to_create = []
                    for a in achievements:
                        if int(a.id) in existing_pairs:
                            continue
                        for i in range(1, n + 1):
                            to_create.append(
                                AchievementActivityColumn(
                                    gradesheet=gradesheet,
                                    achievement_id=int(a.id),
                                    label=f"Actividad {i}",
                                    order=i,
                                    is_active=True,
                                )
                            )
                    if to_create:
                        AchievementActivityColumn.objects.bulk_create(to_create)
                        created_columns = len(to_create)

        recomputed = 0
        if grading_mode == GradeSheet.GRADING_MODE_ACTIVITIES:
            achievement_ids = set(
                self._valid_achievements_for_assignment_period(
                    teacher_assignment=teacher_assignment,
                    period=period,
                ).values_list("id", flat=True)
            )
            enrollment_ids = self._active_enrollment_ids_for_assignment(teacher_assignment=teacher_assignment)
            recomputed = self._recompute_achievement_grades_from_activities(
                gradesheet=gradesheet,
                achievement_ids=achievement_ids,
                enrollment_ids=enrollment_ids,
            )

        return Response(
            {
                "gradesheet": GradeSheetSerializer(gradesheet).data,
                "created_columns": created_columns,
                "recomputed_achievement_grades": recomputed,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="activity-columns")
    def activity_columns(self, request):
        teacher_assignment_id = request.query_params.get("teacher_assignment")
        period_id = request.query_params.get("period")
        if not teacher_assignment_id or not period_id:
            return Response(
                {"error": "teacher_assignment y period son requeridos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )

        achievements = self._valid_achievements_for_assignment_period(
            teacher_assignment=teacher_assignment,
            period=period,
        ).only("id")

        cols = AchievementActivityColumn.objects.filter(
            gradesheet=gradesheet,
            achievement__in=achievements,
        ).order_by("achievement_id", "order", "id")

        return Response(
            {
                "gradesheet": GradeSheetSerializer(gradesheet).data,
                "columns": AchievementActivityColumnSerializer(cols, many=True).data,
            }
        )

    @action(detail=False, methods=["post"], url_path="activity-columns/bulk-upsert")
    def activity_columns_bulk_upsert(self, request):
        serializer = ActivityColumnsBulkUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        teacher_assignment_id = serializer.validated_data["teacher_assignment"]
        period_id = serializer.validated_data["period"]
        columns_payload = serializer.validated_data["columns"]

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        if period.is_closed:
            return Response(
                {"error": "El periodo está cerrado; no se pueden modificar columnas."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = getattr(request, "user", None)
        if not self._teacher_can_edit_structure_after_deadline(
            user=user,
            period=period,
            teacher_assignment=teacher_assignment,
        ):
            return Response(
                {"detail": "La ventana de edición está cerrada para modificar columnas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )

        achievements_qs = self._valid_achievements_for_assignment_period(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        valid_achievement_ids = set(achievements_qs.values_list("id", flat=True))

        # Load existing columns for conflict detection
        existing_cols = list(
            AchievementActivityColumn.objects.filter(
                gradesheet=gradesheet,
                achievement_id__in=list(valid_achievement_ids),
            ).only("id", "achievement_id", "order")
        )
        existing_by_id = {int(c.id): c for c in existing_cols}
        existing_taken: dict[tuple[int, int], int] = {
            (int(c.achievement_id), int(c.order)): int(c.id) for c in existing_cols
        }

        # Precompute next order per achievement when omitted
        max_order_by_ach: dict[int, int] = {}
        for c in existing_cols:
            aid = int(c.achievement_id)
            max_order_by_ach[aid] = max(max_order_by_ach.get(aid, 0), int(c.order))

        desired_keys: dict[tuple[int, int], int | None] = {}
        to_create: list[AchievementActivityColumn] = []
        to_update: list[AchievementActivityColumn] = []

        for item in columns_payload:
            col_id = item.get("id")
            achievement_id = int(item["achievement"])
            if achievement_id not in valid_achievement_ids:
                return Response(
                    {"error": f"Achievement inválido: {achievement_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            order = item.get("order")
            if order in (None, ""):
                max_order_by_ach[achievement_id] = max_order_by_ach.get(achievement_id, 0) + 1
                order = max_order_by_ach[achievement_id]
            order = int(order)

            key = (achievement_id, order)
            if key in desired_keys:
                return Response(
                    {"error": f"Orden duplicado en payload para achievement={achievement_id} order={order}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            desired_keys[key] = int(col_id) if col_id else None

            # Conflict with existing db row not being updated to this id
            taken_id = existing_taken.get(key)
            if taken_id is not None and (not col_id or int(col_id) != int(taken_id)):
                return Response(
                    {"error": f"Ya existe una columna con achievement={achievement_id} y order={order}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            label = item["label"]
            is_active = bool(item.get("is_active", True))

            if col_id:
                existing = existing_by_id.get(int(col_id))
                if not existing or int(existing.achievement_id) != achievement_id:
                    return Response(
                        {"error": f"Columna inválida o no pertenece a la planilla: {col_id}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                existing.label = label
                existing.order = order
                existing.is_active = is_active
                to_update.append(existing)
            else:
                to_create.append(
                    AchievementActivityColumn(
                        gradesheet=gradesheet,
                        achievement_id=achievement_id,
                        label=label,
                        order=order,
                        is_active=is_active,
                    )
                )

        with transaction.atomic():
            if to_create:
                AchievementActivityColumn.objects.bulk_create(to_create)
            if to_update:
                AchievementActivityColumn.objects.bulk_update(to_update, ["label", "order", "is_active", "updated_at"])

        # Keep persisted achievement grades in sync when activities structure changes.
        touched_achievement_ids = {int(item["achievement"]) for item in columns_payload}
        recomputed = 0
        if touched_achievement_ids and gradesheet.grading_mode == GradeSheet.GRADING_MODE_ACTIVITIES:
            enrollment_ids = self._active_enrollment_ids_for_assignment(teacher_assignment=teacher_assignment)
            recomputed = self._recompute_achievement_grades_from_activities(
                gradesheet=gradesheet,
                achievement_ids=touched_achievement_ids,
                enrollment_ids=enrollment_ids,
            )

        cols = AchievementActivityColumn.objects.filter(
            gradesheet=gradesheet,
            achievement_id__in=list(valid_achievement_ids),
        ).order_by("achievement_id", "order", "id")

        return Response(
            {
                "created": len(to_create),
                "updated": len(to_update),
                "recomputed_achievement_grades": recomputed,
                "columns": AchievementActivityColumnSerializer(cols, many=True).data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["post"], url_path="activity-grades/bulk-upsert")
    def activity_grades_bulk_upsert(self, request):
        serializer = ActivityGradesBulkUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        teacher_assignment_id = serializer.validated_data["teacher_assignment"]
        period_id = serializer.validated_data["period"]
        grades_payload = serializer.validated_data["grades"]

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.is_closed:
            return Response(
                {"error": "El periodo está cerrado; no se pueden registrar notas."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        completion_before = self._compute_gradebook_completion_stats(
            teacher_assignment=teacher_assignment,
            period=period,
            gradesheet=gradesheet,
        )

        from students.models import Enrollment

        valid_enrollments = set(
            Enrollment.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                group_id=teacher_assignment.group_id,
            ).values_list("id", flat=True)
        )

        achievements_qs = self._valid_achievements_for_assignment_period(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        valid_achievement_ids = set(achievements_qs.values_list("id", flat=True))

        user = getattr(request, "user", None)
        allowed_enrollment_ids = self._teacher_allowed_enrollment_ids_after_deadline(
            user=user,
            period=period,
            teacher_assignment=teacher_assignment,
        )

        column_ids = sorted({int(g["column"]) for g in grades_payload})
        columns = list(
            AchievementActivityColumn.objects.filter(
                id__in=column_ids,
                gradesheet=gradesheet,
                is_active=True,
            ).only("id", "achievement_id")
        )
        columns_by_id = {int(c.id): c for c in columns}
        if len(columns_by_id) != len(column_ids):
            return Response(
                {"error": "Una o más columnas son inválidas o no pertenecen a la planilla."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        blocked = []
        allowed_by_cell: dict[tuple[int, int], AchievementActivityGrade] = {}
        impacted_enrollment_ids: set[int] = set()
        impacted_achievement_ids: set[int] = set()

        for g in grades_payload:
            enrollment_id = int(g["enrollment"])
            column_id = int(g["column"])
            if enrollment_id not in valid_enrollments:
                return Response(
                    {"error": f"Enrollment inválido: {enrollment_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            column_obj = columns_by_id.get(column_id)
            achievement_id = int(getattr(column_obj, "achievement_id"))
            if achievement_id not in valid_achievement_ids:
                return Response(
                    {"error": f"Achievement inválido para la columna: {achievement_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if allowed_enrollment_ids is not None and enrollment_id not in allowed_enrollment_ids:
                blocked.append(
                    {
                        "enrollment": enrollment_id,
                        "column": column_id,
                        "reason": "EDIT_WINDOW_CLOSED",
                    }
                )
                continue

            impacted_enrollment_ids.add(enrollment_id)
            impacted_achievement_ids.add(achievement_id)

            allowed_by_cell[(column_id, enrollment_id)] = AchievementActivityGrade(
                column_id=column_id,
                enrollment_id=enrollment_id,
                score=g.get("score"),
            )

        to_upsert = list(allowed_by_cell.values())
        if to_upsert:
            with transaction.atomic():
                AchievementActivityGrade.objects.bulk_create(
                    to_upsert,
                    update_conflicts=True,
                    unique_fields=["column", "enrollment"],
                    update_fields=["score", "updated_at"],
                )

                recomputed = self._recompute_achievement_grades_from_activities(
                    gradesheet=gradesheet,
                    achievement_ids=impacted_achievement_ids,
                    enrollment_ids=impacted_enrollment_ids,
                )

        else:
            recomputed = 0

        # Return recomputed final scores for impacted enrollments
        impacted_enrollment_ids_sorted = sorted(impacted_enrollment_ids)

        achievements = achievements_qs.select_related("dimension").order_by("id")

        achievements_by_dimension: dict[int, list[Achievement]] = {}
        for a in achievements:
            if not a.dimension_id:
                continue
            achievements_by_dimension.setdefault(a.dimension_id, []).append(a)

        dimensions = Dimension.objects.filter(
            academic_year_id=teacher_assignment.academic_year_id,
            id__in=list(achievements_by_dimension.keys()),
        ).only("id", "percentage")
        dim_percentage_by_id = {d.id: int(d.percentage) for d in dimensions}

        existing_grades = AchievementGrade.objects.filter(
            gradesheet=gradesheet,
            enrollment_id__in=impacted_enrollment_ids_sorted,
            achievement__in=achievements,
        ).only("enrollment_id", "achievement_id", "score")
        score_by_cell = {(g.enrollment_id, g.achievement_id): g.score for g in existing_grades}

        computed = []
        for enrollment_id in impacted_enrollment_ids_sorted:
            dim_items = []
            for dim_id, dim_achievements in achievements_by_dimension.items():
                items = [
                    (
                        score_by_cell.get((enrollment_id, a.id)),
                        int(a.percentage) if a.percentage else 1,
                    )
                    for a in dim_achievements
                ]
                dim_grade = weighted_average(items) if items else DEFAULT_EMPTY_SCORE
                dim_items.append((dim_grade, dim_percentage_by_id.get(dim_id, 0)))

            final_score = final_grade_from_dimensions(dim_items)
            scale_match = match_scale(teacher_assignment.academic_year_id, final_score)
            computed.append(
                {
                    "enrollment_id": enrollment_id,
                    "final_score": final_score,
                    "scale": scale_match.name if scale_match else None,
                }
            )

        completion_after = self._compute_gradebook_completion_stats(
            teacher_assignment=teacher_assignment,
            period=period,
            gradesheet=gradesheet,
        )
        transitioned_to_complete = (not completion_before["is_complete"]) and completion_after["is_complete"]
        is_teacher = getattr(getattr(request, "user", None), "role", None) == "TEACHER"
        if transitioned_to_complete and is_teacher:
            try:
                payload = self._build_gradebook_payload(
                    gradesheet=gradesheet,
                    teacher_assignment=teacher_assignment,
                    period=period,
                )
                self._send_gradebook_completed_email(
                    teacher=request.user,
                    teacher_assignment=teacher_assignment,
                    period=period,
                    gradesheet=gradesheet,
                    payload=payload,
                )
            except Exception:
                # Email dispatch is best-effort and must not break grade registration.
                pass

        return Response(
            {
                "requested": len(grades_payload),
                "updated": len(to_upsert),
                "recomputed_achievement_grades": recomputed,
                "computed": computed,
                "blocked": blocked,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["post"], url_path="bulk-upsert")
    def bulk_upsert(self, request):
        serializer = GradebookBulkUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        teacher_assignment_id = serializer.validated_data["teacher_assignment"]
        period_id = serializer.validated_data["period"]
        grades = serializer.validated_data["grades"]

        try:
            teacher_assignment = self._get_teacher_assignment(int(teacher_assignment_id))
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.is_closed:
            return Response(
                {"error": "El periodo está cerrado; no se pueden registrar notas."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = getattr(request, "user", None)
        is_teacher = user is not None and getattr(user, "role", None) == "TEACHER"

        # Teacher edit window enforcement (admins/coordinators are not restricted by deadline)
        allowed_enrollment_ids: set[int] | None = None
        # Only enforce deadline when explicitly configured (avoid time-dependent behavior in tests
        # and keep legacy behavior where open periods allow edits).
        effective_deadline = period.grades_edit_until
        if is_teacher and effective_deadline is not None and timezone.now() > effective_deadline:
            active_grants = EditGrant.objects.filter(
                granted_to=user,
                scope=EditRequest.SCOPE_GRADES,
                period_id=period.id,
                teacher_assignment_id=teacher_assignment.id,
                valid_until__gte=timezone.now(),
            )

            has_full = active_grants.filter(grant_type=EditRequest.TYPE_FULL).exists()
            if not has_full:
                allowed_enrollment_ids = set(
                    EditGrantItem.objects.filter(grant__in=active_grants)
                    .values_list("enrollment_id", flat=True)
                )

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        completion_before = self._compute_gradebook_completion_stats(
            teacher_assignment=teacher_assignment,
            period=period,
            gradesheet=gradesheet,
        )

        from students.models import Enrollment

        valid_enrollments = set(
            Enrollment.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                group_id=teacher_assignment.group_id,
            ).values_list("id", flat=True)
        )

        base_achievements = Achievement.objects.filter(
            academic_load=teacher_assignment.academic_load,
            period=period,
        )
        group_achievements = base_achievements.filter(group=teacher_assignment.group)
        if group_achievements.exists():
            valid_achievements = set(group_achievements.values_list("id", flat=True))
        else:
            valid_achievements = set(
                base_achievements.filter(group__isnull=True).values_list("id", flat=True)
            )

        blocked = []
        allowed_by_cell: dict[tuple[int, int], AchievementGrade] = {}
        for g in grades:
            enrollment_id = g["enrollment"]
            achievement_id = g["achievement"]
            if enrollment_id not in valid_enrollments:
                return Response(
                    {"error": f"Enrollment inválido: {enrollment_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if achievement_id not in valid_achievements:
                return Response(
                    {"error": f"Achievement inválido: {achievement_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if allowed_enrollment_ids is not None and enrollment_id not in allowed_enrollment_ids:
                blocked.append(
                    {
                        "enrollment": enrollment_id,
                        "achievement": achievement_id,
                        "reason": "EDIT_WINDOW_CLOSED",
                    }
                )
                continue

            # Deduplicate per cell (last write wins)
            allowed_by_cell[(enrollment_id, achievement_id)] = AchievementGrade(
                gradesheet=gradesheet,
                enrollment_id=enrollment_id,
                achievement_id=achievement_id,
                score=g.get("score"),
            )

        to_upsert = list(allowed_by_cell.values())

        if to_upsert:
            with transaction.atomic():
                AchievementGrade.objects.bulk_create(
                    to_upsert,
                    update_conflicts=True,
                    unique_fields=["gradesheet", "enrollment", "achievement"],
                    update_fields=["score", "updated_at"],
                )

        # Return recomputed final scores for impacted enrollments (for live UI updates)
        impacted_enrollment_ids = sorted({g.enrollment_id for g in to_upsert})

        base_achievements = Achievement.objects.filter(
            academic_load=teacher_assignment.academic_load,
            period=period,
        )

        group_achievements = base_achievements.filter(group=teacher_assignment.group)
        if group_achievements.exists():
            achievements = group_achievements
        else:
            achievements = base_achievements.filter(group__isnull=True)

        achievements = achievements.select_related("dimension").order_by("id")

        achievements_by_dimension: dict[int, list[Achievement]] = {}
        for a in achievements:
            if not a.dimension_id:
                continue
            achievements_by_dimension.setdefault(a.dimension_id, []).append(a)

        dimensions = Dimension.objects.filter(
            academic_year_id=teacher_assignment.academic_year_id,
            id__in=list(achievements_by_dimension.keys()),
        ).only("id", "percentage")
        dim_percentage_by_id = {d.id: int(d.percentage) for d in dimensions}

        existing_grades = AchievementGrade.objects.filter(
            gradesheet=gradesheet,
            enrollment_id__in=impacted_enrollment_ids,
            achievement__in=achievements,
        ).only("enrollment_id", "achievement_id", "score")

        score_by_cell = {(g.enrollment_id, g.achievement_id): g.score for g in existing_grades}

        computed = []
        for enrollment_id in impacted_enrollment_ids:
            dim_items = []
            for dim_id, dim_achievements in achievements_by_dimension.items():
                items = [
                    (
                        score_by_cell.get((enrollment_id, a.id)),
                        int(a.percentage) if a.percentage else 1,
                    )
                    for a in dim_achievements
                ]
                dim_grade = weighted_average(items) if items else DEFAULT_EMPTY_SCORE
                dim_items.append((dim_grade, dim_percentage_by_id.get(dim_id, 0)))

            final_score = final_grade_from_dimensions(dim_items)
            scale_match = match_scale(teacher_assignment.academic_year_id, final_score)
            computed.append(
                {
                    "enrollment_id": enrollment_id,
                    "final_score": final_score,
                    "scale": scale_match.name if scale_match else None,
                }
            )

        completion_after = self._compute_gradebook_completion_stats(
            teacher_assignment=teacher_assignment,
            period=period,
            gradesheet=gradesheet,
        )
        transitioned_to_complete = (not completion_before["is_complete"]) and completion_after["is_complete"]
        if transitioned_to_complete and is_teacher:
            try:
                payload = self._build_gradebook_payload(
                    gradesheet=gradesheet,
                    teacher_assignment=teacher_assignment,
                    period=period,
                )
                self._send_gradebook_completed_email(
                    teacher=user,
                    teacher_assignment=teacher_assignment,
                    period=period,
                    gradesheet=gradesheet,
                    payload=payload,
                )
            except Exception:
                # Email dispatch is best-effort and must not break grade registration.
                pass

        return Response(
            {
                "requested": len(grades),
                "updated": len(to_upsert),
                "computed": computed,
                "blocked": blocked,
            },
            status=status.HTTP_200_OK,
        )


class PreschoolGradebookViewSet(viewsets.ViewSet):
    """Preescolar: planilla cualitativa (SIEE) por logro.

    API dedicada para no mezclar con la planilla numérica.
    - Solo para docentes.
    - Solo para asignaciones cuyo grupo sea de nivel preescolar.
    - Reutiliza las mismas reglas de bloqueo (periodo cerrado, ventana de edición, grants).
    """

    permission_classes = [IsAuthenticated]

    def _require_teacher(self, request):
        user = getattr(request, "user", None)
        if not user or not user.is_authenticated:
            return Response({"detail": "No autenticado."}, status=status.HTTP_401_UNAUTHORIZED)
        if getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Solo docentes."}, status=status.HTTP_403_FORBIDDEN)
        return None

    def _enforce_teacher_current_period(self, *, user, period: Period):
        today = timezone.localdate()
        if today < period.start_date or today > period.end_date:
            return Response(
                {
                    "error": "Solo se pueden diligenciar planillas del periodo actual.",
                    "code": "PERIOD_NOT_CURRENT",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return None

    def _get_teacher_assignment(self, *, user, teacher_assignment_id: int) -> TeacherAssignment:
        return (
            TeacherAssignment.objects.filter(teacher=user)
            .select_related(
                "academic_year",
                "group",
                "group__grade",
                "group__grade__level",
                "academic_load",
                "academic_load__subject",
            )
            .get(id=teacher_assignment_id)
        )

    def _ensure_preschool_assignment(self, teacher_assignment: TeacherAssignment):
        group = getattr(teacher_assignment, "group", None)
        grade = getattr(group, "grade", None) if group is not None else None
        level = getattr(grade, "level", None) if grade is not None else None
        level_type = getattr(level, "level_type", None) if level is not None else None

        if level_type != "PRESCHOOL":
            return Response(
                {"detail": "Esta planilla aplica solo para grupos de preescolar."},
                status=status.HTTP_403_FORBIDDEN,
            )
        return None

    def _teacher_allowed_enrollment_ids_after_deadline(
        self,
        *,
        user,
        period: Period,
        teacher_assignment: TeacherAssignment,
    ) -> set[int] | None:
        effective_deadline = period.grades_edit_until
        if effective_deadline is None:
            return None
        if timezone.now() <= effective_deadline:
            return None

        active_grants = EditGrant.objects.filter(
            granted_to=user,
            scope=EditRequest.SCOPE_GRADES,
            period_id=period.id,
            teacher_assignment_id=teacher_assignment.id,
            valid_until__gte=timezone.now(),
        )

        has_full = active_grants.filter(grant_type=EditRequest.TYPE_FULL).exists()
        if has_full:
            return None

        return set(
            EditGrantItem.objects.filter(grant__in=active_grants).values_list("enrollment_id", flat=True)
        )

    def _valid_achievements_for_assignment_period(
        self,
        *,
        teacher_assignment: TeacherAssignment,
        period: Period,
    ):
        base_achievements = Achievement.objects.filter(
            academic_load=teacher_assignment.academic_load,
            period=period,
        )

        group_achievements = base_achievements.filter(group=teacher_assignment.group)
        if group_achievements.exists():
            return group_achievements
        return base_achievements.filter(group__isnull=True)

    def _get_preschool_labels(self, *, academic_year_id: int):
        preschool = EvaluationScale.objects.filter(
            academic_year_id=academic_year_id,
            scale_type="QUALITATIVE",
            applies_to_level="PRESCHOOL",
        )

        # Backward-compatible fallback: if scales were created before applies_to_level
        # existed (or not configured), allow QUALITATIVE scales with applies_to_level=NULL.
        base = preschool
        if not base.exists():
            base = EvaluationScale.objects.filter(
                academic_year_id=academic_year_id,
                scale_type="QUALITATIVE",
                applies_to_level__isnull=True,
            )

        defaults = base.filter(is_default=True)
        qs = defaults if defaults.exists() else base
        return qs.order_by("order", "id")

    @action(detail=False, methods=["get"], url_path="labels")
    def labels(self, request):
        teacher_resp = self._require_teacher(request)
        if teacher_resp is not None:
            return teacher_resp

        academic_year_id_raw = request.query_params.get("academic_year")
        period_id_raw = request.query_params.get("period")

        academic_year_id = None
        if academic_year_id_raw not in (None, ""):
            try:
                academic_year_id = int(str(academic_year_id_raw))
            except Exception:
                return Response({"detail": "academic_year inválido"}, status=status.HTTP_400_BAD_REQUEST)
        elif period_id_raw not in (None, ""):
            try:
                p = Period.objects.only("id", "academic_year_id").get(id=int(str(period_id_raw)))
                academic_year_id = int(p.academic_year_id)
            except Period.DoesNotExist:
                return Response({"detail": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)
            except Exception:
                return Response({"detail": "period inválido"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"detail": "academic_year o period es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        labels_qs = self._get_preschool_labels(academic_year_id=academic_year_id)
        return Response({"results": PreschoolGradebookLabelSerializer(labels_qs, many=True).data})

    @action(detail=False, methods=["get"], url_path="available")
    def available(self, request):
        teacher_resp = self._require_teacher(request)
        if teacher_resp is not None:
            return teacher_resp

        period_id = request.query_params.get("period")
        if not period_id:
            return Response({"error": "period es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        tas = (
            TeacherAssignment.objects.filter(
                academic_year_id=period.academic_year_id,
                teacher=request.user,
                group__grade__level__level_type="PRESCHOOL",
            )
            .select_related(
                "group",
                "group__grade",
                "academic_load",
                "academic_load__subject",
            )
            .order_by("group__grade__name", "group__name", "academic_load__subject__name")
        )

        from students.models import Enrollment

        items = []
        for ta in tas:
            enrollments_qs = Enrollment.objects.filter(
                academic_year_id=ta.academic_year_id,
                group_id=ta.group_id,
                status="ACTIVE",
            )
            students_count = enrollments_qs.count()

            achievements_qs = self._valid_achievements_for_assignment_period(
                teacher_assignment=ta,
                period=period,
            )
            achievement_ids = list(achievements_qs.values_list("id", flat=True))
            achievements_count = len(achievement_ids)

            total = students_count * achievements_count

            gradesheet_id = (
                GradeSheet.objects.filter(teacher_assignment_id=ta.id, period_id=period.id)
                .values_list("id", flat=True)
                .first()
            )

            filled = 0
            if gradesheet_id and total > 0 and achievement_ids:
                filled = (
                    AchievementGrade.objects.filter(
                        gradesheet_id=gradesheet_id,
                        enrollment__in=enrollments_qs,
                        achievement_id__in=achievement_ids,
                        qualitative_scale__isnull=False,
                    )
                    .only("id")
                    .count()
                )

            percent = int(round((filled / total) * 100)) if total > 0 else 0
            is_complete = total > 0 and filled >= total

            items.append(
                {
                    "teacher_assignment_id": ta.id,
                    "group_id": ta.group_id,
                    "group_name": ta.group.name,
                    "grade_id": ta.group.grade_id,
                    "grade_name": ta.group.grade.name,
                    "academic_load_id": ta.academic_load_id,
                    "subject_name": ta.academic_load.subject.name if ta.academic_load_id else None,
                    "period": {"id": period.id, "name": period.name, "is_closed": period.is_closed},
                    "students_count": students_count,
                    "achievements_count": achievements_count,
                    "completion": {
                        "filled": filled,
                        "total": total,
                        "percent": percent,
                        "is_complete": is_complete,
                    },
                }
            )

        return Response({"results": items})

    @action(detail=False, methods=["get"], url_path="gradebook")
    def gradebook(self, request):
        teacher_resp = self._require_teacher(request)
        if teacher_resp is not None:
            return teacher_resp

        teacher_assignment_id = request.query_params.get("teacher_assignment")
        period_id = request.query_params.get("period")
        if not teacher_assignment_id or not period_id:
            return Response(
                {"error": "teacher_assignment y period son requeridos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            teacher_assignment = self._get_teacher_assignment(
                user=request.user,
                teacher_assignment_id=int(teacher_assignment_id),
            )
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        preschool_resp = self._ensure_preschool_assignment(teacher_assignment)
        if preschool_resp is not None:
            return preschool_resp

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        if getattr(gradesheet, "grading_mode", None) != GradeSheet.GRADING_MODE_QUALITATIVE:
            gradesheet.grading_mode = GradeSheet.GRADING_MODE_QUALITATIVE
            gradesheet.save(update_fields=["grading_mode", "updated_at"])

        achievements = (
            self._valid_achievements_for_assignment_period(
                teacher_assignment=teacher_assignment,
                period=period,
            )
            .select_related("dimension")
            .order_by("id")
        )

        from students.models import Enrollment

        enrollments = (
            Enrollment.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                group_id=teacher_assignment.group_id,
                status="ACTIVE",
            )
            .select_related("student__user")
            .order_by("student__user__last_name", "student__user__first_name")
        )

        existing_grades = AchievementGrade.objects.filter(
            gradesheet=gradesheet,
            enrollment__in=enrollments,
            achievement__in=achievements,
        ).only("enrollment_id", "achievement_id", "qualitative_scale_id")

        qualitative_by_cell = {
            (g.enrollment_id, g.achievement_id): g.qualitative_scale_id for g in existing_grades
        }

        achievement_payload = [
            {
                "id": a.id,
                "description": a.description,
                "dimension": a.dimension_id,
                "dimension_name": a.dimension.name if a.dimension else None,
                "percentage": a.percentage,
            }
            for a in achievements
        ]

        student_payload = [
            {
                "enrollment_id": e.id,
                "student_id": e.student_id,
                "student_name": e.student.user.get_full_name(),
            }
            for e in enrollments
        ]

        cell_payload = [
            {
                "enrollment": e.id,
                "achievement": a.id,
                "qualitative_scale": qualitative_by_cell.get((e.id, a.id)),
            }
            for e in enrollments
            for a in achievements
        ]

        labels_qs = self._get_preschool_labels(academic_year_id=teacher_assignment.academic_year_id)

        payload = {
            "gradesheet": GradeSheetSerializer(gradesheet).data,
            "period": {"id": period.id, "name": period.name, "is_closed": period.is_closed},
            "teacher_assignment": {
                "id": teacher_assignment.id,
                "group": teacher_assignment.group_id,
                "academic_load": teacher_assignment.academic_load_id,
            },
            "achievements": achievement_payload,
            "students": student_payload,
            "cells": cell_payload,
            "labels": PreschoolGradebookLabelSerializer(labels_qs, many=True).data,
        }

        return Response(payload)

    @action(detail=False, methods=["post"], url_path="bulk-upsert")
    def bulk_upsert(self, request):
        teacher_resp = self._require_teacher(request)
        if teacher_resp is not None:
            return teacher_resp

        serializer = PreschoolGradebookBulkUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        teacher_assignment_id = serializer.validated_data["teacher_assignment"]
        period_id = serializer.validated_data["period"]
        grades = serializer.validated_data["grades"]

        try:
            teacher_assignment = self._get_teacher_assignment(
                user=request.user,
                teacher_assignment_id=int(teacher_assignment_id),
            )
        except TeacherAssignment.DoesNotExist:
            return Response({"error": "TeacherAssignment no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        preschool_resp = self._ensure_preschool_assignment(teacher_assignment)
        if preschool_resp is not None:
            return preschool_resp

        try:
            period = Period.objects.select_related("academic_year").get(id=int(period_id))
        except Period.DoesNotExist:
            return Response({"error": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.is_closed:
            return Response(
                {"error": "El periodo está cerrado; no se pueden registrar notas."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if period.academic_year_id != teacher_assignment.academic_year_id:
            return Response(
                {"error": "El periodo no corresponde al año lectivo de la asignación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        current_resp = self._enforce_teacher_current_period(user=request.user, period=period)
        if current_resp is not None:
            return current_resp

        gradesheet, _ = GradeSheet.objects.get_or_create(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        if getattr(gradesheet, "grading_mode", None) != GradeSheet.GRADING_MODE_QUALITATIVE:
            gradesheet.grading_mode = GradeSheet.GRADING_MODE_QUALITATIVE
            gradesheet.save(update_fields=["grading_mode", "updated_at"])

        from students.models import Enrollment

        valid_enrollments = set(
            Enrollment.objects.filter(
                academic_year_id=teacher_assignment.academic_year_id,
                group_id=teacher_assignment.group_id,
            ).values_list("id", flat=True)
        )

        achievements_qs = self._valid_achievements_for_assignment_period(
            teacher_assignment=teacher_assignment,
            period=period,
        )
        valid_achievement_ids = set(achievements_qs.values_list("id", flat=True))

        labels_qs = self._get_preschool_labels(academic_year_id=teacher_assignment.academic_year_id)
        label_by_id = {int(s.id): s for s in labels_qs}

        allowed_enrollment_ids = self._teacher_allowed_enrollment_ids_after_deadline(
            user=request.user,
            period=period,
            teacher_assignment=teacher_assignment,
        )

        blocked = []
        allowed_by_cell: dict[tuple[int, int], AchievementGrade] = {}
        for g in grades:
            enrollment_id = int(g["enrollment"])
            achievement_id = int(g["achievement"])
            qualitative_scale_id = g.get("qualitative_scale")

            if enrollment_id not in valid_enrollments:
                return Response(
                    {"error": f"Enrollment inválido: {enrollment_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if achievement_id not in valid_achievement_ids:
                return Response(
                    {"error": f"Achievement inválido: {achievement_id}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if qualitative_scale_id is not None:
                try:
                    qualitative_scale_id = int(qualitative_scale_id)
                except Exception:
                    return Response(
                        {"error": "qualitative_scale inválido"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if qualitative_scale_id not in label_by_id:
                    return Response(
                        {"error": f"Etiqueta cualitativa inválida: {qualitative_scale_id}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if allowed_enrollment_ids is not None and enrollment_id not in allowed_enrollment_ids:
                blocked.append(
                    {
                        "enrollment": enrollment_id,
                        "achievement": achievement_id,
                        "reason": "EDIT_WINDOW_CLOSED",
                    }
                )
                continue

            internal_score = None
            if qualitative_scale_id is not None:
                internal_score = getattr(label_by_id.get(int(qualitative_scale_id)), "internal_numeric_value", None)

            allowed_by_cell[(enrollment_id, achievement_id)] = AchievementGrade(
                gradesheet=gradesheet,
                enrollment_id=enrollment_id,
                achievement_id=achievement_id,
                qualitative_scale_id=qualitative_scale_id,
                score=internal_score,
            )

        to_upsert = list(allowed_by_cell.values())
        if to_upsert:
            with transaction.atomic():
                AchievementGrade.objects.bulk_create(
                    to_upsert,
                    update_conflicts=True,
                    unique_fields=["gradesheet", "enrollment", "achievement"],
                    update_fields=["qualitative_scale", "score", "updated_at"],
                )

        return Response(
            {
                "requested": len(grades),
                "updated": len(to_upsert),
                "blocked": blocked,
            },
            status=status.HTTP_200_OK,
        )


def _is_admin_like(user) -> bool:
    role = getattr(user, "role", None)
    return role in {"SUPERADMIN", "ADMIN", "COORDINATOR"}


class EditRequestViewSet(viewsets.ModelViewSet):
    queryset = EditRequest.objects.all().select_related(
        "requested_by",
        "period",
        "teacher_assignment",
        "decided_by",
    )
    serializer_class = EditRequestSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["status", "scope", "period"]

    def get_queryset(self):
        qs = super().get_queryset().prefetch_related("items")
        user = getattr(self.request, "user", None)
        if not user or not user.is_authenticated:
            return qs.none()
        if getattr(user, "role", None) == "TEACHER":
            return qs.filter(requested_by=user)
        if _is_admin_like(user):
            return qs
        return qs.none()

    def create(self, request, *args, **kwargs):
        user = getattr(request, "user", None)
        if not user or getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Solo docentes pueden crear solicitudes."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        obj: EditRequest = serializer.save()

        # Notify admin-like users that there is a pending request
        try:
            from notifications.services import admin_like_users_qs, notify_users

            scope_label = "Calificaciones" if obj.scope == EditRequest.SCOPE_GRADES else "Planeación"
            teacher_name = getattr(obj.requested_by, "get_full_name", lambda: "")() or getattr(
                obj.requested_by, "username", "Docente"
            )
            title = f"Solicitud de edición pendiente ({scope_label})"
            body = f"{teacher_name} envió una solicitud para el periodo {obj.period_id}."
            url = (
                "/edit-requests/grades"
                if obj.scope == EditRequest.SCOPE_GRADES
                else "/edit-requests/planning"
            )
            notify_users(
                recipients=admin_like_users_qs(),
                type="EDIT_REQUEST_PENDING",
                title=title,
                body=body,
                url=url,
                dedupe_key=f"EDIT_REQUEST_PENDING:teacher={obj.requested_by_id}:scope={obj.scope}:period={obj.period_id}",
                dedupe_within_seconds=300,
            )
        except Exception:
            # Notifications must not break core flows
            pass

        return obj

    def update(self, request, *args, **kwargs):
        user = getattr(request, "user", None)
        if user and getattr(user, "role", None) == "TEACHER":
            return Response({"detail": "No puedes modificar una solicitud."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        user = getattr(request, "user", None)
        if user and getattr(user, "role", None) == "TEACHER":
            return Response({"detail": "No puedes eliminar una solicitud."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="my")
    def my(self, request):
        user = getattr(request, "user", None)
        if not user or getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Solo docentes."}, status=status.HTTP_403_FORBIDDEN)
        qs = self.get_queryset().order_by("-created_at")
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        user = getattr(request, "user", None)
        if not user or not _is_admin_like(user):
            return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)

        edit_request: EditRequest = self.get_object()
        if edit_request.status != EditRequest.STATUS_PENDING:
            return Response({"detail": "La solicitud ya fue decidida."}, status=status.HTTP_400_BAD_REQUEST)

        decision = EditRequestDecisionSerializer(data=request.data)
        decision.is_valid(raise_exception=True)
        valid_until = decision.validated_data.get("valid_until") or edit_request.requested_until
        if valid_until is None:
            return Response({"valid_until": "Es requerido para aprobar."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            edit_request.status = EditRequest.STATUS_APPROVED
            edit_request.decided_by = user
            edit_request.decided_at = timezone.now()
            edit_request.decision_note = decision.validated_data.get("decision_note", "")
            edit_request.save(update_fields=["status", "decided_by", "decided_at", "decision_note", "updated_at"])

            grant = EditGrant.objects.create(
                scope=edit_request.scope,
                grant_type=edit_request.request_type,
                granted_to=edit_request.requested_by,
                period=edit_request.period,
                teacher_assignment=edit_request.teacher_assignment,
                valid_until=valid_until,
                created_by=user,
                source_request=edit_request,
            )

            if edit_request.request_type == EditRequest.TYPE_PARTIAL:
                items = list(edit_request.items.values_list("enrollment_id", flat=True))
                EditGrantItem.objects.bulk_create(
                    [EditGrantItem(grant=grant, enrollment_id=eid) for eid in items]
                )

            # Notify teacher
            try:
                from notifications.services import create_notification

                scope_label = "Calificaciones" if edit_request.scope == EditRequest.SCOPE_GRADES else "Planeación"
                title = f"Solicitud aprobada ({scope_label})"
                url = "/grades" if edit_request.scope == EditRequest.SCOPE_GRADES else "/planning"
                note = (edit_request.decision_note or "").strip()
                body = f"Aprobada hasta: {valid_until}." + (f"\nNota: {note}" if note else "")
                create_notification(
                    recipient=edit_request.requested_by,
                    type="EDIT_REQUEST_APPROVED",
                    title=title,
                    body=body,
                    url=url,
                )
            except Exception:
                pass

        return Response({"detail": "Solicitud aprobada.", "grant_id": grant.id}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        user = getattr(request, "user", None)
        if not user or not _is_admin_like(user):
            return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)

        edit_request: EditRequest = self.get_object()
        if edit_request.status != EditRequest.STATUS_PENDING:
            return Response({"detail": "La solicitud ya fue decidida."}, status=status.HTTP_400_BAD_REQUEST)

        decision = EditRequestDecisionSerializer(data=request.data)
        decision.is_valid(raise_exception=True)

        edit_request.status = EditRequest.STATUS_REJECTED
        edit_request.decided_by = user
        edit_request.decided_at = timezone.now()
        edit_request.decision_note = decision.validated_data.get("decision_note", "")
        edit_request.save(update_fields=["status", "decided_by", "decided_at", "decision_note", "updated_at"])

        # Notify teacher
        try:
            from notifications.services import create_notification

            scope_label = "Calificaciones" if edit_request.scope == EditRequest.SCOPE_GRADES else "Planeación"
            title = f"Solicitud rechazada ({scope_label})"
            url = "/grades" if edit_request.scope == EditRequest.SCOPE_GRADES else "/planning"
            note = (edit_request.decision_note or "").strip()
            body = (f"Nota: {note}" if note else "")
            create_notification(
                recipient=edit_request.requested_by,
                type="EDIT_REQUEST_REJECTED",
                title=title,
                body=body,
                url=url,
            )
        except Exception:
            pass

        return Response({"detail": "Solicitud rechazada."}, status=status.HTTP_200_OK)


class EditGrantViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = EditGrant.objects.all().select_related(
        "granted_to",
        "period",
        "teacher_assignment",
        "created_by",
        "source_request",
    )
    serializer_class = EditGrantSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["scope", "period", "teacher_assignment", "granted_to"]

    def get_queryset(self):
        qs = super().get_queryset().prefetch_related("items")
        user = getattr(self.request, "user", None)
        if not user or not user.is_authenticated:
            return qs.none()
        if getattr(user, "role", None) == "TEACHER":
            return qs.filter(granted_to=user)
        if _is_admin_like(user):
            return qs
        return qs.none()

    @action(detail=False, methods=["get"], url_path="my")
    def my(self, request):
        user = getattr(request, "user", None)
        if not user or getattr(user, "role", None) != "TEACHER":
            return Response({"detail": "Solo docentes."}, status=status.HTTP_403_FORBIDDEN)
        qs = self.get_queryset().order_by("-created_at")
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)
