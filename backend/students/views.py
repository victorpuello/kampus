from rest_framework import viewsets, filters, serializers
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from django.http import FileResponse, HttpResponse
from django.shortcuts import redirect, render
from django.views import View
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Q, Sum
from django.contrib.auth import get_user_model
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse
from urllib.parse import urlsplit, urlunsplit
import csv
import base64
import io
import os
import random
import tempfile
import re
import unicodedata
import uuid as py_uuid
from decimal import Decimal
from datetime import date, datetime
from academic.models import AcademicLoad, AcademicYear, Grade, Group
from academic.models import (
    Achievement,
    AchievementGrade,
    Dimension,
    EvaluationScale,
    GradeSheet,
    Period,
    TeacherAssignment,
)
from academic.grading import (
    DEFAULT_EMPTY_SCORE,
    final_grade_from_dimensions,
    match_scale,
    weighted_average,
)
from core.models import Institution
from core.models import Campus
from .models import (
    CertificateIssue,
    ConditionalPromotionPlan,
    Enrollment,
    FamilyMember,
    ObserverAnnotation,
    Student,
    StudentDocument,
    StudentNovelty,
)
from rest_framework.permissions import AllowAny, IsAuthenticated, BasePermission

from users.permissions import IsAdministrativeStaff

from audit.services import log_event

try:
    import qrcode  # type: ignore
except Exception:  # pragma: no cover
    qrcode = None

from .serializers import (
    StudentSerializer,
    FamilyMemberSerializer,
    EnrollmentSerializer,
    StudentNoveltySerializer,
    StudentDocumentSerializer,
    ObserverAnnotationSerializer,
)
from .pagination import StudentPagination, EnrollmentPagination
from core.permissions import HasDjangoPermission, KampusModelPermissions
import traceback

from .filters import StudentFilter

from .academic_period_report import compute_certificate_studies_rows, generate_academic_period_report_pdf

from reports.weasyprint_utils import PDF_BASE_CSS, weasyprint_url_fetcher
from students.reports import sort_enrollments_for_enrollment_list

User = get_user_model()


def _director_student_ids(user):
    """Student IDs the given teacher can manage as group director.

    Uses ACTIVE academic year when available. If no ACTIVE year exists,
    falls back to any group where the user is director.
    """

    if user is None or getattr(user, 'role', None) != 'TEACHER':
        return set()

    active_year = AcademicYear.objects.filter(status='ACTIVE').first()
    directed_groups = Group.objects.filter(director=user)
    if active_year:
        directed_groups = directed_groups.filter(academic_year=active_year)

    if not directed_groups.exists():
        return set()

    directed_student_ids = (
        Enrollment.objects.filter(
            group__in=directed_groups,
            status='ACTIVE',
        )
        .values_list('student_id', flat=True)
        .distinct()
    )
    return set(directed_student_ids)


def _teacher_managed_student_ids(user):
    """Student IDs a teacher can manage (director or assigned teacher).

    Uses ACTIVE academic year when available. If no ACTIVE year exists,
    falls back to any group where the user is director or has a TeacherAssignment.
    """

    if user is None or getattr(user, "role", None) != "TEACHER":
        return set()

    active_year = AcademicYear.objects.filter(status="ACTIVE").first()

    directed_groups = Group.objects.filter(director=user)
    assigned_group_ids = TeacherAssignment.objects.filter(teacher=user).values_list("group_id", flat=True)
    assigned_groups = Group.objects.filter(id__in=assigned_group_ids)

    if active_year:
        directed_groups = directed_groups.filter(academic_year=active_year)
        assigned_groups = assigned_groups.filter(academic_year=active_year)

    allowed_groups = directed_groups | assigned_groups
    if not allowed_groups.exists():
        return set()

    year_filter = Q(group__in=allowed_groups, status="ACTIVE")
    if active_year:
        year_filter &= Q(academic_year=active_year)

    managed_student_ids = (
        Enrollment.objects.filter(year_filter)
        .values_list("student_id", flat=True)
        .distinct()
    )
    return set(managed_student_ids)


class IsTeacherDirectorOfStudent(BasePermission):
    """Teacher can access only students from their directed groups."""

    def has_permission(self, request, view):
        user = getattr(request, 'user', None)
        return bool(user and user.is_authenticated and getattr(user, 'role', None) == 'TEACHER')

    def has_object_permission(self, request, view, obj):
        user = getattr(request, 'user', None)
        if not user or getattr(user, 'role', None) != 'TEACHER':
            return False
        return getattr(obj, 'pk', None) in _director_student_ids(user)


class IsTeacherAssignedOrDirectorOfStudent(BasePermission):
    """Teacher can access only students from their directed OR assigned groups."""

    def has_permission(self, request, view):
        user = getattr(request, "user", None)
        return bool(user and user.is_authenticated and getattr(user, "role", None) == "TEACHER")

    def has_object_permission(self, request, view, obj):
        user = getattr(request, "user", None)
        if not user or getattr(user, "role", None) != "TEACHER":
            return False
        return getattr(obj, "pk", None) in _teacher_managed_student_ids(user)


class IsTeacherDirectorOfRelatedStudent(BasePermission):
    """Teacher can access objects that point to a directed student.

    Works for models with a `student_id` attribute.
    For create actions, expects `student` in request.data.
    """

    def has_permission(self, request, view):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated and getattr(user, 'role', None) == 'TEACHER'):
            return False

        if getattr(view, 'action', None) == 'create':
            raw = request.data.get('student')
            try:
                student_id = int(raw)
            except Exception:
                return False
            return student_id in _director_student_ids(user)

        return True

    def has_object_permission(self, request, view, obj):
        user = getattr(request, 'user', None)
        if not user or getattr(user, 'role', None) != 'TEACHER':
            return False
        student_id = getattr(obj, 'student_id', None)
        return student_id in _director_student_ids(user)


class IsTeacherAssignedOrDirectorOfRelatedStudent(BasePermission):
    """Teacher can access objects that point to a student in their directed/assigned groups."""

    def has_permission(self, request, view):
        user = getattr(request, "user", None)
        if not (user and user.is_authenticated and getattr(user, "role", None) == "TEACHER"):
            return False

        if getattr(view, "action", None) == "create":
            raw = request.data.get("student")
            try:
                student_id = int(raw)
            except Exception:
                return False
            return student_id in _teacher_managed_student_ids(user)

        return True

    def has_object_permission(self, request, view, obj):
        user = getattr(request, "user", None)
        if not user or getattr(user, "role", None) != "TEACHER":
            return False
        student_id = getattr(obj, "student_id", None)
        return student_id in _teacher_managed_student_ids(user)



class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.select_related("user").all().order_by("user__last_name", "user__first_name", "user__id")
    serializer_class = StudentSerializer
    permission_classes = [KampusModelPermissions]
    parser_classes = (JSONParser, FormParser, MultiPartParser)
    pagination_class = StudentPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = StudentFilter
    search_fields = ['user__first_name', 'user__last_name', 'document_number']

    def get_permissions(self):
        # Directors (teachers) can manage students in their directed groups,
        # regardless of Django model-permissions.
        if getattr(self.request.user, 'role', None) == 'TEACHER':
            action = getattr(self, 'action', None)
            if action in {'list', 'retrieve'}:
                return [IsAuthenticated(), IsTeacherDirectorOfStudent()]
            if action in {'update', 'partial_update'}:
                return [IsAuthenticated(), IsTeacherDirectorOfStudent()]
        return super().get_permissions()

    def get_queryset(self):
        qs = Student.objects.select_related("user").all().order_by("user__last_name", "user__first_name", "user__id")
        user = getattr(self.request, 'user', None)

        exclude_year_raw = self.request.query_params.get('exclude_active_enrollment_year')
        exclude_year_id = None
        if exclude_year_raw is not None and str(exclude_year_raw).strip() != '':
            try:
                exclude_year_id = int(exclude_year_raw)
            except (TypeError, ValueError):
                exclude_year_id = None

        if user is not None and getattr(user, 'role', None) in {'PARENT', 'STUDENT'}:
            return qs.none()

        if user is not None and getattr(user, 'role', None) == 'TEACHER':
            allowed_ids = _director_student_ids(user)
            if not allowed_ids:
                return qs.none()
            qs = qs.filter(pk__in=allowed_ids)

        if exclude_year_id is not None:
            from students.models import Enrollment

            excluded_student_ids = (
                Enrollment.objects.filter(
                    academic_year_id=exclude_year_id,
                    status='ACTIVE',
                )
                .values_list('student_id', flat=True)
                .distinct()
            )
            qs = qs.exclude(pk__in=excluded_student_ids)

        # Annotate current enrollment status for UI list.
        # Prefer the enrollment from the ACTIVE academic year when present; otherwise use the latest enrollment.
        try:
            from django.db.models import OuterRef, Subquery
            from django.db.models.functions import Coalesce
            from students.models import Enrollment

            active_year = AcademicYear.objects.filter(status='ACTIVE').first()

            latest_status_sq = (
                Enrollment.objects.filter(student_id=OuterRef('pk'))
                .order_by('-academic_year__year', '-id')
                .values('status')[:1]
            )

            latest_grade_ordinal_sq = (
                Enrollment.objects.filter(student_id=OuterRef('pk'))
                .order_by('-academic_year__year', '-id')
                .values('grade__ordinal')[:1]
            )
            latest_grade_name_sq = (
                Enrollment.objects.filter(student_id=OuterRef('pk'))
                .order_by('-academic_year__year', '-id')
                .values('grade__name')[:1]
            )

            if active_year is not None:
                active_year_status_sq = (
                    Enrollment.objects.filter(student_id=OuterRef('pk'), academic_year_id=active_year.id)
                    .values('status')[:1]
                )
                active_year_grade_ordinal_sq = (
                    Enrollment.objects.filter(student_id=OuterRef('pk'), academic_year_id=active_year.id)
                    .values('grade__ordinal')[:1]
                )
                active_year_grade_name_sq = (
                    Enrollment.objects.filter(student_id=OuterRef('pk'), academic_year_id=active_year.id)
                    .values('grade__name')[:1]
                )

                qs = qs.annotate(
                    current_enrollment_status=Coalesce(Subquery(active_year_status_sq), Subquery(latest_status_sq)),
                    current_grade_ordinal=Coalesce(Subquery(active_year_grade_ordinal_sq), Subquery(latest_grade_ordinal_sq)),
                    current_grade_name=Coalesce(Subquery(active_year_grade_name_sq), Subquery(latest_grade_name_sq)),
                )
            else:
                qs = qs.annotate(
                    current_enrollment_status=Subquery(latest_status_sq),
                    current_grade_ordinal=Subquery(latest_grade_ordinal_sq),
                    current_grade_name=Subquery(latest_grade_name_sq),
                )
        except Exception:
            # Best-effort: keep endpoint working even if annotation fails.
            pass

        return qs.order_by("user__last_name", "user__first_name", "user__id")

    def list(self, request, *args, **kwargs):
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            print("ERROR IN STUDENT LIST:")
            traceback.print_exc()
            return Response({"error": str(e), "traceback": traceback.format_exc()}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=["get"], url_path="observer-report")
    def observer_report(self, request, pk=None):
        """Reporte consolidado del Observador/Ficha del estudiante (JSON).

        Diseñado para consumo por el frontend y posterior impresión desde el navegador.
        """

        student: Student = self.get_object()

        # Preferir matrícula activa para resolver sede/institución.
        current_enrollment = (
            Enrollment.objects.select_related(
                "academic_year",
                "grade",
                "group",
                "campus",
                "campus__institution",
            )
            .filter(student=student, status="ACTIVE")
            .order_by("-academic_year__year", "-id")
            .first()
        )

        institution = None
        campus = None
        if current_enrollment is not None:
            campus = getattr(current_enrollment, "campus", None)
            institution = getattr(campus, "institution", None) if campus else None

        if institution is None:
            institution = Institution.objects.first() or Institution()

        def _abs_media_url(rel_or_abs: str | None) -> str | None:
            if not rel_or_abs:
                return None
            try:
                return request.build_absolute_uri(rel_or_abs)
            except Exception:
                return rel_or_abs

        logo_url = None
        try:
            if getattr(institution, "logo", None) and getattr(institution.logo, "url", None):
                logo_url = _abs_media_url(institution.logo.url)
        except Exception:
            logo_url = None

        student_photo_url = None
        try:
            if getattr(student, "photo", None) and getattr(student.photo, "url", None):
                student_photo_url = _abs_media_url(student.photo.url)
        except Exception:
            student_photo_url = None

        family_members = list(
            FamilyMember.objects.filter(student=student)
            .select_related("user")
            .order_by("-is_main_guardian", "id")
        )

        enrollments = list(
            Enrollment.objects.select_related("academic_year", "grade", "group", "campus")
            .filter(student=student)
            .order_by("-academic_year__year", "-id")
        )

        # === Disciplina / Observador ===
        # Respetar reglas de visibilidad similares a DisciplineCaseViewSet.
        user = getattr(request, "user", None)
        role = getattr(user, "role", None)

        from discipline.models import DisciplineCase  # noqa: PLC0415

        cases_qs = (
            DisciplineCase.objects.select_related(
                "enrollment",
                "enrollment__academic_year",
                "enrollment__grade",
                "enrollment__group",
                "created_by",
            )
            .prefetch_related("events")
            .filter(student=student)
            .order_by("-occurred_at", "-id")
        )

        if role == "TEACHER":
            active_year = AcademicYear.objects.filter(status="ACTIVE").first()
            directed_groups = Group.objects.filter(director=user)
            if active_year:
                directed_groups = directed_groups.filter(academic_year=active_year)

            if active_year:
                assigned_group_ids = set(
                    TeacherAssignment.objects.filter(teacher=user, academic_year=active_year).values_list(
                        "group_id", flat=True
                    )
                )
            else:
                assigned_group_ids = set(
                    TeacherAssignment.objects.filter(teacher=user).values_list("group_id", flat=True)
                )

            allowed_group_ids = set(directed_groups.values_list("id", flat=True)) | assigned_group_ids
            if not allowed_group_ids:
                cases_qs = cases_qs.none()
            else:
                cases_qs = cases_qs.filter(enrollment__group_id__in=allowed_group_ids).distinct()

        elif role in {"ADMIN", "SUPERADMIN", "COORDINATOR"}:
            cases_qs = cases_qs

        elif role == "PARENT":
            # Solo si realmente es acudiente del estudiante
            is_guardian = FamilyMember.objects.filter(student=student, user=user).exists()
            if not is_guardian:
                cases_qs = cases_qs.none()

        else:
            cases_qs = cases_qs.none()

        cases = list(cases_qs)

        def _full_name(u) -> str:
            try:
                return u.get_full_name() or getattr(u, "username", "") or ""
            except Exception:
                return ""

        def _fmt_date(value) -> str | None:
            if not value:
                return None
            try:
                # date or datetime
                return value.isoformat()
            except Exception:
                return str(value)

        discipline_entries = []
        for case in cases:
            enrollment = getattr(case, "enrollment", None)
            academic_year = None
            grade_name = ""
            group_name = ""
            try:
                academic_year = getattr(getattr(enrollment, "academic_year", None), "year", None)
                grade_name = getattr(getattr(enrollment, "grade", None), "name", "") or ""
                group_name = getattr(getattr(enrollment, "group", None), "name", "") or ""
            except Exception:
                pass

            events_out = []
            try:
                for ev in list(case.events.all()):
                    events_out.append(
                        {
                            "id": ev.id,
                            "event_type": ev.event_type,
                            "text": ev.text,
                            "created_at": _fmt_date(ev.created_at),
                            "created_by_name": _full_name(getattr(ev, "created_by", None)),
                        }
                    )
            except Exception:
                events_out = []

            discipline_entries.append(
                {
                    "id": case.id,
                    "occurred_at": _fmt_date(case.occurred_at),
                    "location": case.location,
                    "manual_severity": case.manual_severity,
                    "law_1620_type": case.law_1620_type,
                    "status": case.status,
                    "academic_year": academic_year,
                    "grade_name": grade_name,
                    "group_name": group_name,
                    "narrative": case.narrative,
                    "decision_text": case.decision_text,
                    "created_by_name": _full_name(getattr(case, "created_by", None)),
                    "created_at": _fmt_date(case.created_at),
                    "events": events_out,
                }
            )

        observer_number = f"{getattr(student, 'pk', 0):010d}"

        out = {
            "observer_number": observer_number,
            "generated_at": datetime.now().isoformat(),
            "institution": {
                "name": getattr(institution, "name", "") or "",
                "dane_code": getattr(institution, "dane_code", "") or "",
                "nit": getattr(institution, "nit", "") or "",
                "pdf_header_line1": getattr(institution, "pdf_header_line1", "") or "",
                "pdf_header_line2": getattr(institution, "pdf_header_line2", "") or "",
                "pdf_header_line3": getattr(institution, "pdf_header_line3", "") or "",
                "logo_url": logo_url,
            },
            "campus": {
                "name": getattr(campus, "name", "") if campus else "",
                "municipality": getattr(campus, "municipality", "") if campus else "",
            },
            "student": {
                "id": student.pk,
                "full_name": student.user.get_full_name(),
                "first_name": getattr(student.user, "first_name", "") or "",
                "last_name": getattr(student.user, "last_name", "") or "",
                "document_type": student.document_type,
                "document_number": student.document_number,
                "birth_date": _fmt_date(student.birth_date),
                "place_of_issue": student.place_of_issue,
                "neighborhood": student.neighborhood,
                "address": student.address,
                "blood_type": student.blood_type,
                "stratum": student.stratum,
                "sisben_score": student.sisben_score,
                "photo_url": student_photo_url,
            },
            "family_members": [
                {
                    "id": fm.id,
                    "relationship": fm.relationship,
                    "full_name": fm.full_name,
                    "document_number": fm.document_number,
                    "phone": fm.phone,
                    "email": fm.email,
                    "is_main_guardian": fm.is_main_guardian,
                }
                for fm in family_members
            ],
            "enrollments": [
                {
                    "id": e.id,
                    "academic_year": getattr(getattr(e, "academic_year", None), "year", None),
                    "grade_name": getattr(getattr(e, "grade", None), "name", "") or "",
                    "group_name": getattr(getattr(e, "group", None), "name", "") or "",
                    "campus_name": getattr(getattr(e, "campus", None), "name", "") or "",
                    "status": e.status,
                    "final_status": e.final_status,
                    "enrolled_at": _fmt_date(e.enrolled_at),
                }
                for e in enrollments
            ],
            "discipline_entries": discipline_entries,
            "observer_annotations": [
                {
                    "id": a.id,
                    "period": {
                        "id": a.period_id,
                        "name": getattr(getattr(a, "period", None), "name", "") if getattr(a, "period", None) else "",
                        "academic_year": getattr(getattr(getattr(a, "period", None), "academic_year", None), "year", None)
                        if getattr(a, "period", None)
                        else None,
                        "is_closed": bool(getattr(getattr(a, "period", None), "is_closed", False)) if getattr(a, "period", None) else False,
                    }
                    if a.period_id
                    else None,
                    "annotation_type": a.annotation_type,
                    "title": a.title,
                    "text": a.text,
                    "commitments": a.commitments,
                    "commitment_due_date": _fmt_date(a.commitment_due_date),
                    "commitment_responsible": a.commitment_responsible,
                    "is_automatic": bool(a.is_automatic),
                    "created_at": _fmt_date(a.created_at),
                    "updated_at": _fmt_date(a.updated_at),
                    "created_by_name": _full_name(getattr(a, "created_by", None)),
                    "updated_by_name": _full_name(getattr(a, "updated_by", None)),
                }
                for a in ObserverAnnotation.objects.select_related("period", "period__academic_year", "created_by", "updated_by")
                .filter(student=student, is_deleted=False)
                .order_by("-created_at", "-id")
            ],
        }

        return Response(out)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para crear estudiantes."}, status=status.HTTP_403_FORBIDDEN)

        print("Recibiendo datos para crear estudiante:", request.data)
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print("VALIDATION ERRORS:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            print("ERROR IN STUDENT CREATE:")
            traceback.print_exc()
            # If it's a validation error raised by us, return 400
            if "ValidationError" in str(type(e)):
                 return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para editar estudiantes."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para editar estudiantes."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para eliminar estudiantes."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    def _normalize_header(self, value: str) -> str:
        if value is None:
            return ''
        text = str(value).strip()
        text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
        text = text.lower()
        text = re.sub(r'[^a-z0-9]+', '_', text)
        return text.strip('_')

    def _parse_bool(self, value):
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        s = str(value).strip().lower()
        if s in {'1', 'true', 't', 'yes', 'y', 'si', 'sí'}:
            return True
        if s in {'0', 'false', 'f', 'no', 'n'}:
            return False
        return None

    def _parse_date(self, value):
        if value is None or value == '':
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        s = str(value).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    def _extract_value(self, row: dict, *keys, default=None):
        for k in keys:
            if k in row and row[k] not in (None, ''):
                return row[k]
        return default

    def _coerce_str(self, value):
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'true' if value else 'false'
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, float):
            try:
                if value.is_integer():
                    return str(int(value))
            except Exception:
                pass
        return str(value).strip()

    def _map_row_to_student_payload(self, row: dict):
        # Accept common Spanish/English headers.
        first_name = self._extract_value(row, 'first_name', 'nombres', 'nombre', 'name')
        last_name = self._extract_value(row, 'last_name', 'apellidos', 'apellido', 'surname')
        email = self._extract_value(row, 'email', 'correo', 'correo_electronico', 'e_mail')

        document_number = self._extract_value(
            row,
            'document_number',
            'numero_documento',
            'no_documento',
            'documento',
            'identificacion',
            'dni',
        )
        document_type = self._extract_value(row, 'document_type', 'tipo_documento', 'tipo_de_documento')

        sex_raw = self._extract_value(row, 'sex', 'sexo', 'genero')
        sex = None
        if sex_raw is not None and str(sex_raw).strip() != '':
            sx = str(sex_raw).strip().upper()
            if sx in {'M', 'MAS', 'MASCULINO', 'MALE'}:
                sex = 'M'
            elif sx in {'F', 'FEM', 'FEMENINO', 'FEMALE'}:
                sex = 'F'

        payload = {
            'first_name': self._coerce_str(first_name),
            'last_name': self._coerce_str(last_name),
            'email': self._coerce_str(email),
            'document_number': self._coerce_str(document_number),
            'document_type': self._coerce_str(document_type),
        }

        # Optional student fields
        optional_str_fields = {
            'place_of_issue': ('place_of_issue', 'lugar_expedicion', 'lugar_de_expedicion'),
            'nationality': ('nationality', 'nacionalidad'),
            'blood_type': ('blood_type', 'tipo_sangre', 'rh'),
            'address': ('address', 'direccion'),
            'neighborhood': ('neighborhood', 'barrio', 'barrio_vereda'),
            'phone': ('phone', 'telefono', 'celular'),
            'living_with': ('living_with', 'con_quien_vive'),
            'stratum': ('stratum', 'estrato'),
            'ethnicity': ('ethnicity', 'etnia'),
            'sisben_score': ('sisben_score', 'sisben', 'puntaje_sisben'),
            'eps': ('eps',),
            'disability_description': ('disability_description', 'descripcion_discapacidad'),
            'disability_type': ('disability_type', 'tipo_discapacidad'),
            'support_needs': ('support_needs', 'apoyos', 'apoyos_requeridos'),
            'allergies': ('allergies', 'alergias'),
            'emergency_contact_name': ('emergency_contact_name', 'contacto_emergencia_nombre'),
            'emergency_contact_phone': ('emergency_contact_phone', 'contacto_emergencia_telefono'),
            'emergency_contact_relationship': ('emergency_contact_relationship', 'contacto_emergencia_parentesco'),
            'financial_status': ('financial_status', 'estado_financiero'),
        }
        for target, aliases in optional_str_fields.items():
            v = self._extract_value(row, *aliases)
            if v is not None:
                payload[target] = self._coerce_str(v)

        birth_date_raw = self._extract_value(row, 'birth_date', 'fecha_nacimiento', 'nacimiento')
        birth_date = self._parse_date(birth_date_raw)
        if birth_date is not None:
            payload['birth_date'] = birth_date

        if sex is not None:
            payload['sex'] = sex

        is_victim_raw = self._extract_value(row, 'is_victim_of_conflict', 'victima_conflicto', 'victima_del_conflicto')
        is_victim = self._parse_bool(is_victim_raw)
        if is_victim is not None:
            payload['is_victim_of_conflict'] = is_victim

        has_disability_raw = self._extract_value(row, 'has_disability', 'tiene_discapacidad', 'discapacidad')
        has_disability = self._parse_bool(has_disability_raw)
        if has_disability is not None:
            payload['has_disability'] = has_disability

        return payload

    @action(detail=True, methods=["post"], url_path="import-academic-history")
    @transaction.atomic
    def import_academic_history(self, request, pk=None):
        """Import external academic history for a student.

        Creates (or reuses) an Enrollment for the given academic year with group=None,
        and stores imported subject/area finals in EnrollmentPromotionSnapshot.details.

        Body example:
        {
          "academic_year": 2024,
          "grade_name": "Octavo",
          "origin_school": "Colegio X",
          "subjects": [
            {"area": "Matemáticas", "subject": "Álgebra", "final_score": "3.80"},
            {"area": "Ciencias", "subject": "Biología", "final_score": "2.50"}
          ]
        }
        """

        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para importar historial."}, status=status.HTTP_403_FORBIDDEN)

        student = self.get_object()

        year_value = request.data.get("academic_year")
        grade_id = request.data.get("grade")
        grade_name = request.data.get("grade_name")
        origin_school = request.data.get("origin_school", "")
        subjects = request.data.get("subjects") or []

        if not year_value:
            return Response({"detail": "academic_year es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year_int = int(year_value)
        except Exception:
            return Response({"detail": "academic_year inválido"}, status=status.HTTP_400_BAD_REQUEST)

        if not grade_id and not grade_name:
            return Response({"detail": "grade o grade_name es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        if grade_id:
            grade = Grade.objects.filter(id=grade_id).first()
        else:
            grade = Grade.objects.filter(name=grade_name).first()

        if not grade:
            return Response({"detail": "Grado no encontrado"}, status=status.HTTP_400_BAD_REQUEST)

        # Create or reuse AcademicYear by numeric year
        academic_year = AcademicYear.objects.filter(year=year_int).first()
        if not academic_year:
            academic_year = AcademicYear.objects.create(year=year_int, status=AcademicYear.STATUS_PLANNING)

        enrollment, _ = Enrollment.objects.get_or_create(
            student=student,
            academic_year=academic_year,
            defaults={
                "grade": grade,
                "group": None,
                "status": "RETIRED",
                "origin_school": origin_school,
                "final_status": "IMPORTADO",
            },
        )

        # If enrollment already existed, keep grade consistent unless explicitly overridden
        if enrollment.grade_id != grade.id:
            enrollment.grade = grade
        if origin_school:
            enrollment.origin_school = origin_school
        if enrollment.status == "ACTIVE" and academic_year.status != "ACTIVE":
            # Avoid leaving past-year enrollments ACTIVE
            enrollment.status = "RETIRED"
        enrollment.save(update_fields=["grade", "origin_school", "status"])

        # Validate and compute decision from imported subject finals.
        passing_score = Decimal("3.00")
        failed_subjects = []
        failed_areas = set()
        failed_distinct_areas = set()

        normalized_subjects = []
        for idx, row in enumerate(subjects):
            area = (row.get("area") or "").strip()
            subject_name = (row.get("subject") or "").strip()
            score_raw = row.get("final_score")
            if not area or not subject_name or score_raw in (None, ""):
                return Response(
                    {"detail": f"subjects[{idx}] debe incluir area, subject, final_score"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                score = Decimal(str(score_raw)).quantize(Decimal("0.01"))
            except Exception:
                return Response(
                    {"detail": f"subjects[{idx}].final_score inválido"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            normalized_subjects.append({"area": area, "subject": subject_name, "final_score": str(score)})
            if score < passing_score:
                failed_subjects.append({"area": area, "subject": subject_name, "final_score": str(score)})
                failed_areas.add(area)
                failed_distinct_areas.add(area)

        failed_subjects_count = len(failed_subjects)
        failed_areas_count = len(failed_areas)
        failed_subjects_distinct_areas_count = len(failed_distinct_areas)

        # Apply your SIEE decision rules (by names, for imported records)
        if failed_areas_count >= 2:
            decision = "REPEATED"
        elif failed_subjects_count >= 3 and failed_subjects_distinct_areas_count >= 3:
            decision = "REPEATED"
        elif failed_areas_count == 0 and failed_subjects_count == 0:
            decision = "PROMOTED"
        elif failed_areas_count == 1 or failed_subjects_count <= 2:
            decision = "CONDITIONAL"
        else:
            decision = "REPEATED"

        # Persist as snapshot details (no need to create curriculum subjects/areas)
        from academic.models import EnrollmentPromotionSnapshot

        details = {
            "source": "IMPORTED",
            "academic_year": year_int,
            "origin_school": origin_school,
            "grade": {"id": grade.id, "name": grade.name, "ordinal": grade.ordinal},
            "passing_score": str(passing_score),
            "subjects": normalized_subjects,
            "failed_subjects": failed_subjects,
            "failed_areas": sorted(list(failed_areas)),
        }

        EnrollmentPromotionSnapshot.objects.update_or_create(
            enrollment=enrollment,
            defaults={
                "decision": decision,
                "failed_subjects_count": failed_subjects_count,
                "failed_areas_count": failed_areas_count,
                "failed_subjects_distinct_areas_count": failed_subjects_distinct_areas_count,
                "details": details,
            },
        )

        # Keep legacy field simple but informative
        enrollment.final_status = f"IMPORTADO ({decision})"
        enrollment.save(update_fields=["final_status"])

        return Response(
            {
                "enrollment_id": enrollment.id,
                "academic_year": {"id": academic_year.id, "year": academic_year.year},
                "decision": decision,
                "failed_subjects_count": failed_subjects_count,
                "failed_areas_count": failed_areas_count,
            },
            status=status.HTTP_201_CREATED,
        )

    def _read_rows_from_upload(self, upload):
        name = getattr(upload, 'name', '') or ''
        ext = os.path.splitext(name.lower())[1]

        if ext == '.csv':
            raw = upload.read()
            # Try utf-8 with BOM first, then fallback.
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                text = raw.decode('latin-1')
            f = io.StringIO(text)
            reader = csv.DictReader(f)
            return [
                {self._normalize_header(k): v for k, v in (row or {}).items() if k is not None}
                for row in reader
            ]

        if ext in {'.xlsx', '.xls'}:
            if ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(upload, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            else:
                import xlrd
                book = xlrd.open_workbook(file_contents=upload.read())
                sheet = book.sheet_by_index(0)
                rows = [sheet.row_values(r) for r in range(sheet.nrows)]

            if not rows:
                return []

            headers = [self._normalize_header(h) for h in (rows[0] or [])]
            out = []
            for r in rows[1:]:
                if r is None:
                    continue
                row_dict = {}
                empty = True
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    v = r[i] if i < len(r) else None
                    if v not in (None, ''):
                        empty = False
                    row_dict[h] = v
                if not empty:
                    out.append(row_dict)
            return out

        raise ValueError('Formato no soportado. Usa CSV, XLSX o XLS.')

    @action(detail=False, methods=['post'], url_path='bulk-import', parser_classes=(MultiPartParser, FormParser))
    def bulk_import(self, request):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para importar estudiantes."}, status=status.HTTP_403_FORBIDDEN)

        upload = request.FILES.get('file')
        if not upload:
            return Response({"detail": "Archivo requerido (campo 'file')."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = self._read_rows_from_upload(upload)
        except ValueError as ve:
            return Response({"detail": str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": f"No se pudo leer el archivo: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        failed = 0
        errors = []

        # Row numbers are 2-based (1 header + first data row = 2)
        for idx, raw_row in enumerate(rows, start=2):
            try:
                payload = self._map_row_to_student_payload(raw_row)

                # Enforce required fields for bulk import to avoid unique-blank collisions
                if not payload.get('first_name'):
                    raise serializers.ValidationError({"first_name": "Este campo es requerido."})
                if not payload.get('last_name'):
                    raise serializers.ValidationError({"last_name": "Este campo es requerido."})
                if not payload.get('document_number'):
                    raise serializers.ValidationError({"document_number": "Este campo es requerido."})

                serializer = self.get_serializer(data=payload)
                serializer.is_valid(raise_exception=True)
                with transaction.atomic():
                    serializer.save()
                created += 1
            except Exception as e:
                failed += 1
                detail = None
                if hasattr(e, 'detail'):
                    detail = e.detail
                else:
                    detail = str(e)
                errors.append({
                    'row': idx,
                    'error': detail,
                })

        return Response(
            {
                'created': created,
                'failed': failed,
                'errors': errors,
            },
            status=status.HTTP_200_OK,
        )


class FamilyMemberViewSet(viewsets.ModelViewSet):
    queryset = FamilyMember.objects.select_related("student").all().order_by("id")
    serializer_class = FamilyMemberSerializer
    permission_classes = [KampusModelPermissions]
    parser_classes = (JSONParser, FormParser, MultiPartParser)

    def get_permissions(self):
        if getattr(self.request.user, 'role', None) == 'TEACHER':
            return [IsAuthenticated(), IsTeacherDirectorOfRelatedStudent()]
        return super().get_permissions()

    def get_queryset(self):
        role = getattr(self.request.user, 'role', None)
        if role in {'PARENT', 'STUDENT'}:
            return FamilyMember.objects.none()
        if role == 'TEACHER':
            allowed_ids = _director_student_ids(self.request.user)
            if not allowed_ids:
                return FamilyMember.objects.none()
            return FamilyMember.objects.select_related('student').filter(student_id__in=allowed_ids).order_by('id')
        return super().get_queryset()

    def create(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para modificar familiares."}, status=status.HTTP_403_FORBIDDEN)

        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para modificar familiares."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para modificar familiares."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para modificar familiares."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = (
        Enrollment.objects.select_related("student", "student__user", "academic_year", "grade", "group")
        .all()
        .order_by("student__user__last_name", "student__user__first_name", "id")
    )
    serializer_class = EnrollmentSerializer
    permission_classes = [KampusModelPermissions]
    pagination_class = EnrollmentPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = [
        "student__user__first_name",
        "student__user__last_name",
        "student__document_number",
    ]
    filterset_fields = ["student", "academic_year", "grade", "group", "status"]

    def get_permissions(self):
        # Teachers need a safe, filtered way to pick enrollments for their groups.
        if getattr(self, "action", None) in {"my"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        if getattr(self.request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Enrollment.objects.none()
        return super().get_queryset()

    def create(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar matrículas."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar matrículas."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar matrículas."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar matrículas."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='my')
    def my(self, request, *args, **kwargs):
        """List ACTIVE enrollments for the current teacher's groups (directed or assigned).

        This endpoint exists so teachers can register discipline cases without exposing
        the full enrollments index.
        """

        user = getattr(request, 'user', None)
        role = getattr(user, 'role', None)
        if role != 'TEACHER':
            return Response({"detail": "No tienes permisos."}, status=status.HTTP_403_FORBIDDEN)

        active_year = AcademicYear.objects.filter(status='ACTIVE').first()
        if not active_year:
            page = self.paginate_queryset(Enrollment.objects.none())
            if page is not None:
                return self.get_paginated_response([])
            return Response([], status=status.HTTP_200_OK)

        directed_group_ids = set(
            Group.objects.filter(director=user, academic_year=active_year).values_list('id', flat=True)
        )
        assigned_group_ids = set(
            TeacherAssignment.objects.filter(teacher=user, academic_year=active_year).values_list('group_id', flat=True)
        )
        allowed_group_ids = directed_group_ids | assigned_group_ids
        if not allowed_group_ids:
            page = self.paginate_queryset(Enrollment.objects.none())
            if page is not None:
                return self.get_paginated_response([])
            return Response([], status=status.HTTP_200_OK)

        group_raw = request.query_params.get('group_id') or request.query_params.get('group')
        group_id = None
        if group_raw is not None and str(group_raw).strip() != '':
            try:
                group_id = int(group_raw)
            except Exception:
                group_id = None

        student_raw = request.query_params.get('student') or request.query_params.get('student_id')
        student_id = None
        if student_raw is not None and str(student_raw).strip() != '':
            try:
                student_id = int(student_raw)
            except Exception:
                student_id = None

        include_all_years_raw = request.query_params.get('include_all_years')
        include_all_years = str(include_all_years_raw).lower() in {'1', 'true', 'yes', 'y'}

        base_active_qs = Enrollment.objects.filter(
            academic_year=active_year,
            status='ACTIVE',
            group_id__in=allowed_group_ids,
        )

        if include_all_years and student_id is not None:
            # Security: only allow full history for students the teacher can see
            # in the ACTIVE academic year.
            if not base_active_qs.filter(student_id=student_id).exists():
                page = self.paginate_queryset(Enrollment.objects.none())
                if page is not None:
                    return self.get_paginated_response([])
                return Response([], status=status.HTTP_200_OK)

            qs = (
                Enrollment.objects.select_related('student', 'student__user', 'academic_year', 'grade', 'group')
                .filter(student_id=student_id)
                .order_by('-academic_year__year', '-id')
            )
        else:
            qs = (
                Enrollment.objects.select_related('student', 'student__user', 'academic_year', 'grade', 'group')
                .filter(
                    academic_year=active_year,
                    status='ACTIVE',
                    group_id__in=allowed_group_ids,
                )
                .order_by('student__user__last_name', 'student__user__first_name', 'id')
            )

        if group_id is not None:
            qs = qs.filter(group_id=group_id)

        if student_id is not None:
            qs = qs.filter(student_id=student_id)

        q = (request.query_params.get('q') or '').strip()
        if q:
            qs = qs.filter(
                Q(student__user__first_name__icontains=q)
                | Q(student__user__last_name__icontains=q)
                | Q(student__document_number__icontains=q)
            )

        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)

        ser = self.get_serializer(qs, many=True)
        return Response(ser.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def report(self, request, *args, **kwargs):
        # Filters
        def _to_int_or_none(value):
            if value is None:
                return None
            raw = str(value).strip()
            if raw == "":
                return None
            try:
                return int(raw)
            except Exception:
                return None

        year_id = _to_int_or_none(request.query_params.get('year'))
        grade_id = _to_int_or_none(request.query_params.get('grade'))
        group_id = _to_int_or_none(request.query_params.get('group'))

        # IMPORTANT: Do not use query param named `format` here.
        # DRF treats ?format=... as a renderer override and can return 404
        # before reaching this action if the renderer isn't registered.
        report_format = (
            request.query_params.get('export')
            or request.query_params.get('report_format')
            or kwargs.get('format')
            or 'csv'
        )
        report_format = str(report_format).strip().lower()
        want_async = (request.query_params.get("async") or "").strip().lower() in {"1", "true", "yes"}
        
        enrollments_qs = (
            Enrollment.objects.select_related("student", "student__user", "grade", "group", "academic_year")
            .all()
        )
        
        # Filter logic
        year_name = "Todos"
        grade_name = ""
        group_name = ""

        if year_id is not None:
            enrollments_qs = enrollments_qs.filter(academic_year_id=year_id)
            try:
                year_name = AcademicYear.objects.get(pk=year_id).year
            except: pass
        else:
            # Default to active year
            active_year = AcademicYear.objects.filter(status='ACTIVE').first()
            if active_year:
                enrollments_qs = enrollments_qs.filter(academic_year=active_year)
                year_name = active_year.year
        
        if grade_id is not None:
            enrollments_qs = enrollments_qs.filter(grade_id=grade_id)
            try:
                grade_name = Grade.objects.get(pk=grade_id).name
            except: pass

        if group_id is not None:
            enrollments_qs = enrollments_qs.filter(group_id=group_id)
            try:
                group_name = Group.objects.get(pk=group_id).name
            except: pass

        enrollments = sort_enrollments_for_enrollment_list(list(enrollments_qs))
            
        # PDF Generation
        if report_format == 'pdf':
            if want_async:
                from datetime import timedelta  # noqa: PLC0415
                from django.utils import timezone  # noqa: PLC0415
                from reports.models import ReportJob  # noqa: PLC0415
                from reports.serializers import ReportJobSerializer  # noqa: PLC0415
                from reports.tasks import generate_report_job_pdf  # noqa: PLC0415

                ttl_hours = int(getattr(settings, "REPORT_JOBS_TTL_HOURS", 24))
                expires_at = timezone.now() + timedelta(hours=ttl_hours)

                job = ReportJob.objects.create(
                    created_by=request.user,
                    report_type=ReportJob.ReportType.ENROLLMENT_LIST,
                    params={"year_id": year_id, "grade_id": grade_id, "group_id": group_id},
                    expires_at=expires_at,
                )
                generate_report_job_pdf.delay(job.id)
                out = ReportJobSerializer(job, context={"request": request}).data
                return Response(out, status=status.HTTP_202_ACCEPTED)

            try:
                institution = Institution.objects.first() or Institution()

                html_string = render_to_string('students/reports/enrollment_list_pdf.html', {
                    'enrollments': enrollments,
                    'institution': institution,
                    'year_name': year_name,
                    'grade_name': grade_name,
                    'group_name': group_name,
                })

                from reports.weasyprint_utils import WeasyPrintUnavailableError, render_pdf_bytes_from_html  # noqa: PLC0415

                pdf_bytes = render_pdf_bytes_from_html(html=html_string, base_url=str(settings.BASE_DIR))

                response = HttpResponse(pdf_bytes, content_type='application/pdf')
                response['Content-Disposition'] = 'inline; filename="reporte_matriculados.pdf"'
                response["Deprecation"] = "true"
                sunset = getattr(settings, "REPORTS_SYNC_SUNSET_DATE", "2026-06-30")
                response["Sunset"] = str(sunset)
                response["Link"] = '</api/reports/jobs/>; rel="alternate"'
                return response
            except WeasyPrintUnavailableError as e:
                payload = {"error": str(e)}
                if getattr(settings, 'DEBUG', False):
                    payload["traceback"] = traceback.format_exc()
                return Response(payload, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            except Exception as e:
                payload = {"error": "Error generating PDF", "detail": str(e)}
                if getattr(settings, 'DEBUG', False):
                    payload["traceback"] = traceback.format_exc()
                return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # XLSX Generation
        if report_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Matriculados"

            headers = ['Documento', 'Nombres', 'Apellidos', 'Grado', 'Grupo', 'Año', 'Estado', 'Paz y Salvo']
            ws.append(headers)

            for enrollment in enrollments:
                student = enrollment.student
                user = student.user
                ws.append(
                    [
                        student.document_number,
                        (user.first_name or '').upper(),
                        (user.last_name or '').upper(),
                        enrollment.grade.name if enrollment.grade else '',
                        enrollment.group.name if enrollment.group else '',
                        enrollment.academic_year.year,
                        enrollment.get_status_display(),
                        student.get_financial_status_display(),
                    ]
                )

            # Simple column sizing
            for col_idx, header in enumerate(headers, start=1):
                max_len = len(str(header))
                for cell in ws[get_column_letter(col_idx)]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            response = HttpResponse(
                out.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="matriculados.xlsx"'
            return response

        # CSV Generation (Default)
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="matriculados.csv"'

        # UTF-8 BOM helps Excel interpret accents correctly.
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Documento', 'Nombres', 'Apellidos', 'Grado', 'Grupo', 'Año', 'Estado', 'Paz y Salvo'])

        for enrollment in enrollments:
            student = enrollment.student
            user = student.user
            writer.writerow([
                student.document_number,
                (user.first_name or '').upper(),
                (user.last_name or '').upper(),
                enrollment.grade.name if enrollment.grade else '',
                enrollment.group.name if enrollment.group else '',
                enrollment.academic_year.year,
                enrollment.get_status_display(),
                student.get_financial_status_display(),
            ])

        return response

    @action(
        detail=True,
        methods=['get'],
        url_path='academic-report',
        permission_classes=[IsAuthenticated],
    )
    def academic_report(self, request, pk=None):
        """Genera informe académico por periodos para una matrícula.

        GET /api/enrollments/{id}/academic-report/?period=<period_id>
        """
        async_raw = (request.query_params.get("async") or "").strip().lower()
        async_requested = async_raw in {"1", "true", "yes"}

        def _to_int_or_none(value):
            if value is None:
                return None
            raw = str(value).strip()
            if raw == "":
                return None
            try:
                return int(raw)
            except Exception:
                return None

        period_id = _to_int_or_none(request.query_params.get('period'))
        if period_id is None:
            return Response({"detail": "period is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch enrollment without inheriting the restrictive queryset (teachers/parents are blocked there).
        enrollment = (
            Enrollment.objects.select_related(
                'student',
                'student__user',
                'grade',
                'group',
                'group__director',
                'academic_year',
            )
            .filter(id=pk)
            .first()
        )
        if not enrollment:
            return Response({"detail": "Enrollment not found"}, status=status.HTTP_404_NOT_FOUND)

        user = getattr(request, 'user', None)
        role = getattr(user, 'role', None)
        # Allow admin-like roles, and also the teacher who directs the group.
        if role not in {'SUPERADMIN', 'ADMIN', 'COORDINATOR'}:
            if not (role == 'TEACHER' and enrollment.group and enrollment.group.director_id == getattr(user, 'id', None)):
                return Response({"detail": "No tienes permisos para ver este informe."}, status=status.HTTP_403_FORBIDDEN)

        try:
            period = Period.objects.select_related('academic_year').get(id=period_id)
        except Period.DoesNotExist:
            return Response({"detail": "Periodo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        if period.academic_year_id != enrollment.academic_year_id:
            return Response(
                {"detail": "El periodo no corresponde al año lectivo de la matrícula."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if async_requested:
            from reports.models import ReportJob
            from reports.serializers import ReportJobSerializer
            from reports.tasks import generate_report_job_pdf

            from django.utils import timezone
            from datetime import timedelta

            ttl_hours = int(getattr(settings, "REPORT_JOBS_TTL_HOURS", 24))
            expires_at = timezone.now() + timedelta(hours=ttl_hours)

            job = ReportJob.objects.create(
                created_by=request.user,
                report_type=ReportJob.ReportType.ACADEMIC_PERIOD_ENROLLMENT,
                params={"enrollment_id": enrollment.id, "period_id": period.id},
                expires_at=expires_at,
            )
            generate_report_job_pdf.delay(job.id)
            out = ReportJobSerializer(job, context={"request": request}).data
            return Response(out, status=status.HTTP_202_ACCEPTED)

        try:
            pdf_bytes = generate_academic_period_report_pdf(enrollment=enrollment, period=period)
            filename = f"informe-academico-enrollment-{enrollment.id}-period-{period.id}.pdf"
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'inline; filename="{filename}"'

            # Sync PDF is kept for backward compatibility, but jobs are preferred.
            response["Deprecation"] = "true"
            sunset = getattr(settings, "REPORTS_SYNC_SUNSET_DATE", "2026-06-30")
            response["Sunset"] = str(sunset)
            response["Link"] = '</api/reports/jobs/>; rel="alternate"'
            return response
        except Exception as e:
            payload = {"detail": "Error generating PDF", "error": str(e)}
            if getattr(settings, "DEBUG", False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=["get"], url_path="pap-plans")
    def pap_plans(self, request):
        """List PAP plans (conditional promotion plans).

        Query params:
        - status: OPEN|CLEARED|FAILED (optional)
        - academic_year: AcademicYear id (optional)
        - due_period: Period id (optional)
        """

        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para ver PAP."}, status=status.HTTP_403_FORBIDDEN)

        status_raw = (request.query_params.get("status") or "").strip().upper()
        year_id_raw = request.query_params.get("academic_year")
        due_period_raw = request.query_params.get("due_period")

        qs = ConditionalPromotionPlan.objects.select_related(
            "enrollment",
            "enrollment__student",
            "enrollment__student__user",
            "enrollment__academic_year",
            "enrollment__grade",
            "due_period",
            "source_enrollment",
            "source_enrollment__grade",
        ).order_by("-updated_at", "-created_at")

        if status_raw:
            if status_raw not in {
                ConditionalPromotionPlan.STATUS_OPEN,
                ConditionalPromotionPlan.STATUS_CLEARED,
                ConditionalPromotionPlan.STATUS_FAILED,
            }:
                return Response({"detail": "status inválido"}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(status=status_raw)

        if year_id_raw:
            try:
                year_id = int(year_id_raw)
            except Exception:
                return Response({"detail": "academic_year inválido"}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(enrollment__academic_year_id=year_id)

        if due_period_raw:
            try:
                due_period_id = int(due_period_raw)
            except Exception:
                return Response({"detail": "due_period inválido"}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(due_period_id=due_period_id)

        results = []
        for plan in qs[:500]:
            enr = plan.enrollment
            student_user = enr.student.user if enr and enr.student_id else None
            results.append(
                {
                    "id": plan.id,
                    "status": plan.status,
                    "due_period": {
                        "id": plan.due_period_id,
                        "name": plan.due_period.name if plan.due_period_id else None,
                    },
                    "enrollment": {
                        "id": enr.id,
                        "academic_year": {
                            "id": enr.academic_year_id,
                            "year": enr.academic_year.year if enr.academic_year_id else None,
                        },
                        "grade": {"id": enr.grade_id, "name": enr.grade.name if enr.grade_id else None},
                        "student": {
                            "id": enr.student_id,
                            "name": student_user.get_full_name() if student_user else None,
                            "document_number": getattr(enr.student, "document_number", "") if enr.student_id else "",
                        },
                    },
                    "source_enrollment": {
                        "id": plan.source_enrollment_id,
                        "grade": {
                            "id": plan.source_enrollment.grade_id if plan.source_enrollment_id else None,
                            "name": plan.source_enrollment.grade.name if plan.source_enrollment_id else None,
                        }
                        if plan.source_enrollment_id
                        else None,
                    },
                    "pending_subject_ids": plan.pending_subject_ids,
                    "pending_area_ids": plan.pending_area_ids,
                    "notes": plan.notes,
                    "created_at": plan.created_at,
                    "updated_at": plan.updated_at,
                }
            )

        return Response({"results": results})

    @action(detail=True, methods=["get"], url_path="pap")
    def pap(self, request, pk=None):
        """Return conditional promotion plan (PAP) for an enrollment."""

        enrollment = self.get_object()
        plan = getattr(enrollment, "conditional_plan", None)
        if not plan:
            return Response({"detail": "Este enrollment no tiene PAP."}, status=status.HTTP_404_NOT_FOUND)

        return Response(
            {
                "id": plan.id,
                "enrollment_id": enrollment.id,
                "source_enrollment_id": plan.source_enrollment_id,
                "due_period_id": plan.due_period_id,
                "pending_subject_ids": plan.pending_subject_ids,
                "pending_area_ids": plan.pending_area_ids,
                "status": plan.status,
                "notes": plan.notes,
                "created_at": plan.created_at,
                "updated_at": plan.updated_at,
            }
        )

    @action(detail=True, methods=["post"], url_path="pap/resolve")
    @transaction.atomic
    def pap_resolve(self, request, pk=None):
        """Resolve a PAP plan.

        Body: {"status": "CLEARED"|"FAILED", "notes": "..."}
        - CLEARED: keeps current grade, sets Enrollment.final_status.
        - FAILED: reverts Enrollment.grade to source_enrollment.grade (group cleared) and sets Enrollment.final_status.
        """

        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para resolver PAP."}, status=status.HTTP_403_FORBIDDEN)

        enrollment = self.get_object()
        plan = getattr(enrollment, "conditional_plan", None)
        if not plan:
            return Response({"detail": "Este enrollment no tiene PAP."}, status=status.HTTP_404_NOT_FOUND)

        if plan.status != ConditionalPromotionPlan.STATUS_OPEN:
            return Response({"detail": "Este PAP ya fue resuelto."}, status=status.HTTP_400_BAD_REQUEST)

        new_status = (request.data.get("status") or "").strip().upper()
        notes = (request.data.get("notes") or "").strip()

        if new_status not in {ConditionalPromotionPlan.STATUS_CLEARED, ConditionalPromotionPlan.STATUS_FAILED}:
            return Response(
                {"detail": "status debe ser CLEARED o FAILED"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        plan.status = new_status
        if notes:
            plan.notes = notes
        plan.save(update_fields=["status", "notes", "updated_at"])

        if new_status == ConditionalPromotionPlan.STATUS_CLEARED:
            enrollment.final_status = "PAP APROBADO"
            enrollment.save(update_fields=["final_status"])
        else:
            # FAILED => revert grade to source enrollment grade when available.
            source = plan.source_enrollment
            if source and source.grade_id:
                enrollment.grade_id = source.grade_id
                enrollment.group = None
                enrollment.final_status = "PAP NO APROBADO (RETENIDO)"
                enrollment.save(update_fields=["grade", "group", "final_status"])
            else:
                enrollment.final_status = "PAP NO APROBADO"
                enrollment.save(update_fields=["final_status"])

        return Response(
            {
                "enrollment_id": enrollment.id,
                "pap": {"id": plan.id, "status": plan.status, "due_period_id": plan.due_period_id},
                "enrollment": {"grade_id": enrollment.grade_id, "group_id": enrollment.group_id, "final_status": enrollment.final_status},
            },
            status=status.HTTP_200_OK,
        )


class StudentNoveltyViewSet(viewsets.ModelViewSet):
    queryset = StudentNovelty.objects.all().order_by("-date")
    serializer_class = StudentNoveltySerializer
    permission_classes = [KampusModelPermissions]

    def get_queryset(self):
        if getattr(self.request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return StudentNovelty.objects.none()
        return super().get_queryset()

    def create(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para registrar novedades."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para registrar novedades."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para registrar novedades."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'TEACHER', 'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para registrar novedades."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    def perform_create(self, serializer):
        novelty = serializer.save()
        student = novelty.student
        user = student.user
        
        if novelty.novelty_type == "RETIRO":
            user.is_active = False
            user.save()
            # Update active enrollments to RETIRED
            student.enrollment_set.filter(status="ACTIVE").update(status="RETIRED")
            
        elif novelty.novelty_type == "REINGRESO":
            user.is_active = True
            user.save()


class StudentDocumentViewSet(viewsets.ModelViewSet):
    queryset = StudentDocument.objects.all().order_by("-uploaded_at")
    serializer_class = StudentDocumentSerializer
    permission_classes = [KampusModelPermissions]

    def get_permissions(self):
        if getattr(self.request.user, 'role', None) == 'TEACHER':
            return [IsAuthenticated(), IsTeacherDirectorOfRelatedStudent()]
        return super().get_permissions()

    def get_queryset(self):
        role = getattr(self.request.user, 'role', None)
        if role in {'PARENT', 'STUDENT'}:
            return StudentDocument.objects.none()
        if role == 'TEACHER':
            allowed_ids = _director_student_ids(self.request.user)
            if not allowed_ids:
                return StudentDocument.objects.none()
            return StudentDocument.objects.select_related('student').filter(student_id__in=allowed_ids).order_by('-uploaded_at')
        return super().get_queryset()

    def create(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar documentos."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar documentos."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar documentos."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, 'role', None) in {'PARENT', 'STUDENT'}:
            return Response({"detail": "No tienes permisos para gestionar documentos."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class ObserverAnnotationViewSet(viewsets.ModelViewSet):
    queryset = ObserverAnnotation.objects.select_related("student", "period", "created_by", "updated_by").all()
    serializer_class = ObserverAnnotationSerializer
    permission_classes = [KampusModelPermissions]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["student", "period", "annotation_type", "is_automatic"]

    def get_permissions(self):
        if getattr(self.request.user, "role", None) == "TEACHER":
            return [IsAuthenticated(), IsTeacherAssignedOrDirectorOfRelatedStudent()]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset().filter(is_deleted=False)
        role = getattr(self.request.user, "role", None)
        if role in {"PARENT", "STUDENT"}:
            return qs.none()
        if role == "TEACHER":
            allowed_ids = _teacher_managed_student_ids(self.request.user)
            if not allowed_ids:
                return qs.none()
            return qs.filter(student_id__in=allowed_ids)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        instance: ObserverAnnotation = self.get_object()
        if getattr(self.request.user, "role", None) == "TEACHER" and bool(getattr(instance, "is_automatic", False)):
            raise serializers.ValidationError({"detail": "Las anotaciones automáticas no se pueden editar."})
        serializer.save(updated_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        annotation: ObserverAnnotation = self.get_object()
        if getattr(request.user, "role", None) == "TEACHER" and bool(getattr(annotation, "is_automatic", False)):
            return Response({"detail": "Las anotaciones automáticas no se pueden eliminar."}, status=status.HTTP_400_BAD_REQUEST)
        annotation.is_deleted = True
        annotation.deleted_at = datetime.now()
        annotation.deleted_by = request.user
        annotation.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "updated_at"])
        return Response(status=status.HTTP_204_NO_CONTENT)


class BulkEnrollmentView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [HasDjangoPermission]
    required_permission = "students.add_enrollment"

    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        file_obj = request.FILES['file']
        if not file_obj.name.endswith('.csv'):
             return Response({"error": "File must be CSV"}, status=status.HTTP_400_BAD_REQUEST)

        decoded_file = file_obj.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        
        results = {"success": 0, "errors": []}
        
        # Get active academic year
        active_year = AcademicYear.objects.filter(status='ACTIVE').first()
        if not active_year:
             return Response({"error": "No active academic year found"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for row_index, row in enumerate(reader):
                try:
                    # Expected columns: document_number, first_name, last_name, grade_name, group_name (optional)
                    doc_number = row.get('document_number')
                    if not doc_number:
                        continue
                        
                    # Find or Create Student
                    student = Student.objects.filter(document_number=doc_number).first()
                    if not student:
                        # Create basic student
                        first_name = row.get('first_name', 'Unknown')
                        last_name = row.get('last_name', 'Unknown')
                        email = row.get('email', '')
                        
                        # Generate username
                        base_username = f"{first_name[:1]}{last_name}".lower().replace(" ", "")
                        username = base_username
                        counter = 1
                        while User.objects.filter(username=username).exists():
                            username = f"{base_username}{counter}"
                            counter += 1
                            
                        user = User.objects.create_user(username=username, password=doc_number, first_name=first_name, last_name=last_name, email=email, role='STUDENT')
                        student = Student.objects.create(user=user, document_number=doc_number)
                    
                    # Find Grade
                    grade_name = row.get('grade_name')
                    grade = Grade.objects.filter(name=grade_name).first()
                    if not grade:
                        results['errors'].append(f"Row {row_index}: Grade '{grade_name}' not found")
                        continue

                    # Find Group (Optional)
                    group = None
                    group_name = row.get('group_name')
                    if group_name:
                        group = Group.objects.filter(name=group_name, grade=grade, academic_year=active_year).first()
                        if not group:
                             results['errors'].append(f"Row {row_index}: Group '{group_name}' not found for grade '{grade_name}'")
                             # Continue without group or skip? Let's skip enrollment if group specified but not found
                             continue
                    
                    # Check existing enrollment
                    if Enrollment.objects.filter(student=student, academic_year=active_year).exists():
                        results['errors'].append(f"Row {row_index}: Student {doc_number} already enrolled in this year")
                        continue

                    # Create Enrollment
                    Enrollment.objects.create(
                        student=student,
                        academic_year=active_year,
                        grade=grade,
                        group=group,
                        status='ACTIVE'
                    )
                    results['success'] += 1
                    
                except Exception as e:
                    results['errors'].append(f"Row {row_index}: {str(e)}")
        
        return Response(results)


def _format_decimal_score(value: float) -> str:
    try:
        return f"{float(value):.2f}"
    except Exception:
        return str(value)


def _performance_from_score(score: float) -> str:
    # Simple mapping aligned with the provided sample.
    return "ALTO" if score >= 4.0 else "BASICO"


def _qr_data_uri(text: str) -> str:
    if not qrcode:
        return ""
    img = qrcode.make(text)
    buff = io.BytesIO()
    img.save(buff, format="PNG")
    encoded = base64.b64encode(buff.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def _qr_temp_png_path(text: str) -> str:
    """Generate a QR image as a temporary PNG file and return its absolute path.

    PDFs are generated server-side; file paths are reliable for PDF renderers.
    Caller is responsible for deleting the file.
    """

    if not qrcode:
        return ""

    img = qrcode.make(text)
    tmp = tempfile.NamedTemporaryFile(prefix="kampus_qr_", suffix=".png", delete=False)
    tmp.close()
    img.save(tmp.name, format="PNG")
    return tmp.name


def _active_academic_year() -> AcademicYear | None:
    return AcademicYear.objects.filter(status=AcademicYear.STATUS_ACTIVE).first()


def _subjects_for_grade(grade_id: int):
    return (
        AcademicLoad.objects.filter(grade_id=grade_id)
        .select_related("subject", "subject__area")
        .order_by("subject__area__name", "subject__name", "id")
    )


def _normalize_text_for_compare(value: str) -> str:
    text = (value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("utf-8")
    # Keep only alphanumerics for robust comparisons.
    return "".join(ch for ch in text if ch.isalnum())


def _format_certificate_subject_label(
    title: str,
    *,
    grade_level_type: str | None = None,
) -> tuple[str, bool]:
    """Return (label, skip).

    Rules:
    - Prefer showing ONLY the subject name (no area) for most rows.
    - For primaria, keep 'Matemáticas - (Aritmética/Geometría/Estadística)'.
    - Avoid/skip the invalid 'Matemáticas - Matemáticas' row in primaria.
    """

    raw = (title or "").strip()
    if not raw:
        return "", True

    area = None
    subject = raw
    if " - " in raw:
        area, subject = [p.strip() for p in raw.split(" - ", 1)]

    if not area:
        return subject, False

    area_norm = _normalize_text_for_compare(area)
    subject_norm = _normalize_text_for_compare(subject)
    is_primary = (grade_level_type or "").upper() == "PRIMARY"

    math_area_norms = {"matematica", "matematicas"}

    if area_norm in math_area_norms and subject_norm in math_area_norms and is_primary:
        # Primary plan should not include a 'Matemáticas - Matemáticas' entry.
        return "", True

    if area_norm == subject_norm:
        # Avoid duplicates like 'X - X'.
        return subject, False

    if area_norm in math_area_norms and subject_norm in {"aritmetica", "geometria", "estadistica"}:
        # Standardize the label to the plural form users expect.
        return f"Matemáticas - {subject}", False

    # Default: subject only.
    return subject, False


def _parse_score_decimal(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        try:
            return Decimal(str(float(s)))
        except Exception:
            return None


def _certificate_area_rows_from_plan(
    *,
    loads,
    subject_score_by_load_id: dict[int, Decimal],
    academic_year_id: int | None,
) -> list[dict[str, object]]:
    """Build area rows using the grade's study plan.

    - Area definitive score: weighted average of subject scores using AcademicLoad.weight_percentage.
    - Area intensity: sum of AcademicLoad.hours_per_week.
    """

    areas: dict[str, dict[str, object]] = {}

    for load in loads:
        if not load or not getattr(load, "subject", None) or not getattr(load.subject, "area", None):
            continue

        area_name = str(load.subject.area.name or "").strip()
        if not area_name:
            continue

        hours = int(getattr(load, "hours_per_week", 0) or 0)
        weight_pct = int(getattr(load, "weight_percentage", 100) or 100)

        bucket = areas.setdefault(
            area_name,
            {
                "area_subject": area_name,
                "hours_per_week": 0,
                "_weighted_sum": Decimal("0"),
                "_weight_sum": Decimal("0"),
            },
        )

        bucket["hours_per_week"] = int(bucket.get("hours_per_week") or 0) + hours

        score = subject_score_by_load_id.get(int(load.id))
        if score is None:
            continue

        w = Decimal(max(weight_pct, 0))
        bucket["_weighted_sum"] = Decimal(bucket.get("_weighted_sum") or Decimal("0")) + (score * w)
        bucket["_weight_sum"] = Decimal(bucket.get("_weight_sum") or Decimal("0")) + w

    out: list[dict[str, object]] = []
    for area_name in sorted(areas.keys(), key=lambda x: (x or "").lower()):
        b = areas[area_name]
        weight_sum: Decimal = Decimal(b.get("_weight_sum") or Decimal("0"))
        weighted_sum: Decimal = Decimal(b.get("_weighted_sum") or Decimal("0"))

        if weight_sum > 0:
            avg = (weighted_sum / weight_sum).quantize(Decimal("0.01"))
            try:
                perf = match_scale(int(academic_year_id), avg).name if academic_year_id else _performance_from_score(float(avg))
            except Exception:
                perf = _performance_from_score(float(avg))

            out.append(
                {
                    "area_subject": area_name,
                    "hours_per_week": int(b.get("hours_per_week") or 0),
                    "score": f"{avg:.2f}",
                    "performance": perf or "",
                }
            )
        else:
            out.append(
                {
                    "area_subject": area_name,
                    "hours_per_week": int(b.get("hours_per_week") or 0),
                    "score": "—",
                    "performance": "—",
                }
            )

    return out


def _certificate_studies_build_context(request, data: dict, institution: Institution):
    """Build context for the certificate HTML/PDF.

    This is used by both the HTML preview and the PDF issuance.
    """

    enrollment_id = data.get("enrollment_id")
    academic_year_id = data.get("academic_year_id")

    manual_name = (data.get("student_full_name") or "").strip()
    manual_doc_type = (data.get("document_type") or "").strip() or "Documento"
    manual_doc_number = (data.get("document_number") or "").strip()
    manual_grade_id = data.get("grade_id")
    manual_year = data.get("academic_year")
    manual_campus_id = data.get("campus_id")

    enrollment = None
    campus = None

    if enrollment_id:
        enrollment = (
            Enrollment.objects.select_related(
                "student",
                "student__user",
                "grade",
                "grade__level",
                "academic_year",
                "campus",
            )
            .get(pk=int(enrollment_id))
        )

        campus = getattr(enrollment, "campus", None)

        student_full_name = enrollment.student.user.get_full_name() if enrollment.student and enrollment.student.user else ""
        document_type = getattr(enrollment.student, "document_type", "") or "Documento"
        document_number = getattr(enrollment.student, "document_number", "") or ""
        grade = enrollment.grade
        academic_year = enrollment.academic_year.year if enrollment.academic_year else ""
        if academic_year_id:
            try:
                ay = AcademicYear.objects.get(pk=int(academic_year_id))
                academic_year = ay.year
            except Exception:
                pass
    else:
        if not manual_name or not manual_doc_number or not manual_grade_id:
            raise serializers.ValidationError(
                "For manual issuance you must send student_full_name, document_number and grade_id."
            )

        grade = Grade.objects.select_related("level").get(pk=int(manual_grade_id))

        if manual_campus_id:
            try:
                campus = Campus.objects.select_related("institution").get(pk=int(manual_campus_id))
            except Exception:
                campus = None

        student_full_name = manual_name
        document_type = manual_doc_type
        document_number = manual_doc_number

        if manual_year:
            academic_year = manual_year
        else:
            active = _active_academic_year()
            academic_year = active.year if active else date.today().year

    loads = _subjects_for_grade(grade.id)

    subject_score_by_load_id: dict[int, Decimal] = {}
    if enrollment is not None:
        try:
            computed = compute_certificate_studies_rows(enrollment)
        except Exception:
            computed = []

        for r in computed or []:
            if not isinstance(r, dict):
                continue
            load_id = r.get("academic_load_id")
            if not load_id:
                continue
            score = _parse_score_decimal(r.get("score"))
            if score is None:
                continue
            try:
                subject_score_by_load_id[int(load_id)] = score
            except Exception:
                continue
    else:
        # Manual/preview: generate plausible subject scores and aggregate using plan weights.
        for load in loads:
            try:
                subject_score_by_load_id[int(load.id)] = Decimal(str(round(random.uniform(3.0, 4.5), 2)))
            except Exception:
                continue

    rows = _certificate_area_rows_from_plan(
        loads=loads,
        subject_score_by_load_id=subject_score_by_load_id,
        academic_year_id=getattr(enrollment, "academic_year_id", None) if enrollment is not None else None,
    )

    place = ""
    if campus and (campus.municipality or campus.department):
        place = " - ".join([p for p in [campus.municipality, campus.department] if p])
    else:
        place = institution.pdf_header_line3 or institution.name

    signer_name = ""
    signer_role = ""
    if institution.rector:
        signer_name = institution.rector.get_full_name()
        signer_role = "Rector(a)"
    elif institution.secretary:
        signer_name = institution.secretary.get_full_name()
        signer_role = "Secretaría"

    return {
        "enrollment": enrollment,
        "grade": grade,
        "institution": institution,
        "campus": campus,
        "student_full_name": student_full_name,
        "document_type": document_type,
        "document_number": document_number,
        "academic_year": academic_year,
        "grade_name": grade.name,
        "academic_level": getattr(getattr(grade, "level", None), "name", "") or "",
        "rows": rows,
        "conduct": "BUENA",
        "final_status": getattr(enrollment, "final_status", "") or "APROBADO",
        "place": place,
        "issue_date": date.today(),
        "signer_name": signer_name,
        "signer_role": signer_role,
    }


IDENTIFICATION_DOCUMENT_TYPES = (
    ("RC", "Registro Civil de Nacimiento"),
    ("TI", "Tarjeta de Identidad"),
    ("CC", "Cédula de Ciudadanía"),
    ("CE", "Cédula de Extranjería"),
    ("PA", "Pasaporte"),
    ("PEP", "PEP"),
)


class CertificateDocumentTypesView(APIView):
    """List identification document types for archived certificate issuance."""

    permission_classes = [IsAdministrativeStaff]

    def get(self, request, format=None):
        return Response(
            {
                "options": [{"value": label, "label": label} for _, label in IDENTIFICATION_DOCUMENT_TYPES],
                "allow_other": True,
            }
        )


class CertificateStudiesIssueView(APIView):
    permission_classes = [IsAdministrativeStaff]

    def post(self, request, format=None):
        """Issue a 'Certificado de estudios' PDF.

        Payload options:
        - Registered student: { enrollment_id, academic_year_id? }
        - Archivo student: { student_full_name, document_type, document_number, grade_id, academic_year? , campus_id? }
        """

        want_async = (request.query_params.get("async") or "").strip().lower() in {"1", "true", "yes"}

        institution = Institution.objects.first() or Institution()

        try:
            built = _certificate_studies_build_context(request, request.data, institution)
        except serializers.ValidationError as e:
            return Response({"detail": str(e.detail) if hasattr(e, 'detail') else str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Enrollment.DoesNotExist:
            return Response({"detail": "Enrollment not found"}, status=status.HTTP_404_NOT_FOUND)
        except Grade.DoesNotExist:
            return Response({"detail": "Grade not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            payload = {"error": "Error preparing certificate", "detail": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        enrollment = built["enrollment"]
        grade = built["grade"]
        campus = built["campus"]
        student_full_name = built["student_full_name"]
        document_type = built["document_type"]
        document_number = built["document_number"]
        academic_year = built["academic_year"]
        rows = built["rows"]
        place = built["place"]
        signer_name = built["signer_name"]
        signer_role = built["signer_role"]

        try:
            amount_cop = int(getattr(institution, "certificate_studies_price_cop", 10000) or 10000)
        except Exception:
            amount_cop = 10000
        if amount_cop < 0:
            amount_cop = 0

        try:
            issue = CertificateIssue.objects.create(
                certificate_type=CertificateIssue.TYPE_STUDIES,
                status=CertificateIssue.STATUS_PENDING,
                enrollment=enrollment,
                issued_by=request.user if getattr(request, "user", None) and request.user.is_authenticated else None,
                amount_cop=amount_cop,
                payload={
                    "student_full_name": student_full_name,
                    "document_type": document_type,
                    "document_number": document_number,
                    "academic_year": academic_year,
                    "grade_id": grade.id,
                    "grade_name": grade.name,
                    "academic_level": getattr(getattr(grade, "level", None), "name", "") or "",
                    "rows": rows,
                    "conduct": "BUENA",
                    "final_status": getattr(enrollment, "final_status", "") or "APROBADO",
                    "place": place,
                    "issue_date": date.today().isoformat(),
                    "signer_name": signer_name,
                    "signer_role": signer_role,
                },
            )

            # Prefer a deploy-safe verification URL under /api/public/verify/<token>/.
            # This remains user-friendly because the API endpoint can render HTML.
            verify_url = ""
            verify_token = ""
            try:
                from verification.services import get_or_create_for_certificate_issue  # noqa: PLC0415

                public_payload = {
                    "title": "Certificado de estudios",
                    "student_full_name": student_full_name,
                    "document_number": document_number,
                    "academic_year": academic_year,
                    "grade_name": grade.name,
                    "rows": rows,
                    "final_status": getattr(enrollment, "final_status", "") or "APROBADO",
                }

                vdoc = get_or_create_for_certificate_issue(
                    issue_uuid=str(issue.uuid),
                    public_payload=public_payload,
                    seal_hash=getattr(issue, "seal_hash", "") or "",
                )
                verify_token = vdoc.token
                verify_path = reverse("public-verify", kwargs={"token": vdoc.token})
                verify_url = _sanitize_url_path(_public_build_absolute_uri(request, verify_path))
            except Exception:
                # Fallback to legacy certificate verification endpoint.
                verify_path = reverse("public-certificate-verify", kwargs={"uuid": str(issue.uuid)})
                verify_url = _sanitize_url_path(_public_build_absolute_uri(request, verify_path))
        except Exception as e:
            # Common deploy-time failure: database schema is behind (migrations not applied).
            try:
                from django.db.utils import OperationalError, ProgrammingError  # noqa: PLC0415

                if isinstance(e, (OperationalError, ProgrammingError)):
                    return Response(
                        {
                            "error": "Database schema out of date",
                            "detail": "Parece que faltan migraciones en el servidor. Ejecuta `python manage.py migrate` (apps: students, reports).",
                        },
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )
            except Exception:
                pass

            payload = {"error": "Error preparing certificate", "detail": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if want_async:
            from datetime import timedelta  # noqa: PLC0415
            from django.utils import timezone  # noqa: PLC0415
            from reports.models import ReportJob  # noqa: PLC0415
            from reports.serializers import ReportJobSerializer  # noqa: PLC0415
            from reports.tasks import generate_report_job_pdf  # noqa: PLC0415

            ttl_hours = int(getattr(settings, "REPORT_JOBS_TTL_HOURS", 24))
            expires_at = timezone.now() + timedelta(hours=ttl_hours)

            try:
                job = ReportJob.objects.create(
                    created_by=request.user,
                    report_type=ReportJob.ReportType.CERTIFICATE_STUDIES,
                    params={"certificate_uuid": str(issue.uuid), "verify_url": verify_url, "verify_token": verify_token},
                    expires_at=expires_at,
                )
            except Exception as e:
                # Likely missing reports migrations / table doesn't exist.
                try:
                    from django.db.utils import OperationalError, ProgrammingError  # noqa: PLC0415

                    if isinstance(e, (OperationalError, ProgrammingError)):
                        try:
                            issue.delete()
                        except Exception:
                            pass
                        return Response(
                            {
                                "error": "Report jobs unavailable",
                                "detail": "No se pudo crear el ReportJob. Ejecuta migraciones del app `reports` en el servidor: `python manage.py migrate reports`.",
                            },
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        )
                except Exception:
                    pass
                raise

            try:
                generate_report_job_pdf.delay(job.id)
            except Exception as e:
                # Broker/worker misconfiguration should not surface as a generic 500.
                job.mark_failed(error_code="ENQUEUE_FAILED", error_message=str(e))
                return Response(
                    {
                        "error": "Could not enqueue PDF job",
                        "detail": "No se pudo encolar la generación del PDF. Revisa que Celery/Redis estén activos y que `CELERY_BROKER_URL` esté configurado.",
                        "job_id": job.id,
                    },
                    status=status.HTTP_503_SERVICE_UNAVAILABLE,
                )

            out = ReportJobSerializer(job, context={"request": request}).data
            return Response(out, status=status.HTTP_202_ACCEPTED)

        qr_image_src = ""
        qr_tmp_path = ""
        try:
            qr_tmp_path = _qr_temp_png_path(verify_url)
            qr_image_src = qr_tmp_path
        except Exception:
            qr_image_src = ""

        ctx = {
            "institution": institution,
            "campus": campus,
            "student_full_name": student_full_name,
            "document_type": document_type,
            "document_number": document_number,
            "academic_year": academic_year,
            "grade_name": grade.name,
            "academic_level": getattr(getattr(grade, "level", None), "name", "") or "",
            "rows": rows,
            "conduct": "BUENA",
            "final_status": getattr(enrollment, "final_status", "") or "APROBADO",
            "place": place,
            "issue_date": date.today(),
            "signer_name": signer_name,
            "signer_role": signer_role,
            "verify_url": verify_url,
            "qr_image_src": qr_image_src,
            "seal_hash": issue.seal_hash,
        }

        try:
            html_string = render_to_string("students/reports/certificate_studies_pdf.html", ctx)

            from reports.weasyprint_utils import WeasyPrintUnavailableError, render_pdf_bytes_from_html  # noqa: PLC0415

            pdf_bytes = render_pdf_bytes_from_html(html=html_string, base_url=str(settings.BASE_DIR))

            if not pdf_bytes:
                try:
                    issue.delete()
                except Exception:
                    pass
                return Response(
                    {
                        "error": "Error generating PDF",
                        "detail": "Empty PDF output",
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            # Store in private storage (outside MEDIA).
            from pathlib import Path  # noqa: PLC0415

            def _safe_join_private(root: Path, relpath: str) -> Path:
                rel = Path(relpath)
                if rel.is_absolute():
                    raise ValueError("Absolute paths are not allowed")
                final = (root / rel).resolve()
                root_resolved = root.resolve()
                if root_resolved not in final.parents and final != root_resolved:
                    raise ValueError("Invalid path")
                return final

            relpath = f"{settings.PRIVATE_REPORTS_DIR}/certificates/studies/{issue.uuid}.pdf".strip("/")
            out_path = _safe_join_private(Path(settings.PRIVATE_STORAGE_ROOT), relpath)
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_bytes(pdf_bytes)

            issue.pdf_private_relpath = relpath
            issue.pdf_private_filename = f"certificado_estudios_{issue.uuid}.pdf"
            issue.status = CertificateIssue.STATUS_ISSUED
            issue.save(update_fields=["pdf_private_relpath", "pdf_private_filename", "status"])

            try:
                log_event(
                    request,
                    event_type="CERTIFICATE_ISSUED",
                    object_type="CertificateIssue",
                    object_id=str(issue.uuid),
                    status_code=200,
                    metadata={
                        "certificate_type": issue.certificate_type,
                        "amount_cop": issue.amount_cop,
                        "student_full_name": student_full_name,
                        "document_number": document_number,
                        "academic_year": academic_year,
                        "grade_name": grade.name,
                    },
                )
            except Exception:
                # Audit logging must never break issuance.
                pass

            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = 'inline; filename="certificado_estudios.pdf"'
            response["Deprecation"] = "true"
            sunset = getattr(settings, "REPORTS_SYNC_SUNSET_DATE", "2026-06-30")
            response["Sunset"] = str(sunset)
            response["Link"] = '</api/reports/jobs/>; rel="alternate"'
            return response
        except WeasyPrintUnavailableError as e:
            try:
                issue.delete()
            except Exception:
                pass
            payload = {"error": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as e:
            try:
                issue.delete()
            except Exception:
                pass
            payload = {"error": "Error generating PDF", "detail": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if qr_tmp_path:
                try:
                    os.remove(qr_tmp_path)
                except Exception:
                    pass


class CertificateIssuesListView(APIView):
    """List issued certificates for auditing/income reconciliation."""

    permission_classes = [IsAdministrativeStaff]

    def get(self, request, format=None):
        qs = CertificateIssue.objects.all().order_by("-issued_at")

        enrollment_id_raw = (request.query_params.get("enrollment_id") or "").strip()
        if enrollment_id_raw:
            try:
                qs = qs.filter(enrollment_id=int(enrollment_id_raw))
            except Exception:
                pass

        student_id_raw = (request.query_params.get("student_id") or "").strip()
        if student_id_raw:
            try:
                qs = qs.filter(enrollment__student_id=int(student_id_raw))
            except Exception:
                pass

        q = (request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(payload__student_full_name__icontains=q)
                | Q(payload__document_number__icontains=q)
            )

        issued_by_raw = (request.query_params.get("issued_by") or "").strip()
        if issued_by_raw:
            try:
                qs = qs.filter(issued_by_id=int(issued_by_raw))
            except Exception:
                pass

        certificate_type = (request.query_params.get("certificate_type") or "").strip()
        if certificate_type:
            qs = qs.filter(certificate_type=certificate_type)

        status_param = (request.query_params.get("status") or "").strip()
        if status_param:
            qs = qs.filter(status=status_param)

        start_date_raw = (request.query_params.get("start_date") or "").strip()
        end_date_raw = (request.query_params.get("end_date") or "").strip()

        def _parse_date(s: str):
            try:
                return datetime.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return None

        start_date = _parse_date(start_date_raw) if start_date_raw else None
        end_date = _parse_date(end_date_raw) if end_date_raw else None
        if start_date:
            qs = qs.filter(issued_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(issued_at__date__lte=end_date)

        try:
            limit = int(request.query_params.get("limit") or 100)
        except Exception:
            limit = 100
        limit = max(1, min(limit, 500))

        items = []
        for issue in qs.select_related("issued_by", "enrollment").only(
            "uuid",
            "certificate_type",
            "status",
            "issued_at",
            "amount_cop",
            "payload",
            "enrollment_id",
            "enrollment__student_id",
            "pdf_private_relpath",
            "issued_by__id",
            "issued_by__first_name",
            "issued_by__last_name",
        )[:limit]:
            payload = issue.payload or {}
            issued_by = getattr(issue, "issued_by", None)
            items.append(
                {
                    "uuid": str(issue.uuid),
                    "certificate_type": issue.certificate_type,
                    "status": issue.status,
                    "issued_at": issue.issued_at,
                    "amount_cop": issue.amount_cop,
                    "enrollment_id": issue.enrollment_id,
                    "student_id": getattr(getattr(issue, "enrollment", None), "student_id", None),
                    "student_full_name": payload.get("student_full_name") or "",
                    "document_number": payload.get("document_number") or "",
                    "academic_year": payload.get("academic_year") or "",
                    "grade_name": payload.get("grade_name") or "",
                    "issued_by": (
                        {
                            "id": issued_by.id,
                            "name": issued_by.get_full_name(),
                        }
                        if issued_by
                        else None
                    ),
                    "has_pdf": bool(getattr(issue, "pdf_private_relpath", None) or getattr(issue, "pdf_file", None)),
                }
            )

        return Response({"results": items, "count": qs.count(), "limit": limit})


class CertificateIssueDetailView(APIView):
    """Edit (PATCH) and delete/revoke (DELETE) a certificate issue.

    Rules:
    - Only PENDING certificates can be edited.
    - DELETE on ISSUED will revoke (soft-delete) instead of hard delete.
    - DELETE on PENDING/REVOKED will hard delete.
    """

    permission_classes = [IsAdministrativeStaff]

    def _serialize(self, issue: CertificateIssue):
        payload = issue.payload or {}
        issued_by = getattr(issue, "issued_by", None)
        return {
            "uuid": str(issue.uuid),
            "certificate_type": issue.certificate_type,
            "status": issue.status,
            "issued_at": issue.issued_at,
            "amount_cop": issue.amount_cop,
            "student_full_name": payload.get("student_full_name") or "",
            "document_number": payload.get("document_number") or "",
            "academic_year": payload.get("academic_year") or "",
            "grade_name": payload.get("grade_name") or "",
            "issued_by": (
                {
                    "id": issued_by.id,
                    "name": issued_by.get_full_name(),
                }
                if issued_by
                else None
            ),
            "has_pdf": bool(getattr(issue, "pdf_private_relpath", None) or getattr(issue, "pdf_file", None)),
        }

    def patch(self, request, uuid, format=None):
        issue = CertificateIssue.objects.filter(uuid=uuid).select_related("issued_by").first()
        if not issue:
            return Response({"detail": "Certificate not found"}, status=status.HTTP_404_NOT_FOUND)

        if issue.status != CertificateIssue.STATUS_PENDING:
            return Response(
                {"detail": "Solo certificados en estado PENDIENTE pueden editarse."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = request.data or {}

        payload = dict(issue.payload or {})
        for key in ["student_full_name", "document_type", "document_number", "academic_year", "grade_name"]:
            if key in data:
                val = data.get(key)
                if val is None:
                    continue
                payload[key] = val

        update_fields = []
        if payload != (issue.payload or {}):
            issue.payload = payload
            update_fields.append("payload")

        if "amount_cop" in data:
            try:
                amount_cop = int(data.get("amount_cop") or 0)
            except Exception:
                return Response({"detail": "amount_cop inválido"}, status=status.HTTP_400_BAD_REQUEST)
            if amount_cop < 0:
                amount_cop = 0
            if issue.amount_cop != amount_cop:
                issue.amount_cop = amount_cop
                update_fields.append("amount_cop")

        if not update_fields:
            return Response(self._serialize(issue))

        # Keep seal consistent with payload snapshot.
        try:
            issue.seal_hash = issue._compute_seal_hash()
            update_fields.append("seal_hash")
        except Exception:
            # Never block editing for seal issues.
            pass

        issue.save(update_fields=list(dict.fromkeys(update_fields)))
        return Response(self._serialize(issue))

    def delete(self, request, uuid, format=None):
        issue = CertificateIssue.objects.filter(uuid=uuid).select_related("issued_by").first()
        if not issue:
            return Response({"detail": "Certificate not found"}, status=status.HTTP_404_NOT_FOUND)

        # If already revoked or pending: allow hard delete.
        if issue.status in {CertificateIssue.STATUS_PENDING, CertificateIssue.STATUS_REVOKED}:
            try:
                issue.delete()
            except Exception as e:
                payload = {"error": "Error deleting certificate", "detail": str(e)}
                if getattr(settings, 'DEBUG', False):
                    payload["traceback"] = traceback.format_exc()
                return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            return Response(status=status.HTTP_204_NO_CONTENT)

        # For ISSUED: revoke instead of hard delete.
        reason = ""
        try:
            reason = str((request.data or {}).get("reason") or "").strip()
        except Exception:
            reason = ""

        try:
            from django.utils import timezone  # noqa: PLC0415

            issue.status = CertificateIssue.STATUS_REVOKED
            issue.revoked_at = timezone.now()
            issue.revoke_reason = reason
            issue.save(update_fields=["status", "revoked_at", "revoke_reason"])
        except Exception as e:
            payload = {"error": "Error revoking certificate", "detail": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        out = self._serialize(issue)
        out["revoked"] = True
        return Response(out)


class CertificateIssueDownloadPDFView(APIView):
    """Download the stored PDF for an issued certificate."""

    permission_classes = [IsAdministrativeStaff]

    def get(self, request, uuid, format=None):
        issue = CertificateIssue.objects.filter(uuid=uuid).first()
        if not issue:
            return Response({"detail": "Certificate not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            from pathlib import Path  # noqa: PLC0415

            def _safe_join_private(root: Path, relpath: str) -> Path:
                rel = Path(relpath)
                if rel.is_absolute():
                    raise ValueError("Absolute paths are not allowed")
                final = (root / rel).resolve()
                root_resolved = root.resolve()
                if root_resolved not in final.parents and final != root_resolved:
                    raise ValueError("Invalid path")
                return final

            if getattr(issue, "pdf_private_relpath", None):
                abs_path = _safe_join_private(Path(settings.PRIVATE_STORAGE_ROOT), issue.pdf_private_relpath)
                if not abs_path.exists():
                    return Response({"detail": "No stored PDF for this certificate"}, status=status.HTTP_404_NOT_FOUND)

                filename = issue.pdf_private_filename or f"certificado_{issue.uuid}.pdf"
                response = FileResponse(open(abs_path, "rb"), content_type="application/pdf")
                response["Content-Disposition"] = f'inline; filename="{filename}"'
                return response

            if not issue.pdf_file:
                return Response({"detail": "No stored PDF for this certificate"}, status=status.HTTP_404_NOT_FOUND)

            issue.pdf_file.open("rb")
            response = FileResponse(issue.pdf_file, content_type="application/pdf")
            response["Content-Disposition"] = f'inline; filename="certificado_{issue.uuid}.pdf"'
            return response
        except Exception as e:
            payload = {"error": "Error reading stored PDF", "detail": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CertificateRevenueSummaryView(APIView):
    """Revenue summary based on issued certificates (amount_cop snapshot)."""

    permission_classes = [IsAdministrativeStaff]

    def get(self, request, format=None):
        qs = CertificateIssue.objects.filter(status=CertificateIssue.STATUS_ISSUED)

        certificate_type = (request.query_params.get("certificate_type") or "").strip()
        if certificate_type:
            qs = qs.filter(certificate_type=certificate_type)

        start_date_raw = (request.query_params.get("start_date") or "").strip()
        end_date_raw = (request.query_params.get("end_date") or "").strip()

        def _parse_date(s: str):
            try:
                return datetime.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return None

        start_date = _parse_date(start_date_raw) if start_date_raw else None
        end_date = _parse_date(end_date_raw) if end_date_raw else None
        if start_date:
            qs = qs.filter(issued_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(issued_at__date__lte=end_date)

        agg = qs.aggregate(total_amount_cop=Sum("amount_cop"))
        total_amount = int(agg.get("total_amount_cop") or 0)
        total_count = qs.count()

        return Response(
            {
                "total_count": total_count,
                "total_amount_cop": total_amount,
            }
        )


class CertificateStudiesPreviewView(APIView):
    """Render the certificate as HTML for live styling (no PDF, no DB write)."""

    permission_classes = [IsAdministrativeStaff]

    def post(self, request, format=None):
        institution = Institution.objects.first() or Institution()

        try:
            built = _certificate_studies_build_context(request, request.data, institution)
        except serializers.ValidationError as e:
            return Response({"detail": str(e.detail) if hasattr(e, 'detail') else str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Enrollment.DoesNotExist:
            return Response({"detail": "Enrollment not found"}, status=status.HTTP_404_NOT_FOUND)
        except Grade.DoesNotExist:
            return Response({"detail": "Grade not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            payload = {"error": "Error preparing preview", "detail": str(e)}
            if getattr(settings, 'DEBUG', False):
                payload["traceback"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Preview-only token (not persisted): used only to render QR visuals.
        preview_token = secrets.token_urlsafe(16)
        verify_path = reverse("public-verify", kwargs={"token": preview_token})
        verify_url = _sanitize_url_path(_public_build_absolute_uri(request, verify_path))

        # For HTML preview, use a data URI so the browser can render it.
        # (Temp file paths are only resolvable on the server, not by the browser.)
        try:
            qr_image_src = _qr_data_uri(verify_url)
        except Exception:
            qr_image_src = ""

        # Compute a deterministic seal hash for preview (not stored).
        seal_payload = {
            "student_full_name": built["student_full_name"],
            "document_type": built["document_type"],
            "document_number": built["document_number"],
            "academic_year": built["academic_year"],
            "grade_id": built["grade"].id,
            "grade_name": built["grade"].name,
            "rows": built["rows"],
        }
        try:
            seal_hash = hashlib.sha256(json.dumps(seal_payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()
        except Exception:
            seal_hash = ""

        ctx = {
            "institution": institution,
            "campus": built["campus"],
            "student_full_name": built["student_full_name"],
            "document_type": built["document_type"],
            "document_number": built["document_number"],
            "academic_year": built["academic_year"],
            "grade_name": built["grade"].name,
            "academic_level": built["academic_level"],
            "rows": built["rows"],
            "conduct": built["conduct"],
            "final_status": built["final_status"],
            "place": built["place"],
            "issue_date": built["issue_date"],
            "signer_name": built["signer_name"],
            "signer_role": built["signer_role"],
            "verify_url": verify_url,
            "verify_token": preview_token,
            "qr_image_src": qr_image_src,
            "seal_hash": seal_hash,
        }

        html_string = render_to_string("students/reports/certificate_studies_pdf.html", ctx)
        return HttpResponse(html_string, content_type="text/html; charset=utf-8")



class PublicCertificateVerifyView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request, uuid, format=None):
        # If a browser is requesting HTML (typical when scanning a QR), render a
        # user-friendly public page directly. This avoids a fragile dependency on
        # reverse-proxy routing for `/public/`.
        accept = (request.META.get("HTTP_ACCEPT") or "").lower()
        if "text/html" in accept:
            django_request = getattr(request, "_request", request)
            return PublicCertificateVerifyUIView().get(django_request, uuid)

        try:
            issue = CertificateIssue.objects.select_related(
                "enrollment",
                "enrollment__student",
                "enrollment__student__user",
                "enrollment__grade",
                "enrollment__academic_year",
            ).get(uuid=uuid)
        except CertificateIssue.DoesNotExist:
            _audit_certificate_verification_failed(
                request=request,
                attempted_id=str(uuid),
                via="api",
                reason="not_found",
            )
            return Response({"valid": False, "detail": "Certificate not found"}, status=status.HTTP_404_NOT_FOUND)

        _notify_admins_certificate_verification(request=request, issue=issue, via="api")
        _audit_certificate_verification(request=request, issue=issue, via="api")

        payload = issue.payload or {}

        student_name = payload.get("student_full_name")
        document_number = payload.get("document_number")
        academic_year = payload.get("academic_year")
        grade_name = payload.get("grade_name")

        if issue.enrollment:
            try:
                student_name = student_name or issue.enrollment.student.user.get_full_name()
                document_number = document_number or issue.enrollment.student.document_number
                academic_year = academic_year or issue.enrollment.academic_year.year
                grade_name = grade_name or (issue.enrollment.grade.name if issue.enrollment.grade else "")
            except Exception:
                pass

        return Response(
            {
                "valid": issue.status == CertificateIssue.STATUS_ISSUED,
                "status": issue.status,
                "type": issue.certificate_type,
                "uuid": str(issue.uuid),
                "issued_at": issue.issued_at,
                "seal_hash": issue.seal_hash,
                "revoked": issue.status == CertificateIssue.STATUS_REVOKED,
                "revoke_reason": issue.revoke_reason,
                "student_full_name": student_name or "",
                "document_number": document_number or "",
                "academic_year": academic_year or "",
                "grade": grade_name or "",
                "grade_id": payload.get("grade_id"),
            }
        )


def _normalize_uuid_like(value: str) -> py_uuid.UUID:
    """Parse a UUID from common/legacy formats.

    Accepts canonical UUIDs (with hyphens) and also hex-ish strings that may be
    missing separators. We normalize by stripping any non-hex characters and
    parsing the resulting 32-hex UUID.
    """

    raw = str(value or "").strip().lower()
    hex32 = re.sub(r"[^0-9a-f]", "", raw)
    if len(hex32) != 32:
        raise ValueError("Invalid UUID")
    return py_uuid.UUID(hex=hex32)


def _sanitize_url_path(url: str) -> str:
    """Remove accidental whitespace after slashes in a URL path.

    Fixes legacy QR URLs like `/public/  certificates/<id>/`.
    """

    raw = str(url or "").strip()
    if not raw:
        return raw

    try:
        parts = urlsplit(raw)
        clean_path = re.sub(r"/\s+", "/", parts.path)
        return urlunsplit((parts.scheme, parts.netloc, clean_path, parts.query, parts.fragment))
    except Exception:
        return re.sub(r"/\s+", "/", raw)


def _public_build_absolute_uri(request, path: str) -> str:
    """Build an absolute URL for public QR verification.

    In production, the API may be served behind a reverse proxy or on a
    different subdomain than the public site. When KAMPUS_PUBLIC_SITE_URL is
    configured, prefer it as the canonical base.
    """

    try:
        from urllib.parse import urljoin  # noqa: PLC0415

        base = (getattr(settings, "PUBLIC_SITE_URL", "") or "").strip().rstrip("/")
        if base:
            return urljoin(base + "/", str(path or "").lstrip("/"))
    except Exception:
        pass

    # Fallback: depends on correct proxy headers.
    try:
        return request.build_absolute_uri(path)
    except Exception:
        # DRF Request wrapper compatibility
        django_request = getattr(request, "_request", None)
        if django_request is not None:
            return django_request.build_absolute_uri(path)
        return str(path or "")


def _promoted_from_final_status(final_status: str | None) -> bool | None:
    """Infer promotion from Enrollment.final_status-like strings.

    Returns:
    - True / False when it can be inferred confidently
    - None when unknown/ambiguous
    """

    s = (final_status or "").strip().upper()
    if not s:
        return None

    negative_markers = ["NO APROB", "REPROB", "RETEN", "FAILED"]
    if any(m in s for m in negative_markers):
        return False

    positive_markers = ["APROB", "PROMOV"]
    if any(m in s for m in positive_markers):
        return True

    return None


def _get_request_ip_for_public(request) -> str:
    xff = (request.META.get("HTTP_X_FORWARDED_FOR") or "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return (request.META.get("REMOTE_ADDR") or "").strip()


def _notify_admins_certificate_verification(*, request, issue: "CertificateIssue", via: str) -> None:
    """Notify admin-like users that a certificate verification was accessed.

    Best-effort: must never break the public endpoint.
    """

    status_labels = {
        CertificateIssue.STATUS_PENDING: "Pendiente",
        CertificateIssue.STATUS_ISSUED: "Emitido",
        CertificateIssue.STATUS_REVOKED: "Revocado",
    }
    type_labels = {
        CertificateIssue.TYPE_STUDIES: "Certificado de estudios",
    }

    type_label = type_labels.get(issue.certificate_type, issue.certificate_type)
    status_label = status_labels.get(issue.status, issue.status)
    ip = _get_request_ip_for_public(request)

    user_agent = (request.META.get("HTTP_USER_AGENT") or "").strip()
    if len(user_agent) > 200:
        user_agent = user_agent[:200] + "…"

    try:
        verify_path = reverse("public-site-certificate-verify-ui", kwargs={"uuid": str(issue.uuid)})
        verify_url = _sanitize_url_path(_public_build_absolute_uri(request, verify_path))
    except Exception:
        verify_url = ""

    body = (
        f"{type_label} | Estado: {status_label}\n"
        f"UUID: {issue.uuid}\n"
        f"Vía: {via}\n"
        f"IP: {ip}\n"
        f"User-Agent: {user_agent}"
    )
    dedupe_key = f"CERT_VERIFY:{issue.uuid}"
    dedupe_within_seconds = 15 * 60

    # Preferred path: use centralized notification service.
    try:
        from notifications.services import admin_like_users_qs, notify_users

        notify_users(
            recipients=admin_like_users_qs(),
            type="CERTIFICATE_VERIFY",
            title="Verificación de certificado consultada",
            body=body,
            url=verify_url,
            dedupe_key=dedupe_key,
            dedupe_within_seconds=dedupe_within_seconds,
        )
        return
    except Exception:
        # Best-effort fallback: create Notification rows directly.
        pass

    try:
        from datetime import timedelta

        from django.utils import timezone
        from notifications.models import Notification

        UserModel = get_user_model()
        recipients = UserModel.objects.filter(
            role__in=[
                getattr(UserModel, "ROLE_SUPERADMIN", "SUPERADMIN"),
                getattr(UserModel, "ROLE_ADMIN", "ADMIN"),
                getattr(UserModel, "ROLE_COORDINATOR", "COORDINATOR"),
            ],
            is_active=True,
        )

        since = timezone.now() - timedelta(seconds=int(dedupe_within_seconds))
        existing_ids = set(
            Notification.objects.filter(
                recipient__in=recipients,
                dedupe_key=dedupe_key,
                created_at__gte=since,
            ).values_list("recipient_id", flat=True)
        )

        to_create = []
        for u in recipients:
            if u.id in existing_ids:
                continue
            to_create.append(
                Notification(
                    recipient=u,
                    type="CERTIFICATE_VERIFY",
                    title="Verificación de certificado consultada",
                    body=body,
                    url=verify_url,
                    dedupe_key=dedupe_key,
                )
            )

        if to_create:
            Notification.objects.bulk_create(to_create)
    except Exception:
        return


def _audit_certificate_verification(*, request, issue: "CertificateIssue", via: str) -> None:
    """Persist an audit record for a public certificate verification access."""

    try:
        from audit.services import log_public_event
    except Exception:
        return

    try:
        log_public_event(
            request,
            event_type="PUBLIC_CERTIFICATE_VERIFY",
            object_type="CertificateIssue",
            object_id=str(issue.uuid),
            status_code=200,
            metadata={
                "certificate_type": getattr(issue, "certificate_type", ""),
                "certificate_status": getattr(issue, "status", ""),
                "via": via,
            },
        )
    except Exception:
        return


def _audit_certificate_verification_failed(*, request, attempted_id: str, via: str, reason: str) -> None:
    """Audit public verification failures (404/invalid id).

    This helps detect scans of non-existent or malformed QR codes.
    """

    try:
        from audit.services import log_public_event
    except Exception:
        return

    try:
        log_public_event(
            request,
            event_type="PUBLIC_CERTIFICATE_VERIFY",
            object_type="CertificateIssue",
            object_id=str(attempted_id or ""),
            status_code=404,
            metadata={
                "via": via,
                "reason": reason,
            },
        )
    except Exception:
        return


class PublicCertificateVerifyLegacyUIView(View):
    """Redirect legacy certificate URLs to the canonical UUID route."""

    def get(self, request, uuid_str):
        try:
            u = _normalize_uuid_like(uuid_str)
        except Exception:
            _audit_certificate_verification_failed(
                request=request,
                attempted_id=str(uuid_str),
                via="ui",
                reason="invalid_id",
            )
            return render(
                request,
                "students/public/certificate_verify.html",
                {
                    "institution": Institution.objects.first() or Institution(),
                    "found": False,
                    "uuid": str(uuid_str),
                },
                status=404,
            )
        return redirect(reverse("public-site-certificate-verify-ui", kwargs={"uuid": str(u)}))


class PublicCertificateVerifyLegacyView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request, uuid_str, format=None):
        try:
            u = _normalize_uuid_like(uuid_str)
        except Exception:
            _audit_certificate_verification_failed(
                request=request,
                attempted_id=str(uuid_str),
                via="api",
                reason="invalid_id",
            )
            return Response({"valid": False, "detail": "Invalid certificate id"}, status=status.HTTP_404_NOT_FOUND)
        return redirect(reverse("public-site-certificate-verify", kwargs={"uuid": str(u)}))


class PublicCertificateVerifyUIView(View):
    def get(self, request, uuid):
        institution = Institution.objects.first() or Institution()

        status_labels = {
            CertificateIssue.STATUS_PENDING: "Pendiente",
            CertificateIssue.STATUS_ISSUED: "Emitido",
            CertificateIssue.STATUS_REVOKED: "Revocado",
        }
        type_labels = {
            CertificateIssue.TYPE_STUDIES: "Certificado de estudios",
        }

        issue = (
            CertificateIssue.objects.select_related(
                "enrollment",
                "enrollment__student",
                "enrollment__student__user",
                "enrollment__grade",
                "enrollment__academic_year",
            )
            .filter(uuid=uuid)
            .first()
        )

        if not issue:
            _audit_certificate_verification_failed(
                request=request,
                attempted_id=str(uuid),
                via="ui",
                reason="not_found",
            )
            return render(
                request,
                "students/public/certificate_verify.html",
                {
                    "institution": institution,
                    "found": False,
                    "uuid": str(uuid),
                },
                status=404,
            )

        _notify_admins_certificate_verification(request=request, issue=issue, via="ui")
        _audit_certificate_verification(request=request, issue=issue, via="ui")

        payload = issue.payload or {}

        student_name = payload.get("student_full_name")
        document_number = payload.get("document_number")
        academic_year = payload.get("academic_year")
        grade_name = payload.get("grade_name")

        if issue.enrollment:
            try:
                student_name = student_name or issue.enrollment.student.user.get_full_name()
                document_number = document_number or issue.enrollment.student.document_number
                academic_year = academic_year or issue.enrollment.academic_year.year
                grade_name = grade_name or (issue.enrollment.grade.name if issue.enrollment.grade else "")
            except Exception:
                pass

        final_status = payload.get("final_status")
        if not final_status and issue.enrollment:
            final_status = getattr(issue.enrollment, "final_status", "")

        rows = payload.get("rows")
        if not rows and issue.enrollment:
            # Recompute current area rows from the study plan (includes hours/weights).
            try:
                loads = _subjects_for_grade(issue.enrollment.grade_id)
                computed = compute_certificate_studies_rows(issue.enrollment)
            except Exception:
                loads = []
                computed = []

            subject_score_by_load_id: dict[int, Decimal] = {}
            for r in computed or []:
                if not isinstance(r, dict):
                    continue
                load_id = r.get("academic_load_id")
                if not load_id:
                    continue
                score = _parse_score_decimal(r.get("score"))
                if score is None:
                    continue
                try:
                    subject_score_by_load_id[int(load_id)] = score
                except Exception:
                    continue

            rows = _certificate_area_rows_from_plan(
                loads=loads,
                subject_score_by_load_id=subject_score_by_load_id,
                academic_year_id=getattr(issue.enrollment, "academic_year_id", None),
            )

        # If payload rows exist but lack hours, enrich from the current study plan.
        if isinstance(rows, list) and issue.enrollment:
            try:
                loads_for_hours = _subjects_for_grade(issue.enrollment.grade_id)
            except Exception:
                loads_for_hours = []
            area_hours: dict[str, int] = {}
            for load in loads_for_hours:
                try:
                    area_name = str(load.subject.area.name or "").strip()
                    if not area_name:
                        continue
                    area_hours[area_name] = area_hours.get(area_name, 0) + int(getattr(load, "hours_per_week", 0) or 0)
                except Exception:
                    continue

            enriched = []
            for r in rows:
                if not isinstance(r, dict):
                    continue
                area_name = str(r.get("area_subject") or "").strip()
                rr = dict(r)
                if rr.get("hours_per_week") in [None, "", 0]:
                    if area_name in area_hours:
                        rr["hours_per_week"] = area_hours[area_name]
                enriched.append(rr)
            rows = enriched

        # Normalize rows for template safety.
        safe_rows = []
        if isinstance(rows, list):
            for r in rows:
                if not isinstance(r, dict):
                    continue
                safe_rows.append(
                    {
                        "area_subject": str(r.get("area_subject") or "").strip(),
                        "hours_per_week": r.get("hours_per_week") or "",
                        "score": str(r.get("score") or "").strip(),
                        "performance": str(r.get("performance") or "").strip(),
                    }
                )

        api_json_url = request.build_absolute_uri(
            reverse("public-certificate-verify", kwargs={"uuid": str(issue.uuid)})
        )

        context = {
            "institution": institution,
            "found": True,
            "valid": issue.status == CertificateIssue.STATUS_ISSUED,
            "status": issue.status,
            "status_label": status_labels.get(issue.status, issue.status),
            "certificate_type": issue.certificate_type,
            "certificate_type_label": type_labels.get(issue.certificate_type, issue.certificate_type),
            "uuid": str(issue.uuid),
            "issued_at": issue.issued_at,
            "seal_hash": issue.seal_hash,
            "revoked": issue.status == CertificateIssue.STATUS_REVOKED,
            "revoke_reason": issue.revoke_reason,
            "student_full_name": student_name or "",
            "document_number": document_number or "",
            "academic_year": academic_year or "",
            "grade": grade_name or "",
            "grade_id": payload.get("grade_id"),
            "final_status": final_status or "",
            "promoted": _promoted_from_final_status(final_status),
            "rows": safe_rows,
            "api_json_url": f"{api_json_url}?format=json",
        }

        return render(request, "students/public/certificate_verify.html", context)



